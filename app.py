# -*- coding: utf-8 -*-
"""基金日报助手 - 桌面版
本地化的基金持仓看板与日报工具：实时估值/历史净值/回撤修复/同类排名/指数对比，
持仓本地存储，支持粘贴导入、交易记账、已清仓标记。
"""
import sys
import json
import urllib.request
import re
import traceback
import numpy as np
import os
import zipfile
from datetime import datetime, timedelta

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QFrame, QScrollArea, QSizePolicy, QMessageBox,
    QTextEdit, QDialog, QStackedWidget, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QButtonGroup, QGraphicsOpacityEffect,
    QComboBox, QLineEdit, QFileDialog, QSplitter, QCheckBox,
    QRadioButton, QFormLayout, QDateEdit, QGridLayout, QInputDialog,
    QListWidget,
)
from PySide6.QtCore import (
    Qt, QThread, Signal, QTimer, QPropertyAnimation, QAbstractAnimation, QDate,
    QPointF,
)
from PySide6.QtGui import QFont, QColor, QCursor, QPainter, QPen, QPainterPath, QIcon
import pyqtgraph as pg

import data_sources as market_data
import domain as domain_logic
import storage as storage_layer
import ui_theme as theme_ui

# 工作目录固定为程序所在目录，保证相对路径在源码/脚本/打包环境下行为一致
if getattr(sys, 'frozen', False):
    _BASE = os.path.dirname(sys.executable)
    # macOS .app 包：数据文件落在 .app 外部同级目录，不进包内（v1.0.0 双端适配；
    # 首版误写包内已被用户实机捆出，本版修正）
    if sys.platform == "darwin" and _BASE.endswith(".app/Contents/MacOS"):
        _BASE = os.path.dirname(_BASE[:-len("/Contents/MacOS")])
else:
    _BASE = os.path.dirname(os.path.abspath(__file__))
os.chdir(_BASE)

# 跨平台 UI 字体：Windows 用微软雅黑，macOS 用苹方（v1.0.0 双端适配）
FONT = "Microsoft YaHei" if sys.platform == "win32" else "PingFang SC"

FUNDS = []   # 看板基金全部来自「自定义基金.json」；初始为空看板，用户经首页『快速添加』自行建立名单
EXTRA_FILE = "自定义基金.json"

def _atomic_write_json(path, obj, **kw):
    """v2.3 兼容出口：实际原子写盘实现位于 storage.py。"""
    return storage_layer.atomic_write_json(path, obj, **kw)

def _load_json_with_bak(path, default):
    """v2.3 兼容出口：实际损坏回退实现位于 storage.py。"""
    return storage_layer.load_json_with_bak(path, default)

def _load_extra():
    arr = _load_json_with_bak(EXTRA_FILE, [])
    if not isinstance(arr, list): arr = []
    out = []; seen = set()
    for it in arr:
        if not isinstance(it, dict): continue
        c = str(it.get("code", "")).strip(); n = str(it.get("name", "")).strip()
        if re.search(r"^\d{6}$", c) and n and c not in seen:
            out.append({"code": c, "name": n}); seen.add(c)
    return out

def _save_extra(arr):
    try:
        _atomic_write_json(EXTRA_FILE, arr, indent=2)
    except Exception:
        pass

for _it in _load_extra():
    if _it["code"] not in {c for c, _ in FUNDS}:
        FUNDS.append((_it["code"], _it["name"]))
NAME_MAP = {c: n for c, n in FUNDS}

HOLD_FILE = "我的持仓.json"
SHOW_FILE = "基金显示.json"   # 显示层状态(已清仓标记等)；缺席=默认，不碰持仓账本
ACCOUNTS_FILE = "账户.json"    # v0.6 多账户
SETTINGS_FILE = "settings.json"  # v0.7 用户设置（默认备份目录等）
DEFAULT_ACCOUNT = "默认"
APP_VERSION = "2.3.0"
GITHUB_REPO = "GoldenMoon-cell/fund-daily-report"
RED, GREEN, GRAY = "#e53935", "#16a34a", "#888888"
TEAL = "#0891b2"
HL = QColor("#fff7d6")

# ---------- v2.3 主题模块兼容层 ----------
# 令牌与 QSS/QPainter 资产的唯一实现位于 ui_theme.py；app.py 保留旧调用名。
THEMES = theme_ui.THEMES
_THEME = "classic"
def T():
    """当前主题令牌表（v1.7）。"""
    return THEMES.get(_THEME) or THEMES["classic"]
def set_theme(name):
    """v1.7：切换活动主题并同步全局信号色常量；应在构建 UI 前或 _restyle_all 前调用。"""
    global _THEME, RED, GREEN, GRAY
    if name in THEMES: _THEME = name
    t = T(); RED, GREEN, GRAY = t["up"], t["down"], t["flat"]
def card_qss_normal():
    return theme_ui.card_qss_normal(T())
def card_qss_cleared():
    return theme_ui.card_qss_cleared(T())
def panel_qss():
    return theme_ui.panel_qss(T())
def board_qss():
    return theme_ui.board_qss(T())
def ghost_btn_qss(t=None, pad="7px 12px"):
    """v2.0.0：幽灵按钮（FD-HIG 按钮家族 Ghost：透明底+发丝边，导航/工具用）。"""
    return theme_ui.ghost_btn_qss(t or T(), pad)
def primary_btn_qss(t=None):
    """v2.0.0：主操作按钮（FD-HIG Primary：每屏最多一个）。"""
    return theme_ui.primary_btn_qss(t or T())
def soft_btn_qss(t=None):
    """v2.0.0：软按钮（FD-HIG Soft：卡片内轻操作）。"""
    return theme_ui.soft_btn_qss(t or T())
def danger_btn_qss(t=None, pad="8px 14px"):
    """v2.0.0：危险按钮（FD-HIG Danger：破坏性操作，红描边红字）。"""
    return theme_ui.danger_btn_qss(t or T(), pad)
def panel_label_qss(kind="tip"):
    """v2.0.0：提示条标签（tip 中性 / warn 警示 / ok 成功），全主题适配。"""
    return theme_ui.panel_label_qss(T(), kind)
def table_qss(extra=""):
    """v2.0.0：表格主题样式（背景/斑马纹/表头/选中态）。"""
    return theme_ui.table_qss(T(), extra)
def _arrow_png(kind, color):
    """v2.0.0：下拉/日期控件箭头用 QPainter 直画存 PNG，QSS 按绝对路径引用。
       （CSS 边框三角在 Qt 里实测渲染成横条，弃用。）"""
    return theme_ui.arrow_png(kind, color)

def combo_qss(t=None):
    """v2.0.0：下拉框 Win11 风格（圆角/扁平/手绘箭头，去原生斜边框）。"""
    return theme_ui.combo_qss(t or T())
def input_qss(t=None):
    return theme_ui.input_qss(t or T())
def hl_bg():
    """主题感知的“本金误填”行高亮色（暗色主题用暗琥珀，避免刺眼亮黄）。"""
    return theme_ui.highlight_background(T())
def imp_bg():
    """主题感知的“导入/对账待核对”单元格高亮色。"""
    return theme_ui.import_background(T())
def blend_color(bg_hex, fg_hex, alpha):
    """v2.0.0：把 fg 按 alpha 混入 bg（收益日历格子配色用，全主题通用）。"""
    return theme_ui.blend_color(bg_hex, fg_hex, alpha)
def global_qss(t):
    """v2.0.0：弹窗/表格/输入类控件的全局主题样式（四主题均应用）。"""
    return theme_ui.global_qss(t)

def icon_pixmap(kind, color, size=24, stroke=1.6):
    """v2.0.0：原创手绘图标（QPainter 直画，无 emoji 无图标库）。
       FD-HIG 图标规范：24 网格 / 1.5~1.6 线宽 / 圆帽圆角连接 / 单色。"""
    return theme_ui.icon_pixmap(kind, color, size, stroke)
IMP = QColor("#ffe0b2")
REPAIR_THRESHOLD = 0.5
CMP_INDEX = [("1.000300", "沪深300"), ("1.000905", "中证500"), ("1.000016", "上证50"), ("0.399006", "创业板指")]

# 基金类型样例库：仅用于详情页展示基金类型/跟踪指数说明，可自行增删
TRACK = {
    "000001": ("主动混合型", "国内老牌主动混合基金样例（方向以季报为准）"),
    "110011": ("主动混合型(QDII)", "优质企业主题主动选股样例（方向以季报为准）"),
    "161725": ("股票指数型", "跟踪 中证白酒指数（消费行业）"),
    "510300": ("股票指数型", "跟踪 沪深300指数（A股大盘蓝筹）"),
    "000217": ("商品型(黄金)", "跟踪 上海金Au99.99现货合约；跟国内金价走，国内金价≈国际金价×汇率÷31.1035±溢价"),
    "019548": ("股票指数型(QDII)", "跟踪 纳斯达克100指数（美股科技龙头）"),
    "013309": ("股票指数型(QDII)", "跟踪 恒生科技指数（港股科技龙头）"),
    "161120": ("债券指数型", "跟踪 中债-新综合指数（利率债+信用债综合）"),
    "022098": ("股票指数型", "跟踪 中证红利低波动100指数（高股息+低波动双因子）"),
}

_SSL = market_data.DEFAULT_SSL_CONTEXT


def load_accounts():
    arr = _load_json_with_bak(ACCOUNTS_FILE, None)
    if isinstance(arr, list) and arr:
        return [a for a in (x.get("name", "").strip() if isinstance(x, dict) else str(x).strip() for x in arr) if a]
    return [DEFAULT_ACCOUNT]

def save_accounts(names):
    _atomic_write_json(ACCOUNTS_FILE, [{"name": n} for n in names], indent=2)

def load_settings():
    d = _load_json_with_bak(SETTINGS_FILE, {})
    return d if isinstance(d, dict) else {}

def save_settings(d):
    try:
        _atomic_write_json(SETTINGS_FILE, d, indent=2)
    except Exception:
        pass

def _migrate_holdings_if_needed(d):
    """v0.5→v0.6：旧扁平格式 {code: rec} → 新嵌套格式 {"默认": {code: rec}}。"""
    if not d:
        return d
    for k, v in d.items():
        if isinstance(v, dict) and re.search(r'^\d{6}$', k.strip()):
            migrated = {DEFAULT_ACCOUNT: d}
            save_holdings_nested(migrated)
            return migrated
    return d

def load_holdings():
    """读取全部持仓并合并为 {code: rec}（跨账户汇总），兼容 v0.5 旧格式。"""
    d = _load_json_with_bak(HOLD_FILE, {})
    if not isinstance(d, dict): return {}
    d = _migrate_holdings_if_needed(d)
    return merge_holdings(d)

def load_holdings_nested():
    d = _load_json_with_bak(HOLD_FILE, {})
    if not isinstance(d, dict): return {}
    return _migrate_holdings_if_needed(d)

def load_holdings_for_account(account):
    nested = load_holdings_nested()
    h = nested.get(account, {})
    return h if isinstance(h, dict) else {}

def merge_holdings(nested):
    """合并所有账户持仓 → {code: {shares, cost, principal, buy_date}}。"""
    return domain_logic.merge_holdings(nested)

def save_holdings_nested(nested):
    """保存嵌套格式 {account: {code: rec}}。"""
    out = {}
    for acc, holdings in nested.items():
        if isinstance(holdings, dict):
            out[acc] = {k: _round_rec(v) for k, v in holdings.items()}
        else:
            out[acc] = holdings
    _atomic_write_json(HOLD_FILE, out, indent=2)

def _remove_code_from_all_accounts(code):
    """从所有账户中删除某只基金的持仓记录。"""
    nested = load_holdings_nested()
    for acc in list(nested):
        if isinstance(nested[acc], dict):
            nested[acc].pop(code, None)
    save_holdings_nested(nested)


def load_show_state():
    d = _load_json_with_bak(SHOW_FILE, {})
    if not isinstance(d, dict): return set()
    return set(str(c).strip() for c in (d.get("cleared") or []))


def save_show_state(cleared):
    try:
        _atomic_write_json(SHOW_FILE, {"cleared": sorted(cleared)}, indent=2)
    except Exception:
        pass


def _round_rec(rec):
    """落库前抹浮点尾巴：shares4位/cost4位/principal2位；其余键(如buy_date)原样。"""
    return domain_logic.round_holding_record(rec)


def apply_trade(holdings, code, action, *, nav=None, amount=None, shares=None,
                date=None, realized_out=None, cashflow_out=None, account=DEFAULT_ACCOUNT):
    """在 holdings(完整dict) 上原地记一笔。返回该只最新快照 dict。
       约定：调用方必须先 d=load_holdings_for_account(acc) 读全，改完 save_holdings_nested 写全，
       不可拿单只 dict 去保存，否则覆盖其余基金。
       action ∈ {buy, sell, dividend_reinvest, dividend_cash, convert}。
       account 参数指定该笔交易属于哪个账户。
    """
    h = holdings.setdefault(code, {"shares": 0.0, "cost": 0.0, "principal": 0.0})
    sh = float(h.get("shares", 0) or 0)
    cost = float(h.get("cost", 0) or 0)
    prin = float(h.get("principal", 0) or 0)

    if action == "buy":
        if amount:
            add_sh = float(amount) / float(nav)
            new_prin = prin + float(amount)
        elif shares:
            add_sh = float(shares)
            new_prin = prin + add_sh * float(nav)
        else:
            raise ValueError("买入需填金额或份额")
        new_sh = sh + add_sh
        new_cost = (new_prin / new_sh) if new_sh else 0.0
        h["shares"], h["cost"], h["principal"] = new_sh, new_cost, new_prin
        if date and not (h.get("buy_date") or "").strip():
            h["buy_date"] = date.strip()
    elif action == "sell":
        sell_sh = float(shares)
        if sell_sh > sh + 1e-9:
            raise ValueError(f"赎回份额 {sell_sh} 超过持有 {sh}")
        cost_out = cost * sell_sh                 # 按成本摊出本金
        realized = sell_sh * (float(nav) - cost)  # 已实现盈亏
        new_sh = sh - sell_sh
        new_prin = prin - cost_out
        h["shares"], h["principal"] = new_sh, max(new_prin, 0.0)
        if new_sh <= 1e-9:
            h["shares"], h["cost"], h["principal"] = 0.0, 0.0, 0.0
        if realized_out is not None: realized_out[0] = realized
    elif action == "dividend_reinvest":
        add_sh = float(amount) / float(nav)
        h["shares"] = sh + add_sh                 # principal 不动
    elif action == "dividend_cash":
        pass                                      # 三值全不动
        if cashflow_out is not None: cashflow_out[0] = float(amount)
    elif action == "convert":
        sell_sh = float(shares)
        cost_out = cost * sell_sh
        realized = sell_sh * (float(nav) - cost)
        new_sh = sh - sell_sh; new_prin = prin - cost_out
        h["shares"], h["principal"] = new_sh, max(new_prin, 0.0)
        if new_sh <= 1e-9:
            h["shares"], h["cost"], h["principal"] = 0.0, 0.0, 0.0
        if realized_out is not None: realized_out[0] = realized
    else:
        raise ValueError(f"未知 action: {action}")

    # —— 同步流水到 trades.json（收益日历时间线的唯一账本）——
    _rec = None
    if action == "buy":
        _rec = {"side": "buy", "amount": round(float(amount), 2)}
    elif action == "dividend_reinvest":
        _rec = {"side": "dividend_reinvest", "amount": round(float(amount), 2),
                "shares": round(float(amount)/float(nav), 4)}
    elif action == "dividend_cash":
        _rec = {"side": "dividend_cash", "amount": round(float(amount), 2)}
    elif action in ("sell", "convert"):
        _rec = {"side": "sell", "amount": round(float(shares) * float(nav), 2)}
    if _rec is not None:
        _rec["date"] = (date or datetime.now().strftime("%Y-%m-%d"))
        _rec["code"] = code
        _rec["account"] = account
        _all = load_trades()
        if not any(t.get("code") == code and t.get("side") == "open" for t in _all):
            _all.append({"date": (h.get("buy_date") or _rec["date"]), "code": code,
                         "side": "open", "shares": round(sh, 4), "account": account})
        _all.append(_rec)
        save_trades(_all)
    return h


def _http(url):
    return market_data.http_text(url, ssl_context=_SSL)


def resolve_holdings(holdings, price_map):
    return domain_logic.resolve_holdings(holdings, price_map)


# ---------- v1.3 参考信号（纯函数，无副作用，selfcheck 可测；仅信息展示不构成建议） ----------
def nav_percentile(nav_list, nav):
    """当前净值在历史净值序列中的百分位（0~100）；数据不足 60 个或净值无效返回 None。"""
    return domain_logic.nav_percentile(nav_list, nav)

def val_level(pct):
    """估值分档：>=80 过热 / <=20 低估 / 其余中性；None 原样返回。"""
    return domain_logic.val_level(pct)

def take_profit_level(pct):
    """累计收益率触及的止盈参考线：>=20 → 20；>=15 → 15；否则 None。"""
    return domain_logic.take_profit_level(pct)

def concentration_stats(resolved, price_map):
    """持仓集中度：按当前市值算总市值/最大单只占比/前三大占比。无有效持仓返回 None。"""
    return domain_logic.concentration_stats(resolved, price_map, NAME_MAP)


# ---------- v2.1.0 卡片迷你走势（近 60 交易日净值 sparkline：轻量拉取 + 按日缓存） ----------
SPARK_FILE = "走势缓存.json"   # v2.1.0 迷你走势缓存（按日失效；仅缓存层，不属持仓/交易账本）
SPARK_DAYS = market_data.SPARK_DAYS

def load_spark_cache():
    d = _load_json_with_bak(SPARK_FILE, {})
    return d if isinstance(d, dict) else {}

def save_spark_cache(d):
    try:
        _atomic_write_json(SPARK_FILE, d)
    except Exception:
        pass

def spark_from_lsjz(txt):
    """v2.1.0 纯函数：lsjz 接口单页 JSON 文本 → 净值序列（新在前原样返回，合并/截断由调用方做）。
       脏数据/空值丢弃；解析失败返回空列表。"""
    return market_data.spark_from_lsjz(txt)

def _lsjz_page(code, page, size=SPARK_DAYS):
    return market_data._lsjz_page(code, page, size, _SSL)

def fetch_spark(code):
    """v2.1.0：lsjz 分页接口轻量拉近 N 日净值（比 fetch_history 全历史小几十倍）。
       实测 pageSize 被接口忽略、单页固定 20 条 → 拉三页合并再截断，返回旧在前。"""
    return market_data.fetch_spark(code, days=SPARK_DAYS, ssl_context=_SSL)


# ---------- v1.4 估值口径升级（指数基金走跟踪指数 PE/PB 百分位，行业标准口径） ----------
INDEX_VALUATION_URL = market_data.INDEX_VALUATION_URL
PB_FIRST_KEYWORDS = domain_logic.PB_FIRST_KEYWORDS
INDEX_ALIASES = domain_logic.INDEX_ALIASES
NO_VAL_KEYWORDS = domain_logic.NO_VAL_KEYWORDS
DUAL_METRIC_INDICES = domain_logic.DUAL_METRIC_INDICES

def fund_valuation_class(fund_name):
    """v1.4：判无估值口径的基金族（债基/货币/商品）→ 标签；普通基金返回 None。
       债基净值靠票息稳步向上、黄金原油无盈利概念，百分位信号只会误导。"""
    return domain_logic.fund_valuation_class(fund_name)

def fetch_index_valuation(timeout=8):
    """v1.4：拉取蛋卷指数估值（非官方接口，可能随时变动/限流）→ {指数名: {pe, pe_pct, pb, pb_pct}}。"""
    return market_data.fetch_index_valuation(timeout=timeout, ssl_context=_SSL)

def match_index(fund_name, val_map):
    """v1.4：在基金名里匹配跟踪指数名（最长优先，别名兜底）；未命中返回 None。"""
    return domain_logic.match_index(fund_name, val_map)

def pick_index_pct(idx_name, val):
    """v1.4：红利/金融/周期族指数先看 PB 百分位，其余先看 PE 百分位；指标无效自动换另一指标。
       跨行业宽基（DUAL_METRIC_INDICES）PE+PB 结合：返回两项均值定档，metric 为 "PE78/PB53" 双值标签。
       返回 (百分位, 指标名)，指标都无效返回 (None, None)。"""
    return domain_logic.pick_index_pct(idx_name, val)


# ---------- v1.5 估值深化（信号文本/详情文本/卡片排序，纯函数可测） ----------
def val_signal_text(info):
    """v1.5：卡片估值信号行文本。指数源带股息率；无效返回 None。"""
    return domain_logic.val_signal_text(info)

def val_detail_text(info):
    """v1.5：估值完整一行文本（详情页/tooltip 用）。无信息返回 ""。"""
    return domain_logic.val_detail_text(info)

def sort_card_codes(codes, mode, val_map=None, chg_map=None, mv_map=None):
    """v1.5：卡片排序。default/val_asc(低估优先)/val_desc/chg_desc/mv_desc；无数据排最后。"""
    return domain_logic.sort_card_codes(codes, mode, val_map, chg_map, mv_map)


def fetch_one(code):
    return market_data.fetch_one(code, ssl_context=_SSL)


def search_funds(key):
    """按名字/代码模糊搜基金（东财 suggest），返回 [(code, name)]；网络/解析失败返回 []。"""
    return market_data.search_funds(key, ssl_context=_SSL)


def fetch_history(code):
    return market_data.fetch_history(code, ssl_context=_SSL)


def fetch_index_kline(secid):
    return market_data.fetch_index_kline(secid, ssl_context=_SSL)


# ---------- 粘贴文本解析 ----------
def _coerce_item(it):
    """单只基金dict: 各种字段名/别名 → 标准键。给啥认啥, 认不到返回None。"""
    if not isinstance(it, dict):
        return None
    def g(*keys):
        for k in keys:
            if k in it and it[k] not in (None, ""):
                return it[k]
        return None
    def fnum(x):
        try: return float(str(x).replace(",", "").replace("%", "").strip())
        except Exception: return None
    name = g("name", "名称", "基金名称", "fund_name", "title")
    code = g("code", "基金代码", "fund_code", "ts_code")
    if isinstance(code, str):
        m = re.search(r"\d{6}", code.strip()); code = m.group(0) if m else code.strip()
    amt  = g("amount", "持有金额", "市值", "持有市值", "market_value", "hold_amount", "value")
    hp   = g("hold_pnl", "持有收益", "持有盈亏", "累计收益", "累计盈亏", "收益", "pnl", "cum_pnl")
    hpr  = g("hold_pnl_rate_pct", "持有收益率", "累计收益率", "收益率", "收益比例", "rate", "pnl_rate")
    sh   = g("shares", "持有份额", "份额", "hold_shares")
    cost = g("cost", "持仓成本价", "持仓成本", "成本价", "每份成本", "unit_cost", "avg_cost")
    prin = g("principal", "投入本金", "本金", "cost_amount")
    isc  = g("is_cash", "cash")
    typ  = g("asset_class", "类型", "分类", "category")
    if isc is None and typ is not None:
        t = str(typ); isc = ("现金" in t) or ("货币" in t) or ("cash" in t.lower())
    out = {"name": str(name) if name else "", "code": str(code) if code else ""}
    for k, v in (("amount", amt), ("hold_pnl", hp), ("hold_pnl_rate_pct", hpr),
                 ("shares", sh), ("cost", cost), ("principal", prin)):
        fv = fnum(v)
        if fv is not None: out[k] = fv
    if isc is not None:
        out["is_cash"] = bool(isc)
    elif name and ("余额宝" in str(name) or "现金" in str(name)):
        out["is_cash"] = True
    return out if (out.get("name") or out.get("code")) else None

def _normalize_snap_doc(raw):
    """任意json → 标准外壳 {holdings:[标准元素], totals:{amount,n_fund,n_cash}, nav_date}。
       （OCR 已移除，此函数暂保留供将来手动导入 json 复用。）"""
    items_raw = []; nav_date = "?"
    if isinstance(raw, list):
        items_raw = raw
    elif isinstance(raw, dict):
        nav_date = raw.get("nav_date") or raw.get("date") or raw.get("recorded_at") or "?"
        if isinstance(raw.get("holdings"), list): items_raw = raw["holdings"]
        elif isinstance(raw.get("funds"), list):  items_raw = raw["funds"]
        elif isinstance(raw.get("data"), list):   items_raw = raw["data"]
        else:
            meta = {"holdings", "funds", "data", "totals", "total", "summary",
                    "nav_date", "date", "recorded_at"}
            for k, v in raw.items():
                if k in meta or not isinstance(v, dict): continue
                vv = dict(v)
                if not vv.get("code") and not vv.get("name"):
                    vv["code"] = k if re.search(r"^\d{6}$", str(k).strip()) else ""
                    if not vv["code"]: vv["name"] = str(k)
                items_raw.append(vv)
    items = [c for c in (_coerce_item(x) for x in items_raw) if c]
    tot_amt = 0.0; n_cash = 0
    for c in items:
        if c.get("is_cash"): n_cash += 1
        if c.get("amount"): tot_amt += float(c["amount"])
    if isinstance(raw, dict) and isinstance(raw.get("totals"), dict):
        try:
            ta = float(raw["totals"].get("amount", 0) or 0)
            if ta > 0: tot_amt = ta
        except Exception: pass
    return {"holdings": items,
            "totals": {"amount": round(tot_amt, 2), "n_fund": len(items) - n_cash, "n_cash": n_cash},
            "nav_date": nav_date}

def parse_holdings_text(text):
    found = {}
    if not text:
        return found
    positions = []
    for code, name in FUNDS:
        pat = r"\s*".join(re.escape(ch) for ch in name)
        for m in re.finditer(pat, text):
            positions.append((m.start(), code, name))
    if not positions:
        return found
    positions.sort(key=lambda x: x[0])
    for idx, (start, code, name) in enumerate(positions):
        end = positions[idx+1][0] if idx+1 < len(positions) else len(text)
        block = text[start:end]
        shares = cost = None
        m = re.search(r"持有份额[^\d\-]{0,6}([\d,]+\.?\d*)", block)
        if m:
            try: shares = float(m.group(1).replace(",", ""))
            except Exception: pass
        m = re.search(r"持仓成本价?[^\d\-]{0,6}([\d,]+\.?\d*)", block)
        if m:
            try: cost = float(m.group(1).replace(",", ""))
            except Exception: pass
        if shares is not None or cost is not None:
            rec = found.get(code, {})
            if shares is not None: rec["shares"] = shares
            if cost is not None: rec["cost"] = cost
            found[code] = rec
    return found


class Worker(QThread):
    done = Signal(list)
    def run(self):
        from concurrent.futures import ThreadPoolExecutor, as_completed
        def _one(code_name):
            code, name = code_name
            r = fetch_one(code)
            if not r.get("name"): r["name"] = name
            return r
        res = [None] * len(FUNDS)
        with ThreadPoolExecutor(max_workers=8) as ex:
            fut2i = {ex.submit(_one, cn): i for i, cn in enumerate(FUNDS)}
            for fut in as_completed(fut2i):
                i = fut2i[fut]
                try:
                    res[i] = fut.result()
                except Exception as e:
                    res[i] = {"code": FUNDS[i][0], "name": FUNDS[i][1], "nav": 0,
                              "est": 0, "chg": 0, "status": "fail", "err": f"并发异常:{e}"}
        self.done.emit(res)


class HistWorker(QThread):
    done = Signal(str, str, list, object); fail = Signal(str, str)
    def __init__(self, code): super().__init__(); self.code = code
    def run(self):
        try:
            name, hist, rank_by_ts = fetch_history(self.code)
            self.done.emit(self.code, name or NAME_MAP.get(self.code, ""), hist, rank_by_ts)
        except Exception as e:
            self.fail.emit(self.code, str(e))

class PnlWorker(QThread):
    """收益日历后台抓取：每只基金全历史净值 → {日期: 净值}"""
    progress = Signal(int, int)
    done = Signal(dict)
    def __init__(self, codes):
        super().__init__(); self.codes = list(codes)
    def run(self):
        from concurrent.futures import ThreadPoolExecutor
        out = {}
        def one(code):
            try:
                _, hist, _ = fetch_history(code)
                return code, {datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d"): nav for ts, nav, _e in hist}
            except Exception:
                return code, None
        with ThreadPoolExecutor(max_workers=6) as ex:
            for i, (code, m) in enumerate(ex.map(one, self.codes)):
                if m: out[code] = m
                self.progress.emit(i + 1, len(self.codes))
        self.done.emit(out)

class IndexWorker(QThread):
    done = Signal(str, str, list); fail = Signal(str, str, str)
    def __init__(self, secid, name): super().__init__(); self.secid = secid; self.name = name
    def run(self):
        try:
            data = fetch_index_kline(self.secid)
            self.done.emit(self.secid, self.name, data)
        except Exception as e:
            tb = traceback.format_exc(limit=2).replace("\n", " ")
            self.fail.emit(self.secid, self.name, f"{type(e).__name__}:{e} | {tb[-150:]}")


class UpdateWorker(QThread):
    found = Signal(str, str)
    checked = Signal(bool, str, str)  # v1.0.0：“关于”对话框手动检查用（是否有新版, tag, url）；启动自动检查不连此信号，不受影响
    def __init__(self, timeout=3):
        super().__init__(); self._timeout = timeout
    def run(self):
        try:
            req = urllib.request.Request(
                f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
                headers={"User-Agent": "fund-daily-report", "Accept": "application/vnd.github+json"})
            with urllib.request.urlopen(req, timeout=self._timeout, context=_SSL) as r:
                d = json.loads(r.read().decode("utf-8", errors="ignore"))
            tag = (d.get("tag_name") or "").strip().lstrip("vV")
            if tag and self._ver_tuple(tag) > self._ver_tuple(APP_VERSION):
                self.found.emit(tag, d.get("html_url", ""))
                self.checked.emit(True, tag, d.get("html_url", ""))
                return
            self.checked.emit(False, APP_VERSION, "")
        except Exception:
            self.checked.emit(False, "", "")  # 无网/失败：启动自动检查没连 checked，依旧静默跳过；windowed 版无 stdout，绝不可 print
    def _ver_tuple(self, s):
        t = []
        for x in (s or "").strip().split("."):
            n = ""
            for ch in x:
                if ch.isdigit(): n += ch
                else: break
            t.append(int(n) if n else 0)
        while len(t) < 3: t.append(0)
        return tuple(t[:3])


class SearchWorker(QThread):
    """v1.1：名字搜索走后台线程，网络等待不冻界面（耗时操作走子线程铁律）。"""
    done = Signal(list)
    def __init__(self, key):
        super().__init__(); self.key = key
    def run(self):
        self.done.emit(search_funds(self.key))


class ValWorker(QThread):
    """v1.3：估值信号后台拉取，不阻塞主线程。
       v1.4 升级：指数基金用蛋卷指数 PE/PB 百分位（行业标准口径）；
       主动/商品/未覆盖指数回退近 1 年净值百分位；接口失败全部回退。"""
    done = Signal(dict)
    def __init__(self, nav_map, name_map=None):
        super().__init__(); self.nav_map = dict(nav_map); self.name_map = dict(name_map or {})
    def run(self):
        from concurrent.futures import ThreadPoolExecutor
        out = {}
        try:
            val_map = fetch_index_valuation()
        except Exception:
            val_map = {}  # 非官方接口失败 → 全部回退净值百分位，不影响主流程
        nav_todo = {}
        for code, nav in self.nav_map.items():
            name = self.name_map.get(code, "")
            cls = fund_valuation_class(name)  # 债基/货币/商品无估值口径，不出信号避免误导
            if cls:
                out[code] = {"pct": None, "src": "na", "metric": cls}
                continue
            idx = match_index(name, val_map) if val_map else None
            if idx:
                pct, metric = pick_index_pct(idx, val_map[idx])
                if pct is not None:
                    v = val_map[idx]
                    out[code] = {"pct": pct, "metric": metric, "src": "index", "idx": idx,
                                 "pe": v.get("pe"), "pb": v.get("pb"),
                                 "pe_pct": v.get("pe_pct"), "pb_pct": v.get("pb_pct"),
                                 "yield_pct": v.get("yield_pct")}  # v1.5：带全量估值数据供卡片/详情页展示
                    continue
            nav_todo[code] = nav
        def one(code):
            try:
                _, hist, _ = fetch_history(code)
                vals = [nav for _ts, nav, _e in hist if nav and nav > 0][-252:]
                return code, nav_percentile(vals, self.nav_map.get(code, 0))
            except Exception:
                return code, None
        if nav_todo:
            with ThreadPoolExecutor(max_workers=6) as ex:
                for code, pct in ex.map(one, list(nav_todo.keys())):
                    out[code] = {"pct": pct, "metric": "净值", "src": "nav"}
        self.done.emit(out)


class SparkWorker(QThread):
    """v2.1.0：卡片迷你走势后台拉取（近 60 日净值），不阻塞主线程；
       单只失败静默丢弃（走势拉不到不影响卡片其余数据，线位留白）。"""
    done = Signal(dict)
    def __init__(self, codes):
        super().__init__(); self.codes = list(codes)
    def run(self):
        from concurrent.futures import ThreadPoolExecutor
        out = {}
        def one(code):
            try: return code, fetch_spark(code)
            except Exception: return code, None
        if self.codes:
            with ThreadPoolExecutor(max_workers=6) as ex:
                for code, pts in ex.map(one, self.codes):
                    if pts: out[code] = pts
        self.done.emit(out)


class DateAxis(pg.AxisItem):
    def __init__(self, *a, **k):
        super().__init__(*a, **k); self._hist = []; self._fmt = "%m-%d"
    def set_data(self, hist, fmt):
        self._hist = hist; self._fmt = fmt
    def tickStrings(self, values, scale, spacing):
        out = []; n = len(self._hist)
        for v in values:
            i = int(round(v))
            out.append(datetime.fromtimestamp(self._hist[i][0]/1000).strftime(self._fmt) if 0 <= i < n else "")
        return out


class BarChart(pg.PlotWidget):
    def __init__(self):
        super().__init__(); self.setBackground("#ffffff"); self.showGrid(x=False, y=True, alpha=80)
        self.getAxis("left").setLabel("今日涨跌 %"); self.setMouseEnabled(x=False, y=False); self.hideButtons()
        self._bar_item = None; self._anim_timer = None
        self._bars_x = []; self._bars_vals = []; self._bars_colors = []; self._bars_names = []

    def apply_theme(self):
        """v2.0.0/v2.1.0：柱状图按主题令牌着色（背景/双轴刻度文字/柱顶数值标签）。"""
        t = T()
        self.setBackground(QColor(t["card_bg"]))
        for _an in ("left", "bottom"):  # v2.1.0 返工：底轴（基金名）也要刷，此前只刷左轴
            ax = self.getAxis(_an)
            ax.setTextPen(QColor(t["muted"])); ax.setPen(QColor(t["card_border"]))
        for lb in getattr(self, "_val_labels", []):  # 已画出的柱顶数值标签同步重着
            lb.setColor(QColor(t["text"]))

    def draw(self, items, animate=True):
        self.clear(); self._bar_item = None
        if self._anim_timer:
            self._anim_timer.stop(); self._anim_timer = None
        vals = [it[1] for it in items]
        if not any(v != 0 for v in vals):
            t = pg.TextItem("暂无涨跌数据", color=(150,150,150), anchor=(0.5,0.5)); t.setFont(QFont(FONT,11)); self.addItem(t); t.setPos(len(vals)/2,0); return
        names = [it[0][:4] for it in items]; colors = [RED if v >= 0 else GREEN for v in vals]; x = list(range(len(vals)))
        self._bars_x = x; self._bars_vals = vals; self._bars_colors = colors; self._bars_names = names
        self._bar_item = pg.BarGraphItem(x=x, height=[0.0]*len(vals), width=0.6, brushes=colors, pens=colors)
        self.addItem(self._bar_item)
        vmin = min(vals); vmax = max(vals); span = max(vmax - vmin, 0.1); pad = max(span * 0.4, 0.35)
        self.setYRange(vmin - pad, vmax + pad, padding=0)
        self.getAxis("bottom").setTicks([list(zip(x, names))])
        if animate:
            self._step = 0; self._anim_timer = QTimer(self); self._anim_timer.timeout.connect(self._tick); self._anim_timer.start(28)
        else:
            self._finish()

    def _tick(self):
        self._step += 1; p = min(self._step / 12.0, 1.0); e = 1 - (1 - p) ** 3
        if self._bar_item:
            self._bar_item.setOpts(height=[v * e for v in self._bars_vals])
        if p >= 1.0:
            self._anim_timer.stop(); self._anim_timer = None; self._finish()

    def _finish(self):
        if self._bar_item:
            self._bar_item.setOpts(height=self._bars_vals)
        self._val_labels = []
        for xi, v in zip(self._bars_x, self._bars_vals):
            # v2.1.0 返工：柱顶数值标签改主题令牌（原硬编码纯黑，暗色主题下隐形）
            t = pg.TextItem(f"{v:+.2f}%", color=QColor(T()["text"]), anchor=(0.5, 1 if v >= 0 else 0)); t.setFont(QFont(FONT,8)); self.addItem(t); t.setPos(xi, v)
            self._val_labels.append(t)


class ElideLabel(QLabel):
    def __init__(self, *a, **k):
        super().__init__(*a, **k); self._full = ""
    def setText(self, t):
        self._full = t or ""; super().setText(self._full); self._elide()
    def resizeEvent(self, e):
        super().resizeEvent(e); self._elide()
    def _elide(self):
        w = self.width()
        if w <= 0 or not self._full:
            return
        super().setText(self.fontMetrics().elidedText(self._full, Qt.ElideRight, w))


class Sparkline(QWidget):
    """v2.1.0：卡片迷你走势线——近 60 交易日净值 QPainter 直画（FD-HIG 图形资产铁律）。
       区间涨=信号红、跌=信号绿（与涨跌幅语义一致）；灰章置 faint；
       颜色在 paintEvent 实时读 T()，主题切换后 update() 即自动跟随。"""
    def __init__(self):
        super().__init__()
        self._pts = []; self._dim = False
        self.setFixedHeight(28); self.setMinimumWidth(96); self.setMaximumWidth(150)
        self.setToolTip("近 60 交易日净值走势")
    def set_points(self, pts):
        self._pts = [float(p) for p in (pts or []) if p and float(p) > 0]; self.update()
    def set_dim(self, on):
        self._dim = bool(on); self.update()
    def paintEvent(self, ev):
        pts = self._pts
        if len(pts) < 2: return  # 无数据留白（降级不画）
        t = T()
        col = QColor(t["faint"] if self._dim else (t["up"] if pts[-1] >= pts[0] else t["down"]))
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing, True)
        w, h = self.width(), self.height(); pad = 2.0
        lo, hi = min(pts), max(pts); rng = (hi - lo) or 1.0; n = len(pts)
        path = QPainterPath()
        for i, v in enumerate(pts):
            x = pad + (w - 2*pad) * i / (n - 1)
            y = pad + (h - 2*pad) * (1 - (v - lo) / rng)
            if i: path.lineTo(x, y)
            else: path.moveTo(x, y)
        pen = QPen(col, 1.6); pen.setCapStyle(Qt.RoundCap); pen.setJoinStyle(Qt.RoundJoin)
        p.setPen(pen); p.drawPath(path)
        # 尾点小圆点：最新净值位置
        x = pad + (w - 2*pad); y = pad + (h - 2*pad) * (1 - (pts[-1] - lo) / rng)
        p.setPen(Qt.NoPen); p.setBrush(col); p.drawEllipse(QPointF(x, y), 2.2, 2.2)


class FundCard(QFrame):
    def __init__(self, code):
        super().__init__(); self.code = code; self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(card_qss_normal())  # v1.7 令牌化
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        # v2.1.0 逐屏精细化·卡片：三行结构（名称行｜估值+走势行｜持仓摘要+按钮行），涨跌幅升主角
        lay = QVBoxLayout(self); lay.setContentsMargins(14,10,14,10); lay.setSpacing(5)
        r1 = QHBoxLayout(); r1.setSpacing(8)  # 行一：名称+代码·净值 ｜ 大字涨跌幅
        self.lbl_name = ElideLabel("—"); self.lbl_name.setFont(QFont(FONT,11,QFont.Bold))
        self.lbl_name.setStyleSheet(f"color:{T()['text']};")  # v2.1.0 返工：名称无默认主题色，暗底隐形
        self.lbl_name.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred); self.lbl_name.setMinimumWidth(0)
        self.lbl_code = QLabel(code); self.lbl_code.setFont(QFont(FONT,8)); self.lbl_code.setStyleSheet(f"color:{T()['muted']};")
        self.lbl_nav = QLabel("净值 —"); self.lbl_nav.setFont(QFont(FONT,8)); self.lbl_nav.setStyleSheet(f"color:{T()['muted']};")
        r1.addWidget(self.lbl_name); r1.addWidget(self.lbl_code); r1.addWidget(self.lbl_nav); r1.addStretch()
        self.lbl_chg = QLabel("—"); self.lbl_chg.setFont(QFont(FONT,22,QFont.Bold))  # v2.1.0：16→22 大字排版
        self.lbl_chg.setAlignment(Qt.AlignRight|Qt.AlignVCenter)
        r1.addWidget(self.lbl_chg); lay.addLayout(r1)
        r2 = QHBoxLayout(); r2.setSpacing(8)  # 行二：估值信号 ｜ 迷你走势
        self.lbl_val = QLabel(""); self.lbl_val.setFont(QFont(FONT,8))  # v1.3 估值红绿灯
        self.lbl_val.setToolTip("估值参考（v1.4）：指数基金看跟踪指数 PE/PB 百分位（红利/金融族看 PB）；主动基金显示近 1 年净值分位；债基/货币/商品无估值口径不显示信号。≥80% 过热，≤20% 低估。仅信息展示，不构成投资建议。")
        r2.addWidget(self.lbl_val); r2.addStretch()
        self.spark = Sparkline(); r2.addWidget(self.spark); lay.addLayout(r2)
        r3 = QHBoxLayout(); r3.setSpacing(6)  # 行三：持有·今日·累计一行合并 ｜ 按钮
        self.lbl_mv = QLabel(""); self.lbl_mv.setFont(QFont(FONT,9,QFont.Bold)); self.lbl_mv.setStyleSheet(f"color:{T()['text']};")
        self._sep1 = QLabel("·"); self._sep1.setStyleSheet(f"color:{T()['faint']};"); self._sep1.setVisible(False)
        self.lbl_today = QLabel(""); self.lbl_today.setFont(QFont(FONT,9))
        self._sep2 = QLabel("·"); self._sep2.setStyleSheet(f"color:{T()['faint']};"); self._sep2.setVisible(False)
        self.lbl_pnl = QLabel(""); self.lbl_pnl.setFont(QFont(FONT,9))
        for _w in (self.lbl_mv, self._sep1, self.lbl_today, self._sep2, self.lbl_pnl): r3.addWidget(_w)
        r3.addStretch()
        self.btn_clear = QPushButton("标记清仓"); self.btn_clear.setFont(QFont(FONT,8))
        self.btn_clear.setCheckable(True)
        self.btn_detail = QPushButton("详情 →"); self.btn_detail.setFont(QFont(FONT,9))
        self._theme_btns()
        r3.addWidget(self.btn_clear); r3.addWidget(self.btn_detail); lay.addLayout(r3)

    def _sync_seps(self):
        """v2.1.0：持仓行『·』分隔符随相邻标签有无显隐。"""
        self._sep1.setVisible(bool(self.lbl_mv.text()) and bool(self.lbl_today.text()))
        self._sep2.setVisible(bool(self.lbl_today.text()) and bool(self.lbl_pnl.text()))

    def _theme_btns(self):
        """v2.0.0：卡片按钮按当前主题重刷（构建与主题切换共用）。"""
        t = T()
        self.btn_clear.setStyleSheet(f"QPushButton{{padding:4px 8px;border-radius:6px;background:{t['hover_bg']};color:{t['muted']};border:1px solid {t['card_border']};}}QPushButton:hover{{background:{t['card_hover']};}}QPushButton:checked{{background:{t['cleared_bg']};color:{t['text_sub']};border:1px dashed {t['cleared_border']};}}")
        if not getattr(self, "_cleared", False):
            self.btn_detail.setStyleSheet(f"QPushButton{{padding:8px 12px;border-radius:8px;background:{t['accent_soft']};color:{t['accent']};border:none;}}QPushButton:hover{{background:{t['accent_soft_hover']};}}")

    def set_cleared(self, on):
        self._cleared = bool(on)
        self.btn_clear.setChecked(self._cleared)
        self.btn_clear.setText("🚫 已清仓" if self._cleared else "标记清仓")
        self.btn_clear.setToolTip("已清仓: 不参与柱状图/红黑榜/总持仓统计, 卡片仅观察。点『恢复持有』还原。" if self._cleared else "卖出后盖灰章: 该只不再计入总账与榜单, 卡片置灰仅观察; 持仓记录不受影响, 可随时恢复。")
        self.spark.set_dim(self._cleared)  # v2.1.0：走势线随灰章置淡
        if self._cleared:
            self.setStyleSheet(card_qss_cleared())
            self.lbl_chg.setText("已清仓"); self.lbl_chg.setStyleSheet(f"color:{T()['muted']};font-size:13px;")
            self.lbl_val.setText("")  # v1.3：灰章不显示估值信号
            self.lbl_mv.setText("仅观察·不计入总账"); self.lbl_mv.setStyleSheet(f"color:{T()['faint']};")
            self.lbl_today.setText("—"); self.lbl_today.setStyleSheet(f"color:{T()['faint']};")
            self.lbl_pnl.setText("持仓记录保留"); self.lbl_pnl.setStyleSheet(f"color:{T()['faint']};")
            self._sync_seps()
            self.btn_detail.setEnabled(False); self.btn_detail.setStyleSheet(f"QPushButton{{padding:8px 12px;border-radius:8px;background:{T()['btn_disabled_bg']};color:{T()['faint']};border:none;}}")
        else:
            self.setStyleSheet(card_qss_normal())
            self.lbl_chg.setStyleSheet(f"color:{GRAY};font-size:22px;")
            self.btn_detail.setEnabled(True); self.btn_detail.setStyleSheet(f"QPushButton{{padding:8px 12px;border-radius:8px;background:{T()['accent_soft']};color:{T()['accent']};border:none;}}QPushButton:hover{{background:{T()['accent_soft_hover']};}}")

    def set_val_loading(self):
        if getattr(self, "_cleared", False): return
        self.lbl_val.setText("估值 ⏳"); self.lbl_val.setStyleSheet(f"color:{T()['faint']};")

    def set_val(self, info):
        """v1.3/v1.5：估值信号。info={pct, metric, src: index/nav/na, pe/pb/息…} 或 None。"""
        if getattr(self, "_cleared", False): return
        if info and info.get("src") == "na":  # 债基/货币/商品无估值口径，如实标注不出信号
            self.lbl_val.setText(val_detail_text(info)); self.lbl_val.setStyleSheet(f"color:{T()['faint']};")
            self.lbl_val.setToolTip(val_detail_text(info) + "｜仅信息展示，不构成投资建议")
            return
        txt = val_signal_text(info)
        if not txt:
            self.lbl_val.setText("估值 —"); self.lbl_val.setStyleSheet(f"color:{T()['faint']};")
            return
        self.lbl_val.setText(txt)
        col = {"hot": T()["up"], "cold": T()["down"]}.get(val_level(info["pct"]), T()["mid_val"])
        self.lbl_val.setStyleSheet(f"color:{col};")
        tip = val_detail_text(info)  # v1.5：悬停看完整估值数据
        if tip: self.lbl_val.setToolTip(tip + "｜仅信息展示，不构成投资建议")

    def update_data(self, d, resolved):
        if getattr(self, "_cleared", False): return          # 灰章状态: 行情不覆盖灰章
        self.lbl_name.setText(d.get("name") or "—"); nav = d.get("nav",0)
        if d.get("status") != "ok":
            self.lbl_nav.setText("净值 抓取失败"); self.lbl_nav.setStyleSheet("color:#e53935;")
            self.lbl_chg.setText("—"); self.lbl_chg.setStyleSheet(f"color:{GRAY};")
            self.lbl_mv.setText(""); self.lbl_mv.setStyleSheet(f"color:{T()['text']};")
            self.lbl_today.setText("今日 —"); self.lbl_today.setStyleSheet(f"color:{T()['faint']};")
            self.lbl_pnl.setText(f"⚠ {d.get('err','')[:18]}"); self.lbl_pnl.setStyleSheet("color:#e53935;font-size:8px;")
            self._sync_seps(); return
        self.lbl_nav.setText(f"净值 {nav:.4f}"); self.lbl_nav.setStyleSheet(f"color:{T()['muted']};")
        chg = d.get("chg",0); color = RED if chg>0 else (GREEN if chg<0 else GRAY)
        self.lbl_chg.setText(f"{chg:+.2f}%"); self.lbl_chg.setStyleSheet(f"color:{color};")
        nd = (d.get("nav_date") or "").strip()
        today_str = datetime.now().strftime("%Y-%m-%d")
        if nd and nd == today_str:
            day_tag = "今日"
        elif nd:
            day_tag = f"截至{nd[5:]}"
        else:
            day_tag = "今日"
        r2 = resolved.get(self.code)
        if r2 and r2.get("shares") and nav:
            sh = float(r2["shares"]); cost = float(r2.get("cost") or 0)
            mv = sh * nav
            self.lbl_mv.setText(f"持有 ¥{mv:,.2f}"); self.lbl_mv.setStyleSheet(f"color:{T()['text']};")
            if (100 + chg) != 0:
                today_pnl = sh * nav * chg / (100 + chg)
                tc = RED if today_pnl >= 0 else GREEN
                self.lbl_today.setText(f"{day_tag} {today_pnl:+,.2f}元"); self.lbl_today.setStyleSheet(f"color:{tc};")
            else:
                self.lbl_today.setText(f"{day_tag} +0.00元"); self.lbl_today.setStyleSheet(f"color:{GRAY};")
            if cost > 0:
                pnl = (nav-cost)*sh; pct = (nav-cost)/cost*100; pc = RED if pnl>=0 else GREEN
                tag = "🔧" if r2.get("corrected") else ""
                self.lbl_pnl.setText(f"{tag}累计 {pnl:+,.2f}元 ({pct:+.2f}%)"); self.lbl_pnl.setStyleSheet(f"color:{pc};")
            else:
                self.lbl_pnl.setText("累计 未填成本"); self.lbl_pnl.setStyleSheet("color:#bbb;")
        else:
            # 无手填份额 → 提示去管理持仓填写（OCR 快照回退已移除）
            self.lbl_mv.setText(""); self.lbl_mv.setStyleSheet(f"color:{T()['text']};")
            self.lbl_today.setText(f"{day_tag} —"); self.lbl_today.setStyleSheet(f"color:{T()['faint']};")
            self.lbl_pnl.setText("盈亏 未填持仓"); self.lbl_pnl.setStyleSheet(f"color:{T()['faint']};")
        self._sync_seps()

RANGE_DAYS = {"近1月": 30, "近3月": 90, "近6月": 180, "近1年": 365, "全部": None}

class DetailPage(QWidget):
    def __init__(self, on_back):
        super().__init__(); self.setObjectName("detailRoot"); self._hist=[]; self._full=[]; self._code=None; self._my_rec2={}
        self._dd=[]; self._dd_max=0.0; self._dd_max_idx=0
        self._dd_state="none"; self._dd_days=0; self._dd_progress=0.0; self._repair_idx=None; self._view="nav"
        self._buy_date=""
        self._cmp_code=None; self._cmp_name=""; self._cmp_on=False
        self._idx_cache={}; self._idx_err={}; self._fund_y_by_idx={}
        lay = QVBoxLayout(self); lay.setContentsMargins(18,14,18,14); lay.setSpacing(10)
        top = QHBoxLayout()
        self.btn_back = QPushButton("← 返回"); self.btn_back.setFont(QFont(FONT,10))
        self.btn_back.setStyleSheet("QPushButton{padding:8px 14px;border-radius:8px;background:#f0f0f0;border:none;}QPushButton:hover{background:#e3e3e3;}")
        self.btn_back.clicked.connect(on_back); top.addWidget(self.btn_back)
        self.lbl_title = QLabel("基金详情"); self.lbl_title.setFont(QFont(FONT,14,QFont.Bold)); top.addWidget(self.lbl_title)
        top.addStretch(); lay.addLayout(top)
        # v2.2.0 逐屏精细化·详情页：摘要大字区（净值/持有市值/累计盈亏）——详情页的主角
        t0 = T()
        sum_box = QFrame(); sum_box.setStyleSheet(panel_qss()); self.sum_box = sum_box
        sl = QHBoxLayout(sum_box); sl.setContentsMargins(16,10,16,10); sl.setSpacing(0)
        self._sum_caps = []
        def _sum_col(cap):
            box = QVBoxLayout(); box.setSpacing(2)
            a = QLabel(cap); a.setFont(QFont(FONT,8)); a.setAlignment(Qt.AlignCenter); a.setStyleSheet(f"color:{t0['muted']};")
            b = QLabel("—"); b.setFont(QFont(FONT,16,QFont.Bold)); b.setAlignment(Qt.AlignCenter); b.setStyleSheet(f"color:{t0['text']};")
            box.addWidget(a); box.addWidget(b); self._sum_caps.append(a); return box, b
        _c1, self.d_nav_val = _sum_col("最新净值")
        _c2, self.d_mv_val = _sum_col("持有市值")
        _c3, self.d_pnl_val = _sum_col("累计盈亏")
        for _c in (_c1, _c2, _c3): sl.addLayout(_c, 1)
        lay.addWidget(sum_box)
        self.d_sub = QLabel(""); self.d_sub.setFont(QFont(FONT,8)); self.d_sub.setStyleSheet(f"color:{t0['muted']};")
        self.d_sub.setAlignment(Qt.AlignCenter); lay.addWidget(self.d_sub)
        self.lbl_track = QLabel(""); self.lbl_track.setFont(QFont(FONT,8))
        self.lbl_track.setStyleSheet("QLabel{color:#6b7280;background:#f7f9fc;border:1px solid #eef0f3;border-radius:6px;padding:4px 8px;}")
        self.lbl_track.setWordWrap(True); self.lbl_track.hide(); lay.addWidget(self.lbl_track)
        self.lbl_valinfo = QLabel(""); self.lbl_valinfo.setFont(QFont(FONT,8))  # v1.5：详情页完整估值面板
        self.lbl_valinfo.setStyleSheet("QLabel{color:#6b7280;background:#f7f9fc;border:1px solid #eef0f3;border-radius:6px;padding:4px 8px;}")
        self.lbl_valinfo.setWordWrap(True); self.lbl_valinfo.hide(); lay.addWidget(self.lbl_valinfo)
        vbar = QHBoxLayout()
        self._vbg = QButtonGroup(self); self._vbtns = {}
        for i,(k,ico) in enumerate([("nav","净值走势"),("dd","回撤修复"),("rank","同类排名")]):
            b = QPushButton(ico); b.setCheckable(True); b.setFont(QFont(FONT,9))
            b.setStyleSheet("QPushButton{padding:6px 14px;border:1px solid #ddd;border-radius:7px;background:#fff;}"
                            "QPushButton:checked{background:#374151;color:#fff;border-color:#374151;}")
            self._vbg.addButton(b,i); self._vbtns[k]=b; vbar.addWidget(b)
        self._vbtns["nav"].setChecked(True); self._vbg.buttonClicked.connect(lambda _: self._set_view())
        vbar.addStretch()
        self._cmp_lbl = QLabel("对比"); self._cmp_lbl.setStyleSheet("color:#666;font-size:9px;"); vbar.addWidget(self._cmp_lbl)
        self._cmp_combo = QComboBox(); self._cmp_combo.setFont(QFont(FONT,9))
        self._cmp_combo.addItem("不对比", None)
        for secid, nm in CMP_INDEX:
            self._cmp_combo.addItem(nm, secid)
        self._cmp_combo.setStyleSheet("QComboBox{padding:4px 8px;border:1px solid #ddd;border-radius:7px;background:#fff;}")
        self._cmp_combo.currentIndexChanged.connect(self._on_cmp_change); vbar.addWidget(self._cmp_combo)
        self._cmp_hint = QLabel(""); self._cmp_hint.setStyleSheet("color:#b45309;font-size:8px;"); self._cmp_hint.setWordWrap(True); vbar.addWidget(self._cmp_hint)
        lay.addLayout(vbar)
        rbar = QHBoxLayout(); rbar.addStretch(); self._bg = QButtonGroup(self); self._rbtns = {}
        for i,k in enumerate(RANGE_DAYS.keys()):
            b = QPushButton(k); b.setCheckable(True); b.setFont(QFont(FONT,9))
            b.setStyleSheet("QPushButton{padding:6px 12px;border:1px solid #ddd;border-radius:7px;background:#fff;}QPushButton:checked{background:#2563eb;color:#fff;border-color:#2563eb;}")
            self._bg.addButton(b,i); self._rbtns[k]=b; rbar.addWidget(b)
        self._rbtns["近1年"].setChecked(True); self._bg.buttonClicked.connect(lambda _: self._apply_range()); lay.addLayout(rbar)
        self.dd_box = QFrame(); self.dd_box.setStyleSheet("QFrame{background:#fff5f5;border:1px solid #f3c2c2;border-radius:10px;}")
        dl = QHBoxLayout(self.dd_box); dl.setContentsMargins(14,10,14,10)
        self.lbl_dd_max = QLabel("最大回撤  —"); self.lbl_dd_max.setFont(QFont(FONT,11,QFont.Bold)); self.lbl_dd_max.setStyleSheet("color:#c0392b;")
        self.lbl_dd_rep = QLabel("修复  —"); self.lbl_dd_rep.setFont(QFont(FONT,11,QFont.Bold)); self.lbl_dd_rep.setStyleSheet("color:#888;")
        dl.addWidget(self.lbl_dd_max); dl.addSpacing(28); dl.addWidget(self.lbl_dd_rep); dl.addStretch()
        self.dd_box.hide(); lay.addWidget(self.dd_box)
        chart_box = QFrame(); chart_box.setStyleSheet(board_qss())  # v1.7 令牌化
        self.chart_box = chart_box  # v2.2.0 返工：存引用供 _apply_theme 重刷（防切主题后框色冻结）
        cl = QVBoxLayout(chart_box); cl.setContentsMargins(8,8,8,8)
        self._date_axis = DateAxis(orientation='bottom')
        self.plot = pg.PlotWidget(axisItems={'bottom': self._date_axis})
        self.plot.setBackground("#ffffff"); self.plot.showGrid(x=True,y=True,alpha=60)
        self.plot.setMouseEnabled(x=True,y=False); self.plot.hideButtons()
        self.plot.getAxis("bottom").setLabel("日期")
        self.plot.getAxis("left").setLabel("单位净值")
        self.plot.getAxis("left").enableAutoSIPrefix(False)
        self._curve = self.plot.plot(pen=pg.mkPen("#2563eb", width=2))
        self._dd_curve = self.plot.plot(pen=pg.mkPen("#ef4444", width=2))
        self._dd_curve.hide()
        self._cmp_curve = self.plot.plot(pen=pg.mkPen(TEAL, width=2))
        self._cmp_curve.hide()
        self._rank_curve = self.plot.plot(pen=pg.mkPen("#f59e0b", width=2))
        self._rank_curve.hide()
        self._rank_curve.setClipToView(True)
        self._curve.setClipToView(True); self._curve.setDownsampling(auto=True, method='peak')
        self._vline = pg.InfiniteLine(angle=90, movable=False, pen=pg.mkPen("#999",width=1,style=Qt.DashLine))
        self._hline = pg.InfiniteLine(angle=0, movable=False, pen=pg.mkPen("#999",width=1,style=Qt.DashLine))
        self._vline.hide(); self._hline.hide()
        self._dot = pg.ScatterPlotItem(size=10, brush=pg.mkBrush("#2563eb"), pen=pg.mkPen("#fff"))
        self.plot.addItem(self._vline); self.plot.addItem(self._hline); self.plot.addItem(self._dot)
        self._repair_region = pg.LinearRegionItem(values=[0,1], brush=pg.mkBrush(100,116,139,24), pen=pg.mkPen(None), movable=False)
        self._repair_region.setZValue(-1); self._repair_region.hide(); self.plot.addItem(self._repair_region)
        self._zero_line = pg.InfiniteLine(pos=0, angle=0, movable=False, pen=pg.mkPen("#10b981", width=1.6, style=Qt.DashLine))
        self._zero_line.hide(); self.plot.addItem(self._zero_line)
        self._zero_label = pg.TextItem("前高线 0%  回到此线=修复", color=(16,185,129), anchor=(1,0))
        self._zero_label.setFont(QFont(FONT,8,QFont.Bold)); self._zero_label.hide(); self.plot.addItem(self._zero_label)
        self._dd_marker = pg.ScatterPlotItem(size=18, pen=pg.mkPen("#fff",width=2), brush=pg.mkBrush("#dc2626"))
        self._dd_marker.hide(); self.plot.addItem(self._dd_marker)
        self._repair_marker = pg.ScatterPlotItem(size=15, pen=pg.mkPen("#fff",width=2), brush=pg.mkBrush("#16a34a"))
        self._repair_marker.hide(); self.plot.addItem(self._repair_marker)
        self._dd_label = pg.TextItem("", color=(220,38,38), anchor=(0.5,0)); self._dd_label.setFont(QFont(FONT,9,QFont.Bold)); self._dd_label.hide(); self.plot.addItem(self._dd_label)
        self._repair_label = pg.TextItem("", color=(22,163,74), anchor=(0.5,1)); self._repair_label.setFont(QFont(FONT,9,QFont.Bold)); self._repair_label.hide(); self.plot.addItem(self._repair_label)
        self._region_label = pg.TextItem("", color=(180,90,20), anchor=(0.5,0.5)); self._region_label.setFont(QFont(FONT,8,QFont.Bold)); self._region_label.hide(); self.plot.addItem(self._region_label)
        self._cost_line = pg.InfiniteLine(pos=0, angle=0, movable=False, pen=pg.mkPen("#7c3aed", width=1.6, style=Qt.DashLine))
        self._cost_line.hide(); self.plot.addItem(self._cost_line)
        self._cost_label = pg.TextItem("", color=(124,58,237), anchor=(1,0.5)); self._cost_label.setFont(QFont(FONT,8,QFont.Bold)); self._cost_label.hide(); self.plot.addItem(self._cost_label)
        self._buy_line = pg.InfiniteLine(pos=0, angle=90, movable=False, pen=pg.mkPen("#7c3aed", width=1.8, style=Qt.DashLine))
        self._buy_line.hide(); self.plot.addItem(self._buy_line)
        self._buy_dot = pg.ScatterPlotItem(size=16, pen=pg.mkPen("#fff",width=2), brush=pg.mkBrush("#7c3aed"))
        self._buy_dot.hide(); self.plot.addItem(self._buy_dot)
        self._buy_label = pg.TextItem("", color=(124,58,237), anchor=(0,1)); self._buy_label.setFont(QFont(FONT,9,QFont.Bold)); self._buy_label.hide(); self.plot.addItem(self._buy_label)
        self._buy_off = pg.TextItem("", color=(167,139,250), anchor=(0.5,0.5)); self._buy_off.setFont(QFont(FONT,9,QFont.Bold)); self._buy_off.hide(); self.plot.addItem(self._buy_off)
        self._legend = QLabel(""); self._legend.setParent(self.plot)
        self._legend.setStyleSheet("QLabel{background:rgba(255,255,255,225);border:1px solid #ddd;border-radius:6px;padding:4px 8px;color:#333;font-size:9px;}")
        self._legend.setFont(QFont(FONT,9)); self._legend.hide(); self._legend.raise_()
        self._readout = QLabel("移动鼠标看每日数值"); self._readout.setParent(self.plot)
        self._readout.setStyleSheet("QLabel{background:rgba(255,255,255,220);border:1px solid #ddd;border-radius:6px;padding:5px 8px;color:#333;}")
        self._readout.setFont(QFont(FONT,9)); self._readout.move(10,8); self._readout.raise_()
        self.plot.scene().sigMouseMoved.connect(self._mouse_moved)
        self._loading = QLabel("⏳  加载历史净值中…"); self._loading.setFont(QFont(FONT,12)); self._loading.setAlignment(Qt.AlignCenter); self._loading.setStyleSheet("color:#888;")
        self._stack_chart = QStackedWidget(); self._stack_chart.addWidget(self._loading); self._stack_chart.addWidget(self.plot)
        self._stack_chart.setFixedHeight(320); cl.addWidget(self._stack_chart); lay.addWidget(chart_box,3)
        self.table = QTableWidget(0,3); self.table.setHorizontalHeaderLabels(["日期","单位净值","当日涨跌"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers); self.table.setAlternatingRowColors(True); self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(False)
        self.table.setStyleSheet("QTableWidget{font-size:12px;}"); lay.addWidget(self.table,2)
        self._hist_worker = None; self._idx_worker = None
        self._rank_by_ts = {}; self._rank_full = []
        self._apply_theme()  # v2.0.0：按当前主题初始化详情页配色

    def _apply_theme(self):
        """v2.0.0：详情页主题适配（背景/面板/按钮组/图表）。"""
        t = T()
        self.setAttribute(Qt.WA_StyledBackground, True)  # v2.2.0 返工：纯 QWidget 的 QSS 背景必须显式开启才生效，否则填系统默认浅色（半黑灾难根因）
        self.setAutoFillBackground(True)
        self.setStyleSheet(f"QWidget#detailRoot{{background:{t['win_bg']};}}" if t["win_bg"] else "")
        self.lbl_title.setStyleSheet(f"color:{t['text']};")
        _tip = f"QLabel{{color:{t['text_sub']};background:{t['panel_bg']};border:1px solid {t['panel_border']};border-radius:6px;padding:4px 8px;}}"
        self.lbl_track.setStyleSheet(_tip); self.lbl_valinfo.setStyleSheet(_tip)
        self.btn_back.setStyleSheet(ghost_btn_qss(t, pad="8px 14px"))
        _vb = (f"QPushButton{{padding:6px 14px;border:1px solid {t['card_border']};border-radius:7px;background:{t['card_bg']};color:{t['text_sub']};}}"
               f"QPushButton:checked{{background:{t['accent']};color:#ffffff;border-color:{t['accent']};}}")
        for b in self._vbtns.values(): b.setStyleSheet(_vb)
        for b in self._rbtns.values(): b.setStyleSheet(_vb)
        self._cmp_combo.setStyleSheet(combo_qss(t))  # v2.2.0 返工：统一 Win11 风（含箭头子控件，原内联无子控件样式）
        self._cmp_lbl.setStyleSheet(f"color:{t['muted']};font-size:9px;")
        self._cmp_hint.setStyleSheet(f"color:{t['mid_val']};font-size:8px;")
        self.dd_box.setStyleSheet(f"QFrame{{background:{t['hover_bg']};border:1px solid {t['card_border']};border-radius:10px;}}")
        self._update_dd_stats()  # v2.2.0：回撤条文字按新令牌重着
        self.plot.setBackground(QColor(t["card_bg"]))
        for _axn in ("left", "bottom"):
            _a = self.plot.getAxis(_axn); _a.setTextPen(QColor(t["muted"])); _a.setPen(QColor(t["card_border"]))
        self.chart_box.setStyleSheet(board_qss())  # v2.2.0 返工：图表外框容器随主题重刷（用户捆出“走势图周围一圈白色”）
        # v2.2.0 返工：图表曲线/标记色随主题令牌（原主曲线写死 #2563eb，纸账本主题下按钮已变红曲线还是蓝）
        self._curve.setPen(pg.mkPen(t["accent"], width=2))
        self._dot.setBrush(pg.mkBrush(t["accent"]))
        self._dd_curve.setPen(pg.mkPen(t["up"], width=2))
        self._dd_marker.setBrush(pg.mkBrush(t["up"]))
        self._dd_label.setColor(QColor(t["up"]))
        self._zero_line.setPen(pg.mkPen(t["down"], width=1.6, style=Qt.DashLine))
        self._zero_label.setColor(QColor(t["down"]))
        self._repair_marker.setBrush(pg.mkBrush(t["down"]))
        self._repair_label.setColor(QColor(t["down"]))
        self._region_label.setColor(QColor(t["mid_val"]))
        # v2.2.0：摘要区/明细表/加载提示/图表浮层随主题重刷（双路径铁律）
        self.sum_box.setStyleSheet(panel_qss())
        for cap in self._sum_caps: cap.setStyleSheet(f"color:{t['muted']};")
        self._update_my_pnl()  # 数值颜色数据驱动，重算即重着
        self.d_sub.setStyleSheet(f"color:{t['muted']};")
        self.table.setStyleSheet(table_qss("font-size:12px;"))
        self._loading.setStyleSheet(f"color:{t['muted']};")
        _fl = f"QLabel{{background:{t['card_bg']};border:1px solid {t['card_border']};border-radius:6px;padding:4px 8px;color:{t['text_sub']};font-size:9px;}}"
        self._legend.setStyleSheet(_fl)
        self._readout.setStyleSheet(f"QLabel{{background:{t['card_bg']};border:1px solid {t['card_border']};border-radius:6px;padding:5px 8px;color:{t['text_sub']};}}")

    def load(self, code, rec2, val=None, chg=None):
        self._code = code; self._chg = chg; self.lbl_title.setText(f"{NAME_MAP.get(code,'')}  详情")
        vt = val_detail_text(val) if val else ""  # v1.5：完整估值面板
        if vt:
            self.lbl_valinfo.setText("🧭 " + vt + "｜仅信息展示，不构成投资建议"); self.lbl_valinfo.show()
        else:
            self.lbl_valinfo.hide()
        tk = TRACK.get(code)
        if tk:
            self.lbl_track.setText(f"📌 {tk[0]} ｜ {tk[1]}"); self.lbl_track.show()
        else:
            self.lbl_track.hide()
        self._reset_summary(); self._full=[]; self._hist=[]; self._dd=[]
        self._dd_state="none"; self._dd_days=0; self._dd_progress=0.0; self._repair_idx=None
        self._buy_date = load_holdings().get(code, {}).get("buy_date", "") or ""
        self._buy_date = self._buy_date.strip()
        self._cmp_on=False; self._fund_y_by_idx={}
        self._rank_by_ts = {}; self._rank_full = []
        self._curve.setData([]); self._dd_curve.setData([]); self._cmp_curve.setData([]); self._cmp_curve.hide()
        self._dot.setData([]); self._legend.hide()
        self._hide_dd_markers(); self._hide_buy_markers()
        self._vline.hide(); self._hline.hide()
        self._stack_chart.setCurrentWidget(self._loading); self._readout.setText("移动鼠标看每日数值")
        self._my_rec2 = rec2 or {}
        self._hist_worker = HistWorker(code); self._hist_worker.done.connect(self._on_hist); self._hist_worker.fail.connect(self._on_hist_fail); self._hist_worker.start()

    def _on_hist_fail(self, code, err): self._loading.setText(f"⚠ 历史净值加载失败：{err}")
    def _on_hist(self, code, name, hist, rank_by_ts):
        if code != self._code: return
        self._full = hist
        self._rank_by_ts = rank_by_ts or {}
        self._compute_rank()
        if name: self.lbl_title.setText(f"{name} 详情")
        self._apply_range()
        self._stack_chart.setCurrentWidget(self.plot)

    def _set_view(self):
        if self._vbtns["rank"].isChecked(): self._view = "rank"
        elif self._vbtns["dd"].isChecked(): self._view = "dd"
        else: self._view = "nav"
        self._draw_curve(); self._fill_table(); self._update_dd_stats()

    def _apply_range(self):
        days = None
        for k,b in self._rbtns.items():
            if b.isChecked(): days = RANGE_DAYS[k]; break
        if not self._full: return
        if days is None: self._hist = self._full
        else:
            cutoff = (datetime.now()-timedelta(days=days)).timestamp()*1000
            self._hist = [p for p in self._full if p[0] >= cutoff] or self._full
        self._compute_dd()
        self._compute_rank()
        self._draw_curve(); self._fill_table(); self._update_my_pnl(); self._update_dd_stats()

    def _on_cmp_change(self, idx):
        secid = self._cmp_combo.currentData(); name = self._cmp_combo.currentText()
        if not secid:
            self._cmp_code=None; self._cmp_name=""; self._cmp_on=False; self._cmp_hint.setText(""); self._draw_curve(); return
        self._cmp_code=secid; self._cmp_name=name
        if self._idx_cache.get(secid):
            self._cmp_hint.setText(""); self._draw_curve(); return
        self._cmp_hint.setText(f"⏳ 加载{name}（稍候）…"); self._cmp_combo.setEnabled(False)
        self._idx_worker = IndexWorker(secid, name)
        self._idx_worker.done.connect(self._on_idx_done); self._idx_worker.fail.connect(self._on_idx_fail)
        self._idx_worker.start()

    def _on_idx_done(self, secid, name, data):
        if secid != self._cmp_code: return
        self._idx_cache[secid] = data or []; self._cmp_combo.setEnabled(True)
        if not data:
            self._cmp_hint.setText(f"⚠ {name}无数据 secid={secid}")
        else:
            self._cmp_hint.setText("")
        self._draw_curve()

    def _on_idx_fail(self, secid, name, err):
        if secid != self._cmp_code: return
        self._idx_err[secid] = err; self._idx_cache[secid] = []; self._cmp_combo.setEnabled(True)
        self._cmp_hint.setText(f"⚠ {name}失败 secid={secid}：{err[:240]}")
        self._draw_curve()

    def _aligned_for_cmp(self):
        if not self._cmp_code or not self._idx_cache.get(self._cmp_code) or not self._hist:
            return []
        idx_dict = {d: c for d, c in self._idx_cache[self._cmp_code]}
        out = []
        for i, (ts, nav, eqt) in enumerate(self._hist):
            ds = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d")
            c = idx_dict.get(ds)
            if c is not None and c > 0:
                out.append((i, nav, c))
        return out

    def _compute_dd(self):
        self._dd = []; self._dd_max = 0.0; self._dd_max_idx = 0
        self._dd_state = "none"; self._dd_days = 0; self._dd_progress = 0.0; self._repair_idx = None
        peak = -1e18; peaks = []
        for i,(ts,nav,eqt) in enumerate(self._hist):
            if nav > peak: peak = nav
            peaks.append(peak)
            dd = (nav-peak)/peak*100.0 if peak > 0 else 0.0
            self._dd.append(dd)
            if dd < self._dd_max:
                self._dd_max = dd; self._dd_max_idx = i
        if self._dd_max > -0.01:
            self._dd_state = "none"; return
        di = self._dd_max_idx; peak_at_dd = peaks[di]
        for j in range(di+1, len(self._hist)):
            if self._hist[j][1] >= peak_at_dd:
                self._dd_state = "yes"
                self._dd_days = max(int(round((self._hist[j][0]-self._hist[di][0])/86400000.0)), 0)
                self._dd_progress = 1.0; self._repair_idx = j; return
        dmax = self._dd_max; dend = self._dd[-1]
        p = (dend - dmax)/(-dmax) if dmax < 0 else 0.0
        p = max(0.0, min(1.0, p)); self._dd_progress = p; self._repair_idx = len(self._hist)-1
        self._dd_state = "fixing" if p >= REPAIR_THRESHOLD else "no"

    def _compute_rank(self):
        self._rank_full = []
        if not self._rank_by_ts or not self._hist:
            return
        rkeys = sorted(self._rank_by_ts.keys())
        if not rkeys:
            return
        import bisect
        DAY = 86400000
        for i, (ts, nav, eqt) in enumerate(self._hist):
            p = bisect.bisect_left(rkeys, ts)
            best = None
            for cand in (rkeys[p-1] if p > 0 else None,
                         rkeys[p]   if p < len(rkeys) else None):
                if cand is not None and abs(cand - ts) <= DAY + 3600000:
                    if best is None or abs(cand - ts) < abs(best - ts):
                        best = cand
            if best is not None:
                self._rank_full.append((i, self._rank_by_ts[best]))

    def _hide_dd_markers(self):
        for it in (self._zero_line, self._zero_label, self._repair_region, self._dd_marker, self._repair_marker,
                   self._dd_label, self._repair_label, self._region_label):
            it.hide()

    def _hide_buy_markers(self):
        for it in (self._cost_line, self._cost_label, self._buy_line, self._buy_dot, self._buy_label, self._buy_off):
            it.hide()

    def _show_dd_markers(self):
        n = len(self._hist)
        if n == 0 or not self._dd:
            self._hide_dd_markers(); return
        self._zero_line.show(); self._zero_label.setPos(n-1, 0.0); self._zero_label.show()
        st = self._dd_state
        if st == "none":
            for it in (self._repair_region, self._dd_marker, self._repair_marker, self._dd_label, self._repair_label, self._region_label):
                it.hide()
            return
        di = self._dd_max_idx; dv = self._dd[di]
        self._dd_marker.setData([di], [dv]); self._dd_marker.show()
        self._dd_label.setText(f"▼ 最大回撤 {abs(dv):.2f}%"); self._dd_label.setPos(di, dv); self._dd_label.show()
        ri = self._repair_idx if self._repair_idx is not None else n-1
        a = min(di, ri); b = max(di, ri)
        if st == "yes":
            col = (22,163,74); brush = pg.mkBrush(22,163,74,32); rtxt = f"↔ 修复期 {self._dd_days} 天"
            ebrush = "#16a34a"; etxt = "▲ 已修复"; ey = 0.0
        elif st == "fixing":
            pct = self._dd_progress*100
            col = (234,88,12); brush = pg.mkBrush(234,88,12,30); rtxt = f"↔ 修复中 已填{pct:.0f}%"
            ebrush = "#ea580c"; etxt = f"● 修复中 {pct:.0f}%"; ey = self._dd[-1]
        else:
            pct = self._dd_progress*100
            col = (100,116,139); brush = pg.mkBrush(100,116,139,26)
            rtxt = f"↔ 仍在坑中 已填{pct:.0f}%" if pct > 1 else "↔ 仍在坑中"
            ebrush = "#64748b"; etxt = "● 暂未修复"; ey = self._dd[-1]
        if b > a:
            self._repair_region.setRegion([a, b]); self._repair_region.setBrush(brush); self._repair_region.show()
            self._region_label.setText(rtxt); self._region_label.setColor(col); self._region_label.setPos((a+b)/2.0, dv*0.5); self._region_label.show()
        else:
            self._repair_region.hide(); self._region_label.hide()
        self._repair_marker.setBrush(pg.mkBrush(ebrush)); self._repair_marker.setData([ri], [ey]); self._repair_marker.show()
        self._repair_label.setColor(col); self._repair_label.setText(etxt); self._repair_label.setPos(ri, ey); self._repair_label.show()

    def _draw_buy_markers(self):
        cost = float(self._my_rec2.get("cost") or 0)
        n = len(self._hist)
        if self._view == "nav" and not self._cmp_on and cost > 0 and n > 0:
            self._cost_line.setPos(cost); self._cost_line.show()
            self._cost_label.setText(f"成本线 {cost:.4f}"); self._cost_label.setPos(n-1, cost); self._cost_label.show()
        else:
            self._cost_line.hide(); self._cost_label.hide()
        bd = self._buy_date
        if not bd or n == 0:
            self._buy_line.hide(); self._buy_dot.hide(); self._buy_label.hide(); self._buy_off.hide(); return
        try:
            buy_ts = datetime.strptime(bd, "%Y-%m-%d").timestamp()*1000
        except Exception:
            self._buy_line.hide(); self._buy_dot.hide(); self._buy_label.hide(); self._buy_off.hide(); return
        t0 = self._hist[0][0]; t1 = self._hist[-1][0]
        if buy_ts < t0 or buy_ts > t1:
            self._buy_line.hide(); self._buy_dot.hide(); self._buy_label.hide()
            vr = self.plot.viewRange(); ymid = (vr[1][0]+vr[1][1])/2.0
            self._buy_off.setText(f"📍 首笔买入 {bd} 不在当前窗口（切『全部』查看）"); self._buy_off.setPos((n-1)/2.0, ymid); self._buy_off.show()
            return
        i = None
        for k,(ts,nav,eqt) in enumerate(self._hist):
            if ts >= buy_ts: i = k; break
        if i is None:
            self._buy_line.hide(); self._buy_dot.hide(); self._buy_label.hide(); self._buy_off.hide(); return
        self._buy_line.setPos(i); self._buy_line.show()
        yv = self._fund_y_by_idx.get(i)
        txt = f"📍 首笔买入 {bd[5:]}" + (f" @{cost:.4f}" if cost > 0 else "")
        if yv is None:
            self._buy_dot.hide()
            vr = self.plot.viewRange(); self._buy_label.setPos(i, vr[1][1]); self._buy_label.setText(txt); self._buy_label.show()
        else:
            self._buy_dot.setData([i], [yv]); self._buy_dot.show()
            self._buy_label.setText(txt); self._buy_label.setPos(i, yv); self._buy_label.show()
        self._buy_off.hide()

    def _place_legend(self):
        if not self._legend.text():
            self._legend.hide(); return
        self._legend.adjustSize(); self._legend.show(); self._legend.raise_()
        w = self.plot.width()
        if w > 140:
            self._legend.move(w - self._legend.width() - 12, 10)
        else:
            self._legend.move(10, 34)

    def _draw_curve(self):
        n = len(self._hist)
        if n == 0:
            self._curve.setData([]); self._dd_curve.setData([]); self._cmp_curve.setData([]); self._cmp_curve.hide()
            self._rank_curve.setData([]); self._rank_curve.hide()
            self._hide_dd_markers(); self._hide_buy_markers(); self._legend.hide(); return
        self._vline.hide(); self._hline.hide(); self._dot.setData([])
        self._readout.setText("移动鼠标看每日数值")
        xs = np.arange(n)
        span_days = (self._hist[-1][0] - self._hist[0][0]) / 86400000.0
        fmt = "%Y-%m" if span_days > 400 else "%m-%d"
        self._date_axis.set_data(self._hist, fmt)
        self.plot.setXRange(-0.5, n - 1 + 0.5, padding=0)
        aligned = self._aligned_for_cmp() if self._cmp_code else []
        self._cmp_on = bool(self._cmp_code) and len(aligned) >= 2
        if self._cmp_code and not self._cmp_on and self._idx_cache.get(self._cmp_code):
            self._cmp_hint.setText(f"⚠ {self._cmp_name}与本基金无重叠交易日")
        if self._view == "rank":
            self._curve.hide(); self._dd_curve.hide(); self._cmp_curve.hide(); self._legend.hide()
            self._hide_dd_markers(); self._hide_buy_markers()
            self.plot.getAxis("left").setLabel("同类排名（名）")
            self._fund_y_by_idx = {}
            if not self._rank_full:
                self._rank_curve.setData([]); self._rank_curve.hide()
                self.plot.getAxis("left").setLabel("同类排名（暂无数据）")
                return
            rx = np.array([p[0] for p in self._rank_full], dtype=float)
            ry = np.array([p[1] for p in self._rank_full], dtype=float)
            self._rank_curve.show(); self._rank_curve.setData(rx, ry)
            self._fund_y_by_idx = {int(p[0]): int(p[1]) for p in self._rank_full}
            rmin = float(ry.min()); rmax = float(ry.max())
            rpad = max((rmax - rmin) * 0.12, 1.0)
            self.plot.getPlotItem().invertY(True)
            self.plot.setYRange(rmin - rpad, rmax + rpad, padding=0)
            return
        self.plot.getPlotItem().invertY(False)
        self._rank_curve.hide()
        if self._view == "nav":
            self._dd_curve.hide(); self._curve.show(); self._hide_dd_markers()
            if not self._cmp_on:
                self._cmp_curve.hide(); self._legend.hide()
                self.plot.getAxis("left").setLabel("单位净值")
                ys = np.array([p[1] for p in self._hist], dtype=float); self._curve.setData(xs, ys)
                ymin = float(ys.min()); ymax = float(ys.max())
                cost = float(self._my_rec2.get("cost") or 0)
                if cost > 0: ymin = min(ymin, cost); ymax = max(ymax, cost)
                ypad = max((ymax - ymin) * 0.12, 0.005)
                self.plot.setYRange(ymin - ypad, ymax + ypad, padding=0)
                self._fund_y_by_idx = {i: float(self._hist[i][1]) for i in range(n)}
            else:
                self._cmp_curve.show()
                self.plot.getAxis("left").setLabel("区间涨跌幅 %")
                base_nav = aligned[0][1]; base_close = aligned[0][2]
                ax=[]; fp=[]; ip=[]; self._fund_y_by_idx={}
                for (i, nav, close) in aligned:
                    fpc = (nav/base_nav - 1)*100; ipc = (close/base_close - 1)*100
                    ax.append(i); fp.append(fpc); ip.append(ipc); self._fund_y_by_idx[i] = fpc
                self._curve.setData(ax, fp); self._cmp_curve.setData(ax, ip)
                allo = fp + ip; ymin = min(allo); ymax = max(allo); ypad = max((ymax-ymin)*0.12, 0.5)
                self.plot.setYRange(ymin - ypad, ymax + ypad, padding=0)
                self._legend.setText(f'<span style="color:#2563eb">━</span> 本基金 &nbsp; <span style="color:{TEAL}">━</span> {self._cmp_name}（区间涨跌%）')
                self._draw_buy_markers()
        else:
            self._curve.hide(); self._dd_curve.show()
            self.plot.getAxis("left").setLabel("回撤 %")
            ys = np.array(self._dd, dtype=float); self._dd_curve.setData(xs, ys)
            ymin = float(ys.min()) if len(ys) else -1.0
            self._fund_y_by_idx = {i: float(self._dd[i]) for i in range(n)}
            if self._cmp_on:
                self._cmp_curve.show()
                peak = -1e18; ix=[]; ipct=[]
                for (i, nav, close) in aligned:
                    if close > peak: peak = close
                    idd = (close-peak)/peak*100.0 if peak > 0 else 0.0
                    ix.append(i); ipct.append(idd)
                self._cmp_curve.setData(ix, ipct)
                if ipct: ymin = min(ymin, min(ipct))
                self._legend.setText(f'<span style="color:#ef4444">━</span> 本基金回撤 &nbsp; <span style="color:{TEAL}">━</span> {self._cmp_name}回撤')
            else:
                self._cmp_curve.hide(); self._legend.hide()
            ymin = min(ymin, -0.5); ypad = max(abs(ymin) * 0.12, 0.3)
            self.plot.setYRange(ymin - ypad, 0.0 + ypad*0.4, padding=0)
            self._show_dd_markers()
            self._draw_buy_markers()
        self._place_legend()

    def _fill_table(self):
        rows = list(reversed(self._hist))[:200]
        dd_rows = list(reversed(self._dd))[:200] if self._dd else []
        rank_rows = list(reversed(self._rank_full))[:200] if self._rank_full else []
        self.table.setUpdatesEnabled(False)
        third = "同类排名" if self._view == "rank" else ("回撤%" if self._view == "dd" else "当日涨跌")
        self.table.setHorizontalHeaderLabels(["日期", "单位净值", third])
        self.table.setRowCount(len(rows))
        rank_by_i = {int(i): int(r) for (i, r) in self._rank_full} if self._rank_full else {}
        for r,(ts,nav,eqt) in enumerate(rows):
            self.table.setItem(r,0,QTableWidgetItem(datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d")))
            self.table.setItem(r,1,QTableWidgetItem(f"{nav:.4f}"))
            if self._view == "rank":
                idx = (len(self._hist) - 1 - r)
                rv = rank_by_i.get(idx)
                it = QTableWidgetItem(f"第 {rv} 名" if rv is not None else "—")
                it.setForeground(QColor("#b45309") if rv is not None else QColor(GRAY))
            elif self._view == "dd":
                v = dd_rows[r] if r < len(dd_rows) else 0.0
                it = QTableWidgetItem(f"{v:.2f}%")
                it.setForeground(QColor(GRAY) if v > -0.005 else QColor(RED))
            else:
                it = QTableWidgetItem(f"{eqt:+.2f}%"); it.setForeground(QColor(RED) if eqt>=0 else QColor(GREEN))
            self.table.setItem(r,2,it)
        self.table.setUpdatesEnabled(True)

    def _update_dd_stats(self):
        if self._view != "dd":
            self.dd_box.hide(); return
        self.dd_box.show()
        t = T()  # v2.2.0：回撤条文字改主题令牌（原硬编码色暗底失真）
        st = self._dd_state
        if st == "none" or not self._hist:
            self.lbl_dd_max.setText("最大回撤  暂无明显回撤"); self.lbl_dd_max.setStyleSheet(f"color:{t['down']};font-weight:bold;")
            self.lbl_dd_rep.setText("修复  —"); self.lbl_dd_rep.setStyleSheet(f"color:{t['muted']};font-weight:bold;")
        else:
            self.lbl_dd_max.setText(f"最大回撤  {abs(self._dd_max):.2f}%"); self.lbl_dd_max.setStyleSheet(f"color:{t['up']};font-weight:bold;")
            if st == "yes":
                self.lbl_dd_rep.setText(f"修复  已修复 {self._dd_days} 天"); self.lbl_dd_rep.setStyleSheet(f"color:{t['down']};font-weight:bold;")
            elif st == "fixing":
                self.lbl_dd_rep.setText(f"修复  正在修复中 · 已填{self._dd_progress*100:.0f}%"); self.lbl_dd_rep.setStyleSheet(f"color:{t['mid_val']};font-weight:bold;")
            else:
                self.lbl_dd_rep.setText(f"修复  暂未修复 · 已填{self._dd_progress*100:.0f}%"); self.lbl_dd_rep.setStyleSheet(f"color:{t['muted']};font-weight:bold;")

    def _reset_summary(self):
        """v2.2.0：摘要区复位为占位态。"""
        t = T()
        self.d_nav_val.setText("—"); self.d_nav_val.setStyleSheet(f"color:{t['faint']};")
        self.d_mv_val.setText("—"); self.d_mv_val.setStyleSheet(f"color:{t['faint']};")
        self.d_pnl_val.setText("—"); self.d_pnl_val.setStyleSheet(f"color:{t['faint']};")
        self.d_sub.setText("")

    def _update_my_pnl(self):
        """v2.2.0：摘要大字区（净值/市值/累计盈亏）+副行（份额·成本·今日·截至）。"""
        t = T()
        sh = float(self._my_rec2.get("shares") or 0); cost = float(self._my_rec2.get("cost") or 0)
        if not self._hist:
            self._reset_summary(); return
        nav = self._hist[-1][1]
        nd = datetime.fromtimestamp(self._hist[-1][0]/1000).strftime("%m-%d")
        self.d_nav_val.setText(f"{nav:.4f}"); self.d_nav_val.setStyleSheet(f"color:{t['text']};")
        if not sh:
            self.d_mv_val.setText("未填持仓"); self.d_mv_val.setStyleSheet(f"color:{t['faint']};font-size:10px;")
            self.d_pnl_val.setText("未填持仓"); self.d_pnl_val.setStyleSheet(f"color:{t['faint']};font-size:10px;")
            self.d_sub.setText(f"未填持仓（点顶栏「管理持仓」填写） · 截至{nd}"); return
        mv = nav * sh
        self.d_mv_val.setText(f"¥{mv:,.2f}"); self.d_mv_val.setStyleSheet(f"color:{t['text']};")
        parts = [f"份额 {sh:,.2f}"]
        if cost > 0: parts.append(f"成本 {cost:.4f}")
        _chg = getattr(self, "_chg", None)
        if _chg is not None and (100 + _chg) != 0:
            tp = sh * nav * _chg / (100 + _chg)
            parts.append(f"今日 {tp:+,.2f}元")
        parts.append(f"截至{nd}")
        self.d_sub.setText(" · ".join(parts))
        if cost <= 0:
            self.d_pnl_val.setText("未填成本"); self.d_pnl_val.setStyleSheet(f"color:{t['faint']};font-size:10px;"); return
        pnl = (nav-cost)*sh; pct = (nav-cost)/cost*100
        self.d_pnl_val.setText(f"{pnl:+,.2f}元 ({pct:+.2f}%)")
        self.d_pnl_val.setStyleSheet(f"color:{RED if pnl>=0 else GREEN};")

    def _mouse_moved(self, pos):
        if not self._hist: return
        vb = self.plot.getPlotItem().vb
        if not self.plot.sceneBoundingRect().contains(pos): return
        mp = vb.mapSceneToView(pos); idx = int(round(mp.x()))
        if idx < 0 or idx >= len(self._hist): return
        ts,nav,eqt = self._hist[idx]
        self._vline.show(); self._hline.show()
        self._vline.setPos(idx)
        d = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d")
        if self._view == "rank":
            rv = self._fund_y_by_idx.get(idx)
            if rv is None:
                self._vline.hide(); self._hline.hide(); self._dot.setData([])
                self._readout.setText(f"{d} 净值 {nav:.4f} 排名 暂无")
            else:
                self._hline.setPos(rv); self._dot.setBrush(pg.mkBrush("#f59e0b")); self._dot.setData([idx],[rv])
                self._readout.setText(f"{d} 净值 {nav:.4f} 同类排名 第 {rv} 名")
        elif self._view == "nav":
            yv = self._fund_y_by_idx.get(idx, nav)
            self._hline.setPos(yv); self._dot.setBrush(pg.mkBrush("#2563eb")); self._dot.setData([idx],[yv])
            if self._cmp_on:
                self._readout.setText(f"{d} 本基金 {yv:+.2f}% 净值 {nav:.4f}")
            else:
                self._readout.setText(f"{d} 净值 {nav:.4f} 涨跌 {eqt:+.2f}%")
        else:
            dv = self._dd[idx] if idx < len(self._dd) else 0.0
            self._hline.setPos(dv); self._dot.setBrush(pg.mkBrush("#ef4444")); self._dot.setData([idx],[dv])
            self._readout.setText(f"{d} 净值 {nav:.4f} 回撤 {dv:.2f}%")
        self._readout.adjustSize(); self._readout.raise_()


def _ro(text):
    it = QTableWidgetItem(text); it.setFlags(it.flags() & ~Qt.ItemIsEditable); return it


class PasteDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent); self.setWindowTitle("粘贴持仓文字"); self.resize(560,460)
        lay = QVBoxLayout(self)
        _p_lbl = QLabel("把蚂蚁/天天基金里『持有份额、持仓成本价』那段【纯文本】整段复制粘贴到下面，点解析（不是csv/表格）：\n（注：『持有金额/市值』不会被导入——本金列需另行填写或留空。）\n⚠ 只认下面表里已有的基金（即首页看板上的基金）。列表外的新基金粘了会被悄悄跳过——要加新基金，请先回首页用『快速添加』，加完再来粘。")
        _p_lbl.setStyleSheet(f"color:{T()['text_sub']};"); lay.addWidget(_p_lbl)
        self.te = QTextEdit(); self.te.setPlaceholderText("例如：\n某某指数基金C\n持有份额 1234.56\n持仓成本价 1.0000\n……（多只一起粘也行；只认份额+成本价）")
        self.te.setFont(QFont(FONT,9)); lay.addWidget(self.te,3)
        self.lbl = QLabel(""); self.lbl.setStyleSheet(f"color:{T()['text_sub']};"); self.lbl.setWordWrap(True); lay.addWidget(self.lbl)
        bar = QHBoxLayout(); bar.addStretch()
        b = QPushButton("解析"); b.clicked.connect(self._parse)
        b.setStyleSheet(primary_btn_qss())
        bar.addWidget(b); lay.addLayout(bar)
        self.parsed = {}
    def _parse(self):
        try:
            self.parsed = parse_holdings_text(self.te.toPlainText())
        except Exception as e:
            self.parsed = {}; self.lbl.setText(f"⚠ 解析出错：{e}"); return
        n = len(self.parsed)
        if n == 0:
            self.lbl.setText("⚠ 没认出任何一只。请确认文字里含基金全名 + 『持有份额/持仓成本价』字样。认不出的，回到表格手填即可。")
        else:
            names = "、".join(NAME_MAP.get(c, c) for c in self.parsed)
            self.lbl.setText(f"✅ 认出 {n} 只：{names}（仅份额+成本价）。点『采用』填进表格（橙底待你核对），或直接关闭。")
        self.accept()


class HoldDialog(QDialog):
    def __init__(self, resolved, corrected_codes, price_map, parent=None, account=None):
        super().__init__(parent); self.setWindowTitle("管理我的持仓"); self.resize(760,560)
        self._price = price_map; self._busy = True
        self._accounts = load_accounts()
        self._current_account = account or (self._accounts[0] if self._accounts else DEFAULT_ACCOUNT)
        t = T()
        lay = QVBoxLayout(self)
        # —— 账户选择器 ——
        acc_bar = QHBoxLayout()
        _acc_lbl = QLabel("账户："); _acc_lbl.setStyleSheet(f"color:{t['muted']};"); acc_bar.addWidget(_acc_lbl)
        self._acc_combo = QComboBox()
        for a in self._accounts:
            self._acc_combo.addItem(a)
        idx = self._acc_combo.findText(self._current_account)
        if idx >= 0: self._acc_combo.setCurrentIndex(idx)
        self._acc_combo.currentTextChanged.connect(self._switch_account)
        self._acc_combo.setStyleSheet(combo_qss(t)); acc_bar.addWidget(self._acc_combo)
        self._btn_manage_acc = QPushButton("管理账户")
        self._btn_manage_acc.setStyleSheet(ghost_btn_qss(t, pad="4px 10px"))
        self._btn_manage_acc.clicked.connect(self._manage_accounts)
        acc_bar.addWidget(self._btn_manage_acc)
        acc_bar.addStretch()
        lay.addLayout(acc_bar)
        tip = QLabel("  • 「持仓成本价/每份」填每份成本的小数（如 1.0000），不要填持有金额。\n"
                     "  • 「投入本金」列可选：留空即可；要填就填≈份额×成本价(如1000×1.2000≈1200.00)，\n"
                     "    不要填『持有金额/市值』（那是市值=本金+收益，不是本金）。\n"
                     "  • 「买入日期」选填，格式 2026-07-15；填写后详情页画📍首笔买入竖线+🟣成本横线，记不清可留空。\n"
                     "  • 快捷方式：「📋 粘贴导入」只自动填份额+成本价，本金/日期不改动。\n"
                     "  黄底行 = 程序判定填写的是本金，已自动 ÷份额 换算为每份成本。")
        tip.setStyleSheet(panel_label_qss("tip")); lay.addWidget(tip)
        self.lbl_fix = QLabel(""); self.lbl_fix.setStyleSheet(panel_label_qss("warn"))
        self.lbl_fix.setWordWrap(True); self.lbl_fix.hide(); lay.addWidget(self.lbl_fix)
        self.lbl_x = QLabel(""); self.lbl_x.setStyleSheet(panel_label_qss("ok"))
        self.lbl_x.setWordWrap(True); self.lbl_x.hide(); lay.addWidget(self.lbl_x)
        raw_hold = load_holdings_for_account(self._current_account)
        self.table = QTableWidget(len(FUNDS),6)
        self.table.setHorizontalHeaderLabels(["基金名称","代码","持有份额","持仓成本价/每份","投入本金(元,可选)","买入日期(选填)"])
        self.table.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False); self.table.setAlternatingRowColors(True)
        for r,(code,name) in enumerate(FUNDS):
            r2 = resolved.get(code, {})
            sh = r2.get("shares", ""); cost = r2.get("cost", ""); prin = r2.get("principal", "")
            bd = raw_hold.get(code, {}).get("buy_date", "") or ""
            self.table.setItem(r,0,_ro(name)); self.table.setItem(r,1,_ro(code))
            self.table.setItem(r,2,QTableWidgetItem(self._fmt(sh)));
            self.table.setItem(r,3,QTableWidgetItem(self._fmt(cost,4)));
            self.table.setItem(r,4,QTableWidgetItem(self._fmt(prin,2)));
            self.table.setItem(r,5,QTableWidgetItem(bd.strip()))
            if code in corrected_codes:
                for cc in range(6):
                    it = self.table.item(r,cc)
                    if it: it.setBackground(hl_bg())
        self.table.setStyleSheet(table_qss())
        lay.addWidget(self.table)
        if corrected_codes:
            self.lbl_fix.setText(f"⚠ 已自动换算 {len(corrected_codes)} 只：『成本价』列中填写的实为『本金』，已 ÷份额 换算为『每份成本』(黄底行)。核对无误后点保存即永久修正。")
            self.lbl_fix.show()
        self.table.itemChanged.connect(self._on_cell); self._busy = False
        bar = QHBoxLayout()
        b_paste = QPushButton("粘贴填份额·成本"); b_paste.setToolTip("把支付宝/天天里『持有份额·持仓成本价』那段纯文本粘进来，一次灌进表。\n只填数字、不加新基金；新基金请先回首页『快速添加』。")
        b_paste.clicked.connect(self._paste_import)
        b_paste.setStyleSheet(soft_btn_qss(t)); bar.addWidget(b_paste)
        b_x = QPushButton("对账"); b_x.setToolTip("用交易流水逐笔回放，算出每只基金『应有』份额/本金，与手填值比对。\n差异行用橙底填入流水算出的值（待你核对），点保存才落盘；\n手填比流水多的基金会提示去补录流水。")
        b_x.clicked.connect(self._cross_check)
        b_x.setStyleSheet(soft_btn_qss(t)); bar.addWidget(b_x)
        b_move = QPushButton("转移到其他账户"); b_move.setToolTip("把选中的基金持仓从当前账户转移到另一个账户。\n份额/成本/本金原样搬过去，不产生交易流水。")
        b_move.clicked.connect(self._move_to_account)
        b_move.setStyleSheet(soft_btn_qss(t))
        bar.addWidget(b_move); bar.addStretch()
        b_trade = QPushButton("记一笔交易"); b_trade.clicked.connect(self._record_trade)
        b_trade.setStyleSheet(soft_btn_qss(t))
        bar.addWidget(b_trade)
        b_clear = QPushButton("清空全部"); b_clear.clicked.connect(self._clear)
        b_clear.setStyleSheet(danger_btn_qss(t))
        b_cancel = QPushButton("取消"); b_cancel.clicked.connect(self.reject)
        b_cancel.setStyleSheet(ghost_btn_qss(t))
        b_save = QPushButton("保存"); b_save.clicked.connect(self._save)
        b_save.setStyleSheet(primary_btn_qss(t))
        bar.addWidget(b_clear); bar.addWidget(b_cancel); bar.addWidget(b_save); lay.addLayout(bar)
        self.saved = False

    def _paste_import(self):
        dlg = PasteDialog(self); dlg.exec()
        if not dlg.parsed:
            return
        self._busy = True
        try:
            code2row = {code: r for r,(code,name) in enumerate(FUNDS)}
            for code, rec in dlg.parsed.items():
                r = code2row.get(code)
                if r is None:
                    continue
                if "shares" in rec:
                    self.table.setItem(r,2,QTableWidgetItem(self._fmt(rec["shares"])))
                if "cost" in rec:
                    self.table.setItem(r,3,QTableWidgetItem(self._fmt(rec["cost"],4)))
                for cc in (0,1,2,3):
                    it = self.table.item(r,cc)
                    if it: it.setBackground(imp_bg())
        finally:
            self._busy = False

    def _cross_check(self):
        trades = load_trades()
        have = {t["code"] for t in trades}
        if not have:
            QMessageBox.information(self, "对账", "还没有交易流水。\n先在「📒 交易记录」补录，或导入支付宝交易截图的流水，再来对账。")
            return
        self.lbl_x.setText("⏳ 对账中：抓取历史净值…"); self.lbl_x.show()
        QApplication.processEvents()
        same = filled = 0; more_manual = []; failed = []
        self._busy = True
        try:
            for r, (code, name) in enumerate(FUNDS):
                if code not in have: continue
                try:
                    m = _fetch_hist_nav_map(code)
                except Exception:
                    failed.append(name); continue
                d = replay_trades(code, m, trades, self._current_account)
                cur_sh = self._num(r, 2)
                diff = d["shares"] - cur_sh
                tol = max(0.05, d["shares"] * 0.005, cur_sh * 0.005)
                if abs(diff) <= tol:
                    same += 1; continue
                if d["shares"] > cur_sh:
                    self.table.setItem(r, 2, QTableWidgetItem(self._fmt(d["shares"], 4)))
                    if d["cost"] > 0: self.table.setItem(r, 3, QTableWidgetItem(self._fmt(d["cost"], 4)))
                    if d["principal"] > 0: self.table.setItem(r, 4, QTableWidgetItem(self._fmt(d["principal"], 2)))
                    for cc in (2, 3, 4):
                        it = self.table.item(r, cc)
                        if it: it.setBackground(imp_bg())
                    filled += 1
                else:
                    more_manual.append(name)
        finally:
            self._busy = False
        parts = [f"✅ 对账完成：{same} 只一致"]
        if filled: parts.append(f"{filled} 只按流水校准（橙底待核对，点💾保存才落盘；流水估算与支付宝『确认份额』或有小误差，精确以支付宝为准）")
        if more_manual: parts.append(f"⚠ 手填比流水多：{'、'.join(more_manual)}——可能流水有漏录，去「📒 交易记录」补录")
        if failed: parts.append(f"⚠ 没抓到净值跳过：{'、'.join(failed)}")
        self.lbl_x.setText("；".join(parts) + "。")
        self.lbl_x.show()

    def _move_to_account(self):
        """把选中基金的持仓从当前账户转移到另一个账户（立即写盘）。"""
        rows = self.table.selectionModel().selectedRows()
        if not rows:
            cur = self.table.currentRow()
            if cur < 0:
                QMessageBox.information(self, "转移", "请先在表里点一下要转移的那只基金所在行。")
                return
            r = cur
        else:
            r = rows[0].row()
        code = self.table.item(r, 1).text().strip()
        name = self.table.item(r, 0).text().strip()
        # 检查当前账户确实有该基金的持仓（磁盘上）
        cur_holdings = load_holdings_for_account(self._current_account)
        rec = cur_holdings.get(code)
        if not rec or not ((rec.get("shares") or 0) > 0 or (rec.get("principal") or 0) > 0):
            QMessageBox.information(self, "转移", f"「{name}」在账户『{self._current_account}』里没有已保存的持仓，无需转移。")
            return
        # 选择目标账户
        others = [a for a in load_accounts() if a != self._current_account]
        if not others:
            QMessageBox.information(self, "转移", "只有一个账户，无法转移。\n请先点「⚙ 管理账户」新建一个账户。")
            return
        target, ok = QInputDialog.getItem(self, "转移到哪个账户",
                                          f"把「{name}」从『{self._current_account}』转移到：", others, 0, False)
        if not ok or not target:
            return
        # 检查目标账户是否已有同基金
        nested = load_holdings_nested()
        target_holdings = nested.get(target, {})
        if code in target_holdings and ((target_holdings[code].get("shares") or 0) > 0):
            ans = QMessageBox.question(self, "目标已有该基金",
                f"账户『{target}』里已经有「{name}」的持仓。\n转移将合并两边份额/本金，是否继续？",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans != QMessageBox.Yes:
                return
            # 合并
            t_rec = target_holdings[code]
            m_sh = float(t_rec.get("shares") or 0) + float(rec.get("shares") or 0)
            m_prin = float(t_rec.get("principal") or 0) + float(rec.get("principal") or 0)
            bd_t = (t_rec.get("buy_date") or "").strip(); bd_r = (rec.get("buy_date") or "").strip()
            m_bd = min(x for x in [bd_t, bd_r] if x) if (bd_t or bd_r) else ""
            target_holdings[code] = {"shares": round(m_sh, 4),
                                     "cost": round(m_prin / m_sh, 4) if m_sh > 0 else 0.0,
                                     "principal": round(m_prin, 2), "buy_date": m_bd}
        else:
            target_holdings[code] = rec
        nested[target] = target_holdings
        # 从源账户删除
        nested[self._current_account].pop(code, None)
        save_holdings_nested(nested)
        # 同步交易流水的账户标记
        all_t = load_trades(); changed = False
        for t in all_t:
            if t.get("code") == code and (t.get("account") or DEFAULT_ACCOUNT) == self._current_account:
                t["account"] = target; changed = True
        if changed:
            save_trades(all_t)
        # 清空当前表格该行（避免后续保存写回）
        self._busy = True
        try:
            for cc in (2, 3, 4, 5):
                self.table.setItem(r, cc, QTableWidgetItem(""))
        finally:
            self._busy = False
        QMessageBox.information(self, "转移完成",
            f"✅ 「{name}」已从『{self._current_account}』转移到『{target}』。\n"
            f"交易流水已同步归属新账户。\n注意：转移立即生效，其余单元格的未保存修改仍需点「💾保存」。")

    def _record_trade(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows:
            cur = self.table.currentRow()
            if cur < 0:
                QMessageBox.information(self, "记一笔", "请先在表里点一下要记账的那只基金所在行。")
                return
            r = cur
        else:
            r = rows[0].row()
        code = self.table.item(r, 1).text().strip()
        name = self.table.item(r, 0).text().strip()
        dlg = TradeDialog(self, name, code, "buy")
        # 预设当前账户
        idx = dlg.cb_account.findText(self._current_account)
        if idx >= 0: dlg.cb_account.setCurrentIndex(idx)
        if dlg.exec() != QDialog.Accepted or not dlg.result:
            return
        res = dlg.result()
        act, nav = res["kind"], res["price"]
        amount, shares, date = res["amount"], res["share"], res["date"]
        trade_account = res.get("account", self._current_account)
        base_sh   = self._num(r, 2)
        base_cost = self._num(r, 3)
        base_prin = self._num(r, 4)
        base_bd   = (self.table.item(r, 5).text().strip() if self.table.item(r, 5) else "")
        cur = {"shares": base_sh, "cost": base_cost, "principal": base_prin}
        if base_bd:
            cur["buy_date"] = base_bd
        draft = {code: cur}
        realized = [None]; cashflow = [None]
        snap_to = None; to_code = None; src_sh = 0.0; conv_amt = 0.0; to_share = 0.0
        if act == "convert":
            src_sh = shares if shares > 0 else (amount / nav if nav > 0 else 0)
            if src_sh <= 0:
                QMessageBox.warning(self, "转换", "转换请填【份额】（转出份额），或填【金额】按 金额÷净值 反推。")
                return
            conv_amt = round(src_sh * nav, 2)
            to_code = res.get("to_code"); to_share = res.get("to_share", 0)
            if not to_code or to_code == code:
                QMessageBox.warning(self, "转换", "请选择转入基金。")
                return
            try:
                tm = _fetch_hist_nav_map(to_code)
            except Exception:
                tm = {}
            to_nav = tm.get(date)
            if (not to_nav) and to_share > 0:
                to_nav = conv_amt / to_share
            if not to_nav or to_nav <= 0:
                QMessageBox.warning(self, "转换", f"没抓到转入基金 {date} 的净值，请直接填【转入份额】（支付宝交易详情里有）。")
                return
            if to_share <= 0: to_share = conv_amt / to_nav
        if act == "dividend_reinvest" and nav <= 0 and shares > 0 and amount > 0:
            nav = amount / shares
        try:
            if act == "convert":
                snap = apply_trade(draft, code, "convert", nav=nav, shares=src_sh, date=date,
                                   realized_out=realized, cashflow_out=cashflow, account=trade_account)
                snap_to = apply_trade(draft, to_code, "buy", nav=to_nav, amount=conv_amt, date=date,
                                      account=trade_account)
            else:
                snap = apply_trade(draft, code, act, nav=nav, amount=amount,
                                   shares=shares, date=date,
                                   realized_out=realized, cashflow_out=cashflow, account=trade_account)
        except Exception as e:
            QMessageBox.warning(self, "记账失败", f"{e}")
            return
        self._busy = True
        try:
            self.table.setItem(r, 2, QTableWidgetItem(self._fmt(snap.get("shares", 0))))
            self.table.setItem(r, 3, QTableWidgetItem(self._fmt(snap.get("cost", 0), 4)))
            self.table.setItem(r, 4, QTableWidgetItem(self._fmt(snap.get("principal", 0), 2)))
            bd = (snap.get("buy_date") or "").strip()
            self.table.setItem(r, 5, QTableWidgetItem(bd))
            if snap_to is not None:
                rr = next((i for i in range(self.table.rowCount())
                           if self.table.item(i, 1) and self.table.item(i, 1).text().strip() == to_code), None)
                if rr is not None:
                    self.table.setItem(rr, 2, QTableWidgetItem(self._fmt(snap_to.get("shares", 0))))
                    self.table.setItem(rr, 3, QTableWidgetItem(self._fmt(snap_to.get("cost", 0), 4)))
                    self.table.setItem(rr, 4, QTableWidgetItem(self._fmt(snap_to.get("principal", 0), 2)))
        finally:
            self._busy = False
        if act == "convert":
            to_name = NAME_MAP.get(to_code, to_code)
            QMessageBox.information(self, "已记入草稿",
                f"✅ 转换已记入本窗草稿（尚未存盘）：{name} 转出 {src_sh:.4f} 份（≈{conv_amt:.2f}元）→ {to_name} 转入 {to_share:.4f} 份。\n"
                f"⚠️ 点右下角『💾保存』才会写入文件；直接关窗＝放弃本笔。")
            return
        extra = ""
        if realized[0] is not None: extra = f" 已实现盈亏 {realized[0]:+.4f}"
        if cashflow[0] is not None: extra = f" 现金流入 {cashflow[0]:+.2f}"
        QMessageBox.information(self, "已记入草稿",
            f"✅ {name}({code}) 已记入本窗草稿（尚未存盘）。\n"
            f"现 份额={snap.get('shares',0):.4f} 成本={snap.get('cost',0):.4f} 本金={snap.get('principal',0):.2f}{extra}\n"
            f"⚠️ 点右下角『💾保存』才会写入文件；直接关窗＝放弃本笔。")

    def _fmt(self, v, nd=0):
        if v in ("", None): return ""
        try: f = float(v)
        except Exception: return str(v)
        return (f"{f:.{nd}f}" if nd else (str(int(f)) if f == int(f) else str(f)))

    def _num(self, r, c):
        it = self.table.item(r,c)
        if not it: return 0.0
        try: return float(it.text().strip())
        except Exception: return 0.0

    def _on_cell(self, item):
        if self._busy: return
        r = item.row(); c = item.column()
        if c not in (2,3,4): return
        self._busy = True
        try:
            sh = self._num(r,2); cost = self._num(r,3); prin = self._num(r,4)
            if c == 3 and sh > 0:
                self.table.setItem(r,4,QTableWidgetItem(f"{cost*sh:.2f}" if cost else ""))
            elif c == 4 and sh > 0:
                self.table.setItem(r,3,QTableWidgetItem(f"{prin/sh:.4f}" if prin else ""))
        finally:
            self._busy = False

    def _clear(self):
        self._busy = True
        for r in range(self.table.rowCount()):
            self.table.setItem(r,2,QTableWidgetItem("")); self.table.setItem(r,3,QTableWidgetItem("")); self.table.setItem(r,4,QTableWidgetItem("")); self.table.setItem(r,5,QTableWidgetItem(""))
        self._busy = False

    def _switch_account(self, name):
        if self._busy or not name:
            return
        self._busy = True
        try:
            self._current_account = name
            self._load_for_account()
        finally:
            self._busy = False

    def _load_for_account(self):
        """根据当前账户重新加载表格数据。"""
        holdings = load_holdings_for_account(self._current_account)
        price_map = self._price
        resolved, corrected = resolve_holdings(holdings, price_map)
        self._busy = True
        try:
            for r, (code, name) in enumerate(FUNDS):
                r2 = resolved.get(code, {})
                sh = r2.get("shares", ""); cost = r2.get("cost", ""); prin = r2.get("principal", "")
                bd = holdings.get(code, {}).get("buy_date", "") or ""
                self.table.setItem(r,2,QTableWidgetItem(self._fmt(sh)))
                self.table.setItem(r,3,QTableWidgetItem(self._fmt(cost,4)))
                self.table.setItem(r,4,QTableWidgetItem(self._fmt(prin,2)))
                self.table.setItem(r,5,QTableWidgetItem(bd.strip()))
                for cc in range(6):
                    it = self.table.item(r,cc)
                    if it:
                        it.setBackground(hl_bg() if code in corrected else QColor(255,255,255,0))
        finally:
            self._busy = False

    def _manage_accounts(self):
        dlg = QDialog(self); dlg.setWindowTitle("管理账户"); dlg.resize(400, 300)
        lay = QVBoxLayout(dlg)
        _m_lbl = QLabel("每个账户对应一个平台（如支付宝、天天基金、招行等）。")
        _m_lbl.setStyleSheet(f"color:{T()['muted']};"); lay.addWidget(_m_lbl)
        list_box = QVBoxLayout(); lay.addLayout(list_box)
        def rebuild():
            while list_box.count():
                it = list_box.takeAt(0); w = it.widget()
                if w: w.deleteLater()
            for a in load_accounts():
                # 每行套一个 QWidget：否则旧行 takeAt 时 widget() 为 None，
                # 旧行控件删不掉变幽灵行（v1.1 修复：重复显示/删后残留）
                roww = QWidget()
                row = QHBoxLayout(roww); row.setContentsMargins(0,2,0,2)
                lb = QLabel(a); lb.setFont(QFont(FONT, 10)); lb.setStyleSheet(f"color:{T()['text']};")
                row.addWidget(lb, 1)
                if a != DEFAULT_ACCOUNT:
                    btn_rename = QPushButton("改名")
                    btn_rename.setStyleSheet(ghost_btn_qss(T(), pad="3px 10px"))
                    btn_rename.clicked.connect(lambda _, x=a: do_rename(x))
                    btn_del = QPushButton("删除")
                    btn_del.setStyleSheet(danger_btn_qss(T(), pad="3px 10px"))
                    btn_del.clicked.connect(lambda _, x=a: do_delete(x))
                    row.addWidget(btn_rename); row.addWidget(btn_del)
                list_box.addWidget(roww)
        def do_rename(old):
            new, ok = QInputDialog.getText(dlg, "改名", "新账户名：", text=old)
            if not ok or not new.strip() or new.strip() == old:
                return
            new = new.strip()
            if new in load_accounts():
                QMessageBox.warning(dlg, "重复", f"账户「{new}」已存在。"); return
            accs = load_accounts()
            accs[accs.index(old)] = new
            save_accounts(accs)
            nested = load_holdings_nested()
            if old in nested:
                nested[new] = nested.pop(old)
                save_holdings_nested(nested)
            all_t = load_trades()
            for t in all_t:
                if t.get("account") == old:
                    t["account"] = new
            save_trades(all_t)
            self._current_account = new
            rebuild()
        def do_delete(name):
            h = load_holdings_for_account(name)
            has_data = any((v.get("shares") or 0) > 0 or (v.get("principal") or 0) > 0 for v in h.values() if isinstance(v, dict))
            if has_data:
                QMessageBox.warning(dlg, "无法删除", f"账户「{name}」还有持仓数据，请先清空。"); return
            if QMessageBox.question(dlg, "删除账户", f"确定删除账户「{name}」？") != QMessageBox.Yes:
                return
            accs = load_accounts(); accs.remove(name); save_accounts(accs)
            nested = load_holdings_nested(); nested.pop(name, None); save_holdings_nested(nested)
            # 删掉的恰好是当前选中账户：切到第一个账户并重载表格（v1.1 修复）
            if self._current_account == name:
                new_cur = accs[0] if accs else DEFAULT_ACCOUNT
                self._acc_combo.blockSignals(True)
                idx = self._acc_combo.findText(new_cur)
                if idx >= 0: self._acc_combo.setCurrentIndex(idx)
                self._acc_combo.blockSignals(False)
                self._current_account = new_cur
                self._load_for_account()
            rebuild()
        add_bar = QHBoxLayout()
        ed_new = QLineEdit(); ed_new.setPlaceholderText("新账户名称"); ed_new.setStyleSheet(input_qss())
        btn_add = QPushButton("添加"); btn_add.setStyleSheet(soft_btn_qss())
        def do_add():
            n = ed_new.text().strip()
            if not n: return
            if n in load_accounts():
                QMessageBox.warning(dlg, "重复", f"账户「{n}」已存在。"); return
            accs = load_accounts(); accs.append(n); save_accounts(accs)
            ed_new.clear(); rebuild()
        btn_add.clicked.connect(do_add)
        add_bar.addWidget(ed_new, 1); add_bar.addWidget(btn_add)
        lay.addLayout(add_bar)
        btn_close = QPushButton("关闭"); btn_close.clicked.connect(dlg.accept)
        btn_close.setStyleSheet(primary_btn_qss())
        lay.addWidget(btn_close)
        rebuild()
        dlg.exec()
        self._accounts = load_accounts()
        self._acc_combo.blockSignals(True)
        self._acc_combo.clear()
        for a in self._accounts:
            self._acc_combo.addItem(a)
        idx = self._acc_combo.findText(self._current_account)
        if idx >= 0:
            self._acc_combo.setCurrentIndex(idx)
        else:
            self._current_account = self._accounts[0] if self._accounts else DEFAULT_ACCOUNT
            self._acc_combo.setCurrentIndex(0)
        self._acc_combo.blockSignals(False)
        self._load_for_account()

    def _save(self):
        d = {}
        for r,(code,name) in enumerate(FUNDS):
            sh = self._num(r,2); cost = self._num(r,3); prin = self._num(r,4)
            bd = (self.table.item(r,5).text().strip() if self.table.item(r,5) else "")
            if bd:
                try:
                    datetime.strptime(bd, "%Y-%m-%d")
                except Exception:
                    QMessageBox.warning(self,"买入日期格式",
                        f"「{name}」的买入日期『{bd}』格式不对。\n请用 YYYY-MM-DD，例如 2026-07-15；记不清就清空这一格。")
                    return
            if sh == 0 and cost == 0 and prin == 0 and not bd: continue
            if cost == 0 and prin > 0 and sh > 0: cost = prin/sh
            if prin == 0 and cost > 0 and sh > 0: prin = cost*sh
            nav = self._price.get(code,0)
            if cost > 0 and nav > 0 and (cost/nav > 3 or cost/nav < 0.33):
                ans = QMessageBox.question(self,"成本价异常",
                    f"「{name}」的每份成本 {cost:.4f} 与现价 {nav:.4f} 相差 {cost/nav:.0f} 倍。\n"
                    f"你是不是把『本金/金额』填进了『每份成本』？\n\n选『否』返回修改；选『是』仍按当前值保存。",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if ans != QMessageBox.Yes: return
            d[code] = {"shares": sh, "cost": round(cost,4), "principal": round(prin,2), "buy_date": bd}
        nested = load_holdings_nested()
        nested[self._current_account] = d
        save_holdings_nested(nested)
        self.saved = True
        self.accept()


# ---- 历史净值缓存：补录记账自动带净值 ----
_HIST_CACHE = {}  # {基金代码: {日期: 单位净值}}

def _fetch_hist_nav_map(code):
    if code in _HIST_CACHE:
        return _HIST_CACHE[code]
    m, errs, raws = {}, [], []
    UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    def get(url, ref=None):
        h = dict(UA)
        if ref:
            h["Referer"] = ref
        with urllib.request.urlopen(
                urllib.request.Request(url, headers=h), timeout=10) as r:
            return r.read().decode("utf-8", "ignore")

    code = (code or "").strip()
    if len(code) != 6 or not code.isdigit():
        raise RuntimeError(f"基金代码无效:[{code}]")

    # 优先全历史净值（lsjz 分页接口可能只返回最近一页，回放/对账会少算）
    try:
        _, hist, _ = fetch_history(code)
        m = {datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d"): nv for ts, nv, _e in hist}
    except Exception as e:
        errs.append(f"全历史:{e}")
    if m:
        _HIST_CACHE[code] = m
        return m

    try:
        for ps in (500, 20):
            txt = get(f"https://api.fund.eastmoney.com/f10/lsjz"
                      f"?fundCode={code}&pageIndex=1&pageSize={ps}",
                      "https://fundf10.eastmoney.com/")
            raws.append(f"东财1:len={len(txt)}")
            data = json.loads(txt)
            for it in ((data.get("Data") or {}).get("LSJZList") or []):
                try:
                    m[str(it.get("FSRQ", "")).strip()] = float(it.get("DWJZ"))
                except Exception:
                    pass
            if m:
                break
    except Exception as e:
        errs.append(f"东财1:{e}")

    if not m:
        try:
            txt = get(f"https://fund.eastmoney.com/pingzhongdata/{code}.js",
                      "https://fund.eastmoney.com/")
            raws.append(f"东财3:len={len(txt)}")
            seg = re.search(r'Data_netWorthTrend\s*=\s*(\[.*?\]);', txt, re.S)
            if seg:
                for x, y in re.findall(r'"x"\s*:\s*(\d+)\s*,\s*"y"\s*:\s*(-?[\d.]+)',
                                       seg.group(1)):
                    ds = datetime.date.fromtimestamp(int(x) / 1000).strftime("%Y-%m-%d")
                    m.setdefault(ds, float(y))
        except Exception as e:
            errs.append(f"东财3:{e}")

    if not m:
        raise RuntimeError(f"code={code}；{'；'.join(errs) or '可达但解析0条'}"
                           f"；{' / '.join(raws)}")
    _HIST_CACHE[code] = m
    return m


class TradeDialog(QDialog):
    """交易录入：买入/卖出 + 日期 + 金额/份额 + 净值（支持补录历史交易）"""
    def __init__(self, parent, name, code, kind):
        super().__init__(parent)
        self.setWindowTitle("记一笔 - " + name)
        self.resize(560, 360)
        self._code = code
        self._price = getattr(parent, "_price", {})
        L = QVBoxLayout(self); L.setSpacing(8)

        tip = QLabel("口径：15:00 前提交按【当天】净值确认，15:00 后/节假日按【下一交易日】净值确认。\n补录历史交易：日期填确认净值对应的交易日，净值/份额照抄支付宝交易详情即可。\n分红：现金分红只记金额（计入当日收益）；红利再投记金额+份额（份额并入持仓）。")
        tip.setWordWrap(True)
        tip.setStyleSheet(f"color:{T()['muted']}; font-size:11px;")
        L.addWidget(tip)

        H = QHBoxLayout()
        self.rb_buy = QRadioButton("买入")
        self.rb_sell = QRadioButton("卖出")
        self.rb_conv = QRadioButton("转换")
        self.rb_divc = QRadioButton("现金分红")
        self.rb_divr = QRadioButton("红利再投")
        (self.rb_buy if kind == "buy" else self.rb_sell).setChecked(True)
        for _rb in (self.rb_buy, self.rb_sell, self.rb_conv, self.rb_divc, self.rb_divr):
            _rb.setStyleSheet(f"color:{T()['text']};"); H.addWidget(_rb)
        H.addStretch()
        L.addLayout(H)

        F = QFormLayout(); F.setLabelAlignment(Qt.AlignRight)
        self.cb_account = QComboBox()
        for a in load_accounts():
            self.cb_account.addItem(a)
        self.cb_account.setStyleSheet(combo_qss())
        F.addRow("账户", self.cb_account)
        self.ed_date = QDateEdit(QDate.currentDate())
        self.ed_date.setCalendarPopup(True)
        self.ed_date.setDisplayFormat("yyyy-MM-dd")
        self.ed_date.setStyleSheet(input_qss())
        F.addRow("交易日期", self.ed_date)
        self.ed_amt = QLineEdit()
        self.ed_amt.setPlaceholderText("买入=花的钱 卖出=到账的钱")
        self.ed_amt.setStyleSheet(input_qss())
        F.addRow("金额(元)", self.ed_amt)
        self.ed_share = QLineEdit()
        self.ed_share.setPlaceholderText("选填：可直接抄支付宝【确认份额】；不填按 金额÷净值 算")
        self.ed_share.setStyleSheet(input_qss())
        F.addRow("份额", self.ed_share)
        self.ed_price = QLineEdit()
        p0 = self._price.get(code, 0)
        if p0 > 0: self.ed_price.setText(f"{p0:.4f}")
        self.ed_price.setStyleSheet(input_qss())
        F.addRow("成交净值", self.ed_price)
        L.addLayout(F)

        self.w_to = QWidget()
        tf = QFormLayout(self.w_to); tf.setContentsMargins(0, 0, 0, 0); tf.setLabelAlignment(Qt.AlignRight)
        self.cb_to = QComboBox()
        for c, n in FUNDS:
            if c != code: self.cb_to.addItem(f"{n}（{c}）", c)
        self.cb_to.setStyleSheet(combo_qss())
        self.ed_to_share = QLineEdit()
        self.ed_to_share.setPlaceholderText("选填：不填按 转出金额÷转入净值 算")
        self.ed_to_share.setStyleSheet(input_qss())
        tf.addRow("转入基金", self.cb_to); tf.addRow("转入份额", self.ed_to_share)
        self.w_to.hide()
        L.addWidget(self.w_to)
        self.rb_conv.toggled.connect(self.w_to.setVisible)

        self.lbl_hint = QLabel("")
        self.lbl_hint.setWordWrap(True)
        self.lbl_hint.setStyleSheet(panel_label_qss("warn"))
        self.lbl_hint.hide()
        L.addWidget(self.lbl_hint)

        BB = QHBoxLayout()
        btn_ok = QPushButton("确定")
        btn_ok.clicked.connect(self._try_accept)
        btn_ok.setStyleSheet(primary_btn_qss())
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_cancel.setStyleSheet(ghost_btn_qss())
        BB.addStretch(); BB.addWidget(btn_ok); BB.addWidget(btn_cancel)
        L.addLayout(BB)

        self.ed_date.dateChanged.connect(self._on_date_changed)

    def _on_date_changed(self, d):
        if d == QDate.currentDate():
            p0 = self._price.get(self._code, 0)
            self.ed_price.setText(f"{p0:.4f}" if p0 > 0 else "")
            self.lbl_hint.hide()
            return
        ds = d.toString("yyyy-MM-dd")
        err = ""
        try:
            m = _fetch_hist_nav_map(self._code)
        except Exception as e:
            m, err = {}, str(e)
        nav, used = m.get(ds), ds
        if not nav:
            nd = QDate(d)
            for _ in range(10):
                nd = nd.addDays(1)
                v = m.get(nd.toString("yyyy-MM-dd"))
                if v:
                    nav, used = v, nd.toString("yyyy-MM-dd")
                    break
        if nav:
            self.ed_price.setText(f"{nav:.4f}")
            if used == ds:
                txt = f"✅ 已带出 {ds} 的确认净值：{nav:.4f}；份额可直接抄支付宝【确认份额】（不符可手改）。"
            else:
                txt = f"✅ {ds} 为非交易日，按规则已带出下一交易日 {used} 的确认净值：{nav:.4f}。"
            self.lbl_hint.setStyleSheet("color:#1b5e20; background:#e8f5e9; padding:4px; border-radius:4px;")
        else:
            self.ed_price.setText("")
            txt = f"⚠ 没抓到净值（{err or '未知原因'}）：请先按支付宝交易详情手填。"
            self.lbl_hint.setStyleSheet("color:#8a6d00; background:#fff8e1; padding:4px; border-radius:4px;")
        self.lbl_hint.setText(txt)
        self.lbl_hint.show()

    def _kind(self):
        if self.rb_conv.isChecked(): return "convert"
        if self.rb_divc.isChecked(): return "dividend_cash"
        if self.rb_divr.isChecked(): return "dividend_reinvest"
        return "buy" if self.rb_buy.isChecked() else "sell"

    def _try_accept(self):
        kind = self._kind()
        try: price = float(self.ed_price.text())
        except ValueError: price = 0
        try: amount = float(self.ed_amt.text() or 0)
        except ValueError:
            QMessageBox.warning(self, "提示", "金额必须是数字。"); return
        try: share = float(self.ed_share.text() or 0)
        except ValueError:
            QMessageBox.warning(self, "提示", "份额必须是数字。"); return
        if amount < 0 or share < 0:
            QMessageBox.warning(self, "提示", "金额/份额不能为负数。"); return
        if kind == "dividend_cash":
            if amount <= 0:
                QMessageBox.warning(self, "提示", "现金分红请填【金额】（到账的分红钱）。"); return
            self.accept(); return
        if kind == "dividend_reinvest":
            if amount <= 0:
                QMessageBox.warning(self, "提示", "红利再投请填【金额】。"); return
            if price <= 0 and share <= 0:
                QMessageBox.warning(self, "提示", "红利再投需要【成交净值】或【份额】之一，用来算再投份额。"); return
            self.accept(); return
        if price <= 0:
            QMessageBox.warning(self, "提示", "成交净值必须大于 0。\n补录历史交易请填该笔的确认净值，不要用今天的净值。"); return
        if amount <= 0 and share <= 0:
            QMessageBox.warning(self, "提示", "【金额】和【份额】至少填一个：买入填金额，卖出可只填份额。"); return
        self.accept()

    def result(self):
        kind = self._kind()
        date = self.ed_date.date().toString("yyyy-MM-dd")
        amount = float(self.ed_amt.text() or 0)
        share = float(self.ed_share.text() or 0)
        price = float(self.ed_price.text() or 0)
        account = self.cb_account.currentText() or DEFAULT_ACCOUNT
        out = {"kind": kind, "date": date, "amount": amount, "share": share, "price": price, "account": account}
        if kind == "convert":
            out["to_code"] = self.cb_to.currentData()
            try: out["to_share"] = float(self.ed_to_share.text() or 0)
            except ValueError: out["to_share"] = 0
        return out

TRADES_FILE = os.path.join(_BASE, "trades.json")

def load_trades():
    d = _load_json_with_bak(TRADES_FILE, [])
    return d if isinstance(d, list) else []

def save_trades(t):
    _atomic_write_json(TRADES_FILE, t, indent=1)

def replay_trades(code, nav_map, trades, account=None):
    """按流水逐笔回放，推导该只“应有”持仓三值(shares/cost/principal)。
       account=None 表示不过滤账户（合并回放）。"""
    return domain_logic.replay_trades(code, nav_map, trades, account, DEFAULT_ACCOUNT)

class PnlDialog(QDialog):
    """收益明细：收益总览 + 收益日历(红涨绿跌) + 当日明细。
       口径：当日盈亏=Σ当日持有份额×(当日净值−前一日净值)；份额按交易记录逐日回溯；
       起点=第一笔交易/最早买入日期(之前不记录)；休市日=0；清仓后=0。"""
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("收益明细"); self.resize(900, 720)
        self._hist = {}; self._days = []; self._sel = None
        self._ym = (datetime.now().year, datetime.now().month)
        self._worker = None; self._account_filter = "__all__"
        t = T()
        L = QVBoxLayout(self); L.setSpacing(8)
        tip = QLabel("口径：当日盈亏 = Σ 当日持有份额 ×（当日净值 − 前一日净值），份额按交易记录逐日回溯。从第一笔交易/买入日期开始记录，之前不计；休市日收益为 0；全部卖出后为 0。无交易记录且未填买入日期的基金，从起点按当前份额近似（仅供参考）。")
        tip.setWordWrap(True); tip.setStyleSheet(f"color:{t['muted']};font-size:10px;"); L.addWidget(tip)
        # 账户筛选
        acc_bar = QHBoxLayout()
        _af_lbl = QLabel("账户筛选："); _af_lbl.setStyleSheet(f"color:{t['muted']};"); acc_bar.addWidget(_af_lbl)
        self._acc_filter = QComboBox()
        self._acc_filter.addItem("全部账户", "__all__")
        for a in load_accounts():
            self._acc_filter.addItem(a, a)
        self._acc_filter.currentIndexChanged.connect(self._on_filter_changed)
        self._acc_filter.setStyleSheet(combo_qss(t)); acc_bar.addWidget(self._acc_filter)
        acc_bar.addStretch()
        L.addLayout(acc_bar)
        ov = QFrame(); ov.setStyleSheet(panel_qss())
        ol = QHBoxLayout(ov); ol.setContentsMargins(14,10,14,10)
        self._ov_lbls = {}
        for key, name in (("yest","昨日收益"), ("month","本月收益"), ("monthpct","本月收益率"), ("year","本年累计")):
            box = QVBoxLayout(); box.setSpacing(2)
            a = QLabel(name); a.setStyleSheet(f"color:{t['muted']};font-size:9px;"); a.setAlignment(Qt.AlignCenter)
            b = QLabel("—"); b.setFont(QFont(FONT,13,QFont.Bold)); b.setAlignment(Qt.AlignCenter); b.setStyleSheet(f"color:{t['text']};")
            box.addWidget(a); box.addWidget(b); ol.addLayout(box)
            self._ov_lbls[key] = b
        L.addWidget(ov)
        nav = QHBoxLayout()
        self.btn_prev = QPushButton("‹"); self.btn_next = QPushButton("›")
        for b in (self.btn_prev, self.btn_next):
            b.setFixedWidth(34); b.setFont(QFont(FONT,11,QFont.Bold))
            b.setStyleSheet(ghost_btn_qss(t, pad="4px 0px"))
        self.btn_prev.clicked.connect(lambda: self._shift(-1)); self.btn_next.clicked.connect(lambda: self._shift(1))
        self.lbl_month = QLabel(""); self.lbl_month.setFont(QFont(FONT,12,QFont.Bold)); self.lbl_month.setAlignment(Qt.AlignCenter)
        self.lbl_month.setStyleSheet(f"color:{t['text']};")
        self.lbl_mtot = QLabel(""); self.lbl_mtot.setFont(QFont(FONT,12,QFont.Bold))  # v2.2.0：当月合计置顶
        nav.addWidget(self.btn_prev); nav.addWidget(self.lbl_month,1); nav.addWidget(self.lbl_mtot); nav.addWidget(self.btn_next); L.addLayout(nav)
        self.grid = QGridLayout(); self.grid.setSpacing(4)
        for c, wt in enumerate(("日","一","二","三","四","五","六")):
            h = QLabel(wt); h.setAlignment(Qt.AlignCenter); h.setStyleSheet(f"color:{t['muted']};font-weight:bold;")
            self.grid.addWidget(h, 0, c)
        self._cells = []
        for r in range(6):
            row = []
            for c in range(7):
                cell = QLabel(""); cell.setAlignment(Qt.AlignCenter)
                cell.setFixedSize(104, 60); cell.setFont(QFont(FONT,10,QFont.Bold))  # v2.2.0：格子加大、数字加粗
                cell.setCursor(QCursor(Qt.PointingHandCursor))
                cell.mousePressEvent = lambda ev, rr=r, cc=c: self._click(rr, cc)
                self.grid.addWidget(cell, r+1, c); row.append(cell)
            self._cells.append(row)
        L.addLayout(self.grid)
        self.lbl_day = QLabel("点日历上某一天，看每只基金当天贡献 ↓")
        self.lbl_day.setFont(QFont(FONT,10,QFont.Bold)); self.lbl_day.setStyleSheet(f"color:{t['text']};")
        L.addWidget(self.lbl_day)
        self.tbl = QTableWidget(0,4); self.tbl.setHorizontalHeaderLabels(["基金","当日盈亏(元)","当日涨跌","备注"])
        self.tbl.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.verticalHeader().setVisible(False)
        self.tbl.setAlternatingRowColors(True); self.tbl.setStyleSheet(table_qss("font-size:12px;"))
        L.addWidget(self.tbl,1)
        self.lbl_status = QLabel(""); self.lbl_status.setStyleSheet(f"color:{t['muted']};font-size:10px;"); L.addWidget(self.lbl_status)
        bb = QHBoxLayout(); bb.addStretch()
        b = QPushButton("关闭"); b.clicked.connect(self.reject)
        b.setStyleSheet(primary_btn_qss(t))
        bb.addWidget(b); L.addLayout(bb)

    def start(self):
        all_trades = load_trades()
        if self._account_filter != "__all__":
            filtered = [t for t in all_trades if (t.get("account") or DEFAULT_ACCOUNT) == self._account_filter]
        else:
            filtered = all_trades
        trades_by = {}
        for t in filtered:
            trades_by.setdefault(t["code"], []).append(t)
        # 现金分红按日合计（v1.1 分红入账），计入日历/总览
        self._div = {}
        for t in filtered:
            if t.get("side") == "dividend_cash" and t.get("date"):
                self._div[t["date"]] = self._div.get(t["date"], 0.0) + float(t.get("amount") or 0)
        # 兑底持仓也按账户筛选（无交易记录的基金用当前份额近似）
        if self._account_filter != "__all__":
            self._hold_fb = load_holdings_for_account(self._account_filter)
        else:
            self._hold_fb = load_holdings()
        hold = self._hold_fb
        codes = []
        for c, _n in FUNDS:
            sh = float(hold.get(c, {}).get("shares") or 0)
            bd = (hold.get(c) or {}).get("buy_date", "")
            if trades_by.get(c) or sh > 0 or bd:
                codes.append(c)
        if not codes:
            self.lbl_status.setText("⚠ 没有可统计的持仓基金。"); return
        dates = [t.get("date", "") for ts in trades_by.values() for t in ts if t.get("date")]
        dates += [(hold.get(c) or {}).get("buy_date", "") for c in codes if (hold.get(c) or {}).get("buy_date")]
        self._start = min(dates) if dates else ""
        self._codes = codes
        self.lbl_status.setText(f"⏳ 抓取历史净值 0/{len(codes)} …")
        self._worker = PnlWorker(codes)
        self._worker.progress.connect(self._on_prog)
        self._worker.done.connect(self._on_done)
        self._worker.start()

    def _on_prog(self, i, n): self.lbl_status.setText(f"⏳ 抓取历史净值 {i}/{n} …")

    def _on_done(self, out):
        self._hist = out
        self._per = {}; self._pnav = {}
        for code, m in out.items():
            dd = sorted(m); per = {}; pnav = {}; prev = None
            for ds in dd:
                if prev is not None:
                    per[ds] = (m[ds]-prev)/prev*100 if prev else 0.0
                    pnav[ds] = prev
                prev = m[ds]
            self._per[code] = per; self._pnav[code] = pnav
        self._days = sorted(set().union(*[set(p) for p in self._per.values()])) if self._per else []
        if getattr(self, "_div", None):
            self._days = sorted(set(self._days) | set(self._div))
        self._build_timeline()
        if not self._days:
            self.lbl_status.setText("⚠ 没抓到任何历史净值。"); return
        self._ym = (int(self._days[-1][:4]), int(self._days[-1][5:7]))
        self._refresh_month(); self._update_overview()
        if self._start:
            self.lbl_status.setText(f"✅ 已加载 {len(out)}/{len(self._codes)} 只基金的历史净值。收益自 {self._start} 起记录，之前不计。")
        else:
            self.lbl_status.setText(f"✅ 已加载 {len(out)}/{len(self._codes)} 只基金的历史净值。⚠ 无交易记录/买入日期，按当前份额回溯仅供参考，补录交易或填买入日期可修正。")

    def _on_filter_changed(self):
        self._account_filter = self._acc_filter.currentData()
        # 刷新兑底持仓
        if self._account_filter != "__all__":
            self._hold_fb = load_holdings_for_account(self._account_filter)
        else:
            self._hold_fb = load_holdings()
        if self._hist:
            self._build_timeline()
            self._refresh_month(); self._update_overview()

    def _build_timeline(self):
        """每只基金的份额时间线：有交易记录→按记录推；没有→用持仓buy_date+当前份额近似。
           注意：trades.json 里 side 存的是英文 buy/sell/open（见 apply_trade）。"""
        self._tl = {}
        all_trades = load_trades()
        if self._account_filter != "__all__":
            all_trades = [t for t in all_trades if (t.get("account") or DEFAULT_ACCOUNT) == self._account_filter]
        trades = {}
        for t in all_trades:
            trades.setdefault(t["code"], []).append(t)
        for code in self._hist:
            m = self._hist[code]; dds = sorted(m)
            ts = sorted(trades.get(code, []), key=lambda x: x["date"])
            if ts:
                shares = 0.0; pts = []
                for t in ts:
                    ds = t["date"]
                    if t["side"] == "open":
                        shares = float(t.get("shares", 0)); pts.append((ds, shares)); continue
                    if ds not in m:
                        cand = [d for d in dds if d <= ds]
                        ds = cand[-1] if cand else dds[0]
                    amt = t.get("amount")
                    if not amt or m[ds] <= 0:
                        pts.append((t["date"], max(shares, 0.0))); continue
                    dsh = amt/m[ds]
                    if t["side"] in ("buy", "dividend_reinvest"):
                        shares += dsh
                    elif t["side"] in ("sell", "convert"):
                        shares -= dsh
                    else:
                        continue   # dividend_cash 等：份额不变，不产生时间点
                    pts.append((t["date"], max(shares, 0.0)))
                if pts: self._tl[code] = pts
                continue
            rec = self._hold_fb.get(code, {})
            sh = float(rec.get("shares") or 0); bd = (rec.get("buy_date") or "")
            if sh > 0 and bd: self._tl[code] = [(bd, sh)]

    def _shares_on(self, code, ds):
        if getattr(self, "_start", "") and ds < self._start: return 0.0
        tl = self._tl.get(code)
        if not tl: return float(getattr(self, "_hold_fb", {}).get(code, {}).get("shares") or 0)
        sh = 0.0
        for d, s in tl:
            if d <= ds: sh = s
            else: break
        return sh

    def _day_pnl(self, ds):
        tot = 0.0
        for code, per in self._per.items():
            if ds not in per: continue
            sh = self._shares_on(code, ds); pn = self._pnav[code].get(ds)
            if sh and pn: tot += sh*pn*per[ds]/100
        tot += getattr(self, "_div", {}).get(ds, 0.0)   # 现金分红当日计入
        return tot

    def _set_ov(self, key, val, pct=False):
        lb = self._ov_lbls[key]
        if val is None:
            lb.setText("—"); lb.setStyleSheet(f"color:{T()['muted']};font-weight:bold;"); return
        lb.setText(f"{val:+.2f}%" if pct else f"{val:+,.2f}")
        lb.setStyleSheet(f"color:{RED if val>=0 else GREEN};font-weight:bold;")

    def _update_overview(self):
        if not self._days: return
        last = self._days[-1]
        self._set_ov("yest", self._day_pnl(last))
        ym = last[:7]
        mp = sum(self._day_pnl(ds) for ds in self._days if ds[:7] == ym)
        base = 0.0
        first = next((d for d in self._days if d[:7] == ym and d >= (self._start or d)), None)
        if first:
            for code in self._hist:
                sh = self._shares_on(code, first); pv = self._pnav.get(code, {}).get(first)
                if sh and pv: base += sh*pv
        self._set_ov("month", mp); self._set_ov("monthpct", (mp/base*100) if base else None, pct=True)
        self._set_ov("year", sum(self._day_pnl(ds) for ds in self._days if ds[:4] == last[:4]))

    def _shift(self, d):
        y, m = self._ym; m += d
        if m < 1: y, m = y-1, 12
        if m > 12: y, m = y+1, 1
        self._ym = (y, m); self._refresh_month()

    def _refresh_month(self):
        import calendar as _cal
        y, m = self._ym
        self.lbl_month.setText(f"{y}年{m}月")
        day_pnl = {ds: self._day_pnl(ds) for ds in self._days if ds[:7] == f"{y:04d}-{m:02d}"}
        vmax = max([abs(v) for v in day_pnl.values() if abs(v) > 1e-9] + [0.01])
        wd, ndays = _cal.monthrange(y, m)
        first_wd = (wd + 1) % 7      # monthrange 周一=0 → 日历列 周日=0
        today = datetime.now().strftime("%Y-%m-%d")
        t = T()  # v2.0.0：日历格子全主题适配（底色混合而非固定浅色）
        _empty = f"QLabel{{background:{t['panel_bg']};border-radius:8px;color:{t['faint']};}}"
        # v2.2.0：当月合计置顶（红绿信号色，随 set_theme 同步的 RED/GREEN）
        _mtot = sum(day_pnl.values())
        self.lbl_mtot.setText(f"合计 {_mtot:+,.2f}元")
        self.lbl_mtot.setStyleSheet(f"color:{RED if _mtot > 1e-9 else (GREEN if _mtot < -1e-9 else GRAY)};")
        for row in self._cells:
            for cell in row:
                cell.setText(""); cell.setProperty("ds", ""); cell.setStyleSheet(_empty)
        for d in range(1, ndays+1):
            ds = f"{y:04d}-{m:02d}-{d:02d}"; idx = first_wd + d - 1
            cell = self._cells[idx//7][idx%7]; cell.setProperty("ds", ds)
            _tb = f"border:2px solid {t['accent']};" if ds == today else ""  # v2.2.0：今日描边
            if getattr(self, "_start", "") and ds < self._start:
                cell.setText(f"{d}\n—"); cell.setStyleSheet(_empty)
            elif ds in day_pnl and abs(day_pnl[ds]) > 1e-9:
                v = day_pnl[ds]; a = min(abs(v)/vmax, 1.0)
                bg = blend_color(t["card_bg"], t["up"] if v >= 0 else t["down"], 0.18+0.45*a)
                cell.setText(f"{d}\n{v:+.2f}")
                cell.setStyleSheet(f"QLabel{{background:{bg};border-radius:8px;color:{t['text']};font-weight:bold;{_tb}}}QLabel:hover{{border:2px solid {t['accent']};}}")
            elif (idx % 7) in (0, 6) or ds < today:
                cell.setText(f"{d}\n0.00"); cell.setStyleSheet(f"QLabel{{background:{t['hover_bg']};border-radius:8px;color:{t['muted']};{_tb}}}QLabel:hover{{border:2px solid {t['accent']};}}")
            else:
                cell.setText(f"{d}\n—"); cell.setStyleSheet(_empty)
        if self._sel and self._sel[:7] == f"{y:04d}-{m:02d}": self._show_day(self._sel)

    def _click(self, r, c):
        ds = self._cells[r][c].property("ds") or ""
        if not ds: return
        self._sel = ds; self._show_day(ds)

    def _show_day(self, ds):
        rows = []
        for code, per in self._per.items():
            if ds not in per: continue
            sh = self._shares_on(code, ds); pn = self._pnav[code].get(ds)
            if not (sh and pn): continue
            rows.append((NAME_MAP.get(code, code), sh*pn*per[ds]/100, per[ds]))
        div = getattr(self, "_div", {}).get(ds, 0.0)
        if div: rows.append(("现金分红", div, None))
        if not rows:
            self.lbl_day.setText(f"{ds} 休市 / 无持仓"); self.tbl.setRowCount(0); return
        rows.sort(key=lambda x: -x[1]); tot = sum(x[1] for x in rows)
        self.lbl_day.setText(f"{ds} 合计 {tot:+,.2f} 元 （{len(rows)} 条数据）")
        self.tbl.setRowCount(len(rows))
        for r, (nm, pnl, pct) in enumerate(rows):
            self.tbl.setItem(r,0,QTableWidgetItem(nm))
            it = QTableWidgetItem(f"{pnl:+,.2f}"); it.setForeground(QColor(RED if pnl>=0 else GREEN)); self.tbl.setItem(r,1,it)
            it = QTableWidgetItem(f"{pct:+.2f}%" if pct is not None else ""); it.setForeground(QColor(RED if (pct or 0)>=0 else GREEN)); self.tbl.setItem(r,2,it)
            self.tbl.setItem(r,3,QTableWidgetItem(""))

class TradesDialog(QDialog):
    """交易记录：查看流水 + 补录历史。
       补录只写 trades.json，不动当前持仓（历史交易已含在手动持仓里，再过一遍会双计）。
       新交易在「管理持仓」记一笔时自动进流水，不用在这里重复录。
       （OCR 截图导入已移除。）"""
    SIDES = [("buy", "买入"), ("sell", "卖出/赎回"), ("dividend_reinvest", "红利再投"),
             ("dividend_cash", "现金分红"), ("open", "期初快照")]

    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("交易记录")
        self.resize(780, 560)
        root = QVBoxLayout(self)
        _t_lbl = QLabel("补录＝把支付宝里过去的交易抄进来，只影响收益日历，不影响当前持仓。份额必填（照抄支付宝「确认份额」），金额选填。")
        _t_lbl.setStyleSheet(f"color:{T()['text_sub']};"); root.addWidget(_t_lbl)

        self.tbl = QTableWidget(0, 6)
        self.tbl.setHorizontalHeaderLabels(["日期", "基金", "类型", "金额(元)", "份额", "账户"])
        self.tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tbl.setStyleSheet(table_qss())
        root.addWidget(self.tbl, 1)

        form = QFormLayout()
        self.d_date = QDateEdit(QDate.currentDate())
        self.d_date.setCalendarPopup(True); self.d_date.setDisplayFormat("yyyy-MM-dd")
        self.d_date.setStyleSheet(input_qss())
        self.d_code = QComboBox()
        for c in sorted(NAME_MAP, key=lambda c: NAME_MAP[c]):
            self.d_code.addItem(f"{NAME_MAP[c]}（{c}）", c)
        self.d_code.setStyleSheet(combo_qss())
        self.d_side = QComboBox()
        for k, v in self.SIDES:
            self.d_side.addItem(v, k)
        self.d_side.setStyleSheet(combo_qss())
        self.d_account = QComboBox()
        for a in load_accounts():
            self.d_account.addItem(a)
        self.d_account.setStyleSheet(combo_qss())
        self.d_shares = QLineEdit(); self.d_shares.setPlaceholderText("必填：支付宝里的确认份额"); self.d_shares.setStyleSheet(input_qss())
        self.d_amount = QLineEdit(); self.d_amount.setPlaceholderText("选填：金额(元)"); self.d_amount.setStyleSheet(input_qss())
        form.addRow("日期", self.d_date);  form.addRow("基金", self.d_code)
        form.addRow("类型", self.d_side);  form.addRow("账户", self.d_account)
        form.addRow("份额", self.d_shares); form.addRow("金额", self.d_amount)
        root.addLayout(form)

        bar = QHBoxLayout()
        b_add = QPushButton("补录一笔"); b_del = QPushButton("删除选中行")
        b_add.setStyleSheet(soft_btn_qss()); b_del.setStyleSheet(danger_btn_qss())
        b_add.clicked.connect(self._add); b_del.clicked.connect(self._del)
        bar.addWidget(b_add); bar.addWidget(b_del); bar.addStretch()
        root.addLayout(bar)
        self._reload()

    def _reload(self):
        ts = sorted(load_trades(), key=lambda t: t.get("date", ""))
        self.tbl.setRowCount(len(ts))
        for r, t in enumerate(ts):
            side_txt = dict(self.SIDES).get(t.get("side"), t.get("side", ""))
            if t.get("note") == "转换": side_txt = "转换·" + side_txt
            vals = [t.get("date", ""),
                    NAME_MAP.get(t.get("code", "")) or t.get("name") or t.get("code", ""),
                    side_txt,
                    str(t["amount"]) if t.get("amount") is not None else "",
                    str(t["shares"]) if t.get("shares") is not None else "",
                    t.get("account") or DEFAULT_ACCOUNT]
            for c, v in enumerate(vals):
                self.tbl.setItem(r, c, QTableWidgetItem(v))

    def _add(self):
        code = self.d_code.currentData(); side = self.d_side.currentData()
        date = self.d_date.date().toString("yyyy-MM-dd")
        shs = self.d_shares.text().strip(); amt = self.d_amount.text().strip()
        if side != "dividend_cash" and not shs:
            QMessageBox.warning(self, "提示", "请先填份额（现金分红除外）。"); return
        rec = {"date": date, "code": code, "side": side, "account": self.d_account.currentText() or DEFAULT_ACCOUNT}
        if shs: rec["shares"] = float(shs)
        if amt: rec["amount"] = round(float(amt), 2)
        all_t = load_trades(); all_t.append(rec); save_trades(all_t)
        self.d_shares.clear(); self.d_amount.clear()
        self._reload()

    def _del(self):
        r = self.tbl.currentRow()
        if r < 0: return
        ts = sorted(load_trades(), key=lambda t: t.get("date", ""))
        t = ts[r]
        name = NAME_MAP.get(t.get("code", ""), t.get("code", ""))
        if QMessageBox.question(self, "删除", f"删除 {t.get('date')} {name} 的这笔记录？") == QMessageBox.Yes:
            all_t = load_trades(); all_t.remove(t); save_trades(all_t)
            self._reload()


# ---------- 导出 Excel ----------
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def _write_xlsx(path, sections, parent):
    """sections: list of (key, title) ; parent: MainWindow。每个 section 一个 sheet。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除默认 sheet
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
    EMPTY_FONT = Font(italic=True, color="999999")

    def style_sheet(ws, headers, rows):
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = HEADER_ALIGN
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        for i, _ in enumerate(headers, 1):
            col = get_column_letter(i)
            maxlen = max((len(str(ws.cell(r, i).value or "")) for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[col].width = min(maxlen + 2, 40)

    def write_empty(ws, msg):
        ws.cell(1, 1, msg).font = EMPTY_FONT

    for key, title in sections:
        ws = wb.create_sheet(title=title[:31])  # sheet 名 <=31 字符
        if key == "holdings":
            nested = load_holdings_nested()
            if not nested or all(not v for v in nested.values()):
                write_empty(ws, "暂无持仓记录（在「💼 管理持仓」填写后导出）")
                continue
            rows = []
            for acc in sorted(nested):
                holdings = nested[acc]
                if not isinstance(holdings, dict): continue
                for code in sorted(holdings):
                    r = holdings[code]
                    if not isinstance(r, dict): continue
                    sh = r.get("shares", 0); cost = r.get("cost", 0)
                    prin = r.get("principal", 0); bd = r.get("buy_date", "")
                    rows.append([acc, code, sh, cost, prin if prin is not None else "", bd])
            style_sheet(ws, ["账户", "基金代码", "份额", "每份成本", "本金", "买入日期"], rows)
        elif key == "trades":
            t = load_trades()
            if not t:
                write_empty(ws, "暂无交易记录（在「📒 交易记录」补录后导出）")
                continue
            side_map = {"buy": "买入", "sell": "卖出", "open": "期初",
                        "dividend_reinvest": "红利再投", "dividend_cash": "现金分红"}
            rows = []
            for it in sorted(t, key=lambda x: x.get("date", "")):
                rows.append([it.get("date", ""), it.get("code", ""),
                             side_map.get(it.get("side", ""), it.get("side", "")),
                             it.get("amount", ""), it.get("shares", ""),
                             it.get("account") or DEFAULT_ACCOUNT])
            style_sheet(ws, ["日期", "基金代码", "类型", "金额", "份额", "账户"], rows)
        elif key == "pnl":
            pnl = getattr(parent, "_pnl_dialog", None)
            hist = getattr(pnl, "_hist", None) if pnl else None
            days = getattr(pnl, "_days", []) if pnl else []
            if not hist or not days:
                write_empty(ws, "暂无收益明细（请先打开「📅 收益明细」并等待抓取完成）")
                continue
            per = pnl._per; pnav = pnl._pnav; tl = pnl._tl
            name_map = NAME_MAP
            rows = []
            for ds in days:
                for code in sorted(hist):
                    if ds not in per.get(code, {}):
                        continue
                    sh = pnl._shares_on(code, ds)
                    pn = pnav[code].get(ds)
                    if not (sh and pn):
                        continue
                    pnl_val = sh * pn * per[code][ds] / 100
                    rows.append([ds, code, name_map.get(code, code), hist[code].get(ds, ""),
                                 pn, per[code][ds], round(pnl_val, 2)])
            if not rows:
                write_empty(ws, "收益明细抓取完成但无可用数据")
                continue
            style_sheet(ws, ["日期", "基金代码", "基金名称", "当日净值", "前一日净值", "涨跌幅%", "当日盈亏"], rows)
        elif key == "snapshot":
            results = getattr(parent, "last_results", [])
            if not results:
                write_empty(ws, "暂无行情快照（点「🔄 刷新数据」后再导出）")
                continue
            rows = []
            for r in results:
                rows.append([r.get("code", ""), r.get("name", ""), r.get("nav", 0),
                             r.get("est", 0), r.get("chg", 0), r.get("nav_date", ""),
                             r.get("status", ""), r.get("via", "")])
            style_sheet(ws, ["基金代码", "名称", "净值", "估值", "涨跌幅%", "净值日期", "状态", "来源"], rows)
    wb.save(path)


class ExportDialog(QDialog):
    SECTIONS = [("holdings", "持仓"), ("trades", "交易流水"),
                ("pnl", "收益明细"), ("snapshot", "行情快照")]

    def __init__(self, parent=None):
        super().__init__(parent); self.setWindowTitle("导出 Excel"); self.resize(520, 380)
        self._parent = parent
        lay = QVBoxLayout(self); lay.setSpacing(12)
        _e_lbl = QLabel("勾选要导出的数据，选好保存路径，点「导出」生成一个 xlsx 文件（每个数据一个 sheet）。")
        _e_lbl.setStyleSheet(f"color:{T()['text_sub']};"); lay.addWidget(_e_lbl)
        # 复选框
        self._checks = {}
        pnl = getattr(parent, "_pnl_dialog", None)
        has_pnl = bool(pnl) and bool(getattr(pnl, "_hist", None))
        for key, title in self.SECTIONS:
            cb = QCheckBox(title); cb.setFont(QFont(FONT, 10))
            cb.setStyleSheet(f"color:{T()['text']};")
            if key == "pnl" and not has_pnl:
                cb.setEnabled(False)
                cb.setToolTip("请先打开「收益明细」并等待抓取完成，再回来导出")
                cb.setText(f"{title}（需先打开收益明细）")
            elif key == "snapshot" and not getattr(parent, "last_results", None):
                cb.setToolTip("点「刷新数据」后再导出")
            cb.setChecked(key != "pnl" or has_pnl)
            lay.addWidget(cb); self._checks[key] = cb
        # 路径选择
        path_row = QHBoxLayout()
        _sv_lbl = QLabel("保存到："); _sv_lbl.setStyleSheet(f"color:{T()['muted']};"); path_row.addWidget(_sv_lbl)
        self._ed_path = QLineEdit()
        default_name = f"基金日报导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self._ed_path.setText(os.path.join(_BASE, default_name))
        self._ed_path.setStyleSheet(input_qss())
        path_row.addWidget(self._ed_path, 1)
        b_browse = QPushButton("浏览…")
        b_browse.clicked.connect(self._browse)
        b_browse.setStyleSheet(ghost_btn_qss(T(), pad="6px 12px"))
        path_row.addWidget(b_browse)
        lay.addLayout(path_row)
        # 状态
        self._lbl_status = QLabel(""); self._lbl_status.setWordWrap(True)
        self._lbl_status.setStyleSheet(f"color:{T()['muted']};"); lay.addWidget(self._lbl_status)
        # 按钮
        bar = QHBoxLayout(); bar.addStretch()
        b_cancel = QPushButton("取消"); b_cancel.clicked.connect(self.reject)
        b_cancel.setStyleSheet(ghost_btn_qss())
        bar.addWidget(b_cancel)
        b_export = QPushButton("导出"); b_export.clicked.connect(self._do_export)
        b_export.setStyleSheet(primary_btn_qss())
        bar.addWidget(b_export); lay.addLayout(bar)
        self._b_export = b_export

    def _browse(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "选择保存位置", self._ed_path.text(), "Excel 文件 (*.xlsx)")
        if path:
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self._ed_path.setText(path)

    def _do_export(self):
        sections = [(k, t) for k, t in self.SECTIONS if self._checks[k].isChecked()]
        if not sections:
            self._lbl_status.setStyleSheet("color:#e53935;")
            self._lbl_status.setText("⚠ 请至少勾选一项要导出的数据")
            return
        path = self._ed_path.text().strip()
        if not path:
            self._lbl_status.setStyleSheet("color:#e53935;")
            self._lbl_status.setText("⚠ 请选择保存路径")
            return
        self._b_export.setEnabled(False); self._lbl_status.setStyleSheet("color:#888;")
        self._lbl_status.setText("⏳ 导出中…"); QApplication.processEvents()
        try:
            _write_xlsx(path, sections, self._parent)
            size = os.path.getsize(path)
            size_str = f"{size/1024:.1f} KB" if size < 1024 * 1024 else f"{size/1024/1024:.2f} MB"
            self._lbl_status.setStyleSheet("color:#16a34a;")
            self._lbl_status.setText(f"✅ 导出成功：{path}（{size_str}，{len(sections)} 个 sheet）")
            QMessageBox.information(self, "导出成功", f"已导出到： {path}  |  大小：{size_str}  |  包含 {len(sections)} 个 sheet。")
        except Exception as e:
            self._lbl_status.setStyleSheet("color:#e53935;")
            self._lbl_status.setText(f"⚠ 导出失败：{e}")
        finally:
            self._b_export.setEnabled(True)


class SettingsDialog(QDialog):
    """v1.7：设置对话框（外观组：主题切换）。完整分组设置中心见 3.0.0。"""
    def __init__(self, parent=None):
        super().__init__(parent); self.setWindowTitle("设置"); self.resize(470, 210)
        lay = QVBoxLayout(self); lay.setContentsMargins(16,14,16,14); lay.setSpacing(10)
        grp = QLabel("外观"); grp.setFont(QFont(FONT, 11, QFont.Bold)); lay.addWidget(grp)
        row = QHBoxLayout()
        lab = QLabel("应用主题"); lab.setFont(QFont(FONT, 10)); row.addWidget(lab)
        self.combo = QComboBox(); self.combo.setFont(QFont(FONT, 10))
        for k, t in THEMES.items(): self.combo.addItem(t["name"], k)
        ix = self.combo.findData(_THEME)
        if ix >= 0: self.combo.setCurrentIndex(ix)
        row.addWidget(self.combo, 1); lay.addLayout(row)
        hint = QLabel("主题切换立即生效并记住。四套主题均已全屏适配（顶栏/卡片/榜单/详情页/弹窗）。")
        hint.setFont(QFont(FONT, 8)); hint.setStyleSheet(f"color:{T()['muted']};"); hint.setWordWrap(True); lay.addWidget(hint)
        lay.addStretch()
        brow = QHBoxLayout(); brow.addStretch()
        ok = QPushButton("确定"); ok.clicked.connect(self.accept)
        canc = QPushButton("取消"); canc.clicked.connect(self.reject)
        brow.addWidget(ok); brow.addWidget(canc); lay.addLayout(brow)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__(); self.setWindowTitle("基金日报"); self.resize(1320,900)
        s = QApplication.primaryScreen().geometry(); self.move(max(0,(s.width()-1320)//2),max(0,(s.height()-900)//2))
        pg.setConfigOptions(antialias=True)
        set_theme((load_settings() or {}).get("theme", "b_dark"))  # v2.0.0：默认主题=深空暗（FD-HIG 定案），老用户设置优先
        self._apply_global_qss()
        root = QWidget(); self.setCentralWidget(root)
        outer = QVBoxLayout(root); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        self.stack = QStackedWidget(); outer.addWidget(self.stack)
        self.home = self._build_home(); self.detail = DetailPage(on_back=self._go_home)
        self.stack.addWidget(self.home); self.stack.addWidget(self.detail)
        self.worker = None; self.last_results = []; self.resolved = {}; self.corrected_codes = set()
        self._cleared_codes = load_show_state()
        self._pnl_dialog = None
        self._val_cache = {}   # v1.3 估值红绿灯缓存 {"date": 当日, "pct": {code: 百分位}}
        self._spark_cache = load_spark_cache()  # v2.1.0 迷你走势缓存 {"date": 当日, "data": {code: [净值…]}}
        self._spark_workers = []  # v2.1.0：持有工作线程引用防 GC
        self._refresh_home()
        self._check_update()

    def _check_update(self):
        self._upd_worker = UpdateWorker(); self._upd_worker.found.connect(self._on_update_found); self._upd_worker.start()

    def _on_update_found(self, tag, url):
        msg = QMessageBox(self)
        msg.setWindowTitle("发现新版本")
        msg.setIcon(QMessageBox.Information)
        msg.setText(f"🎉 v{tag} 已发布（当前 v{APP_VERSION}）。\n点「去下载」打开下载页。")
        yes = msg.addButton("去下载", QMessageBox.YesRole)
        msg.addButton("跳过", QMessageBox.NoRole)
        msg.exec()
        if msg.clickedButton() is yes and url:
            import webbrowser; webbrowser.open(url)

    def _build_home(self):
        w = QWidget(); w.setObjectName("homeRoot"); w.setAutoFillBackground(True)  # v1.7：主题窗口底
        w.setAttribute(Qt.WA_StyledBackground, True)  # v2.2.0 返工：同详情页，纯 QWidget 的 QSS 背景需显式开启
        _t0 = T()
        if _t0["win_bg"]: w.setStyleSheet(f"#homeRoot{{background:{_t0['win_bg']};}}")
        self._home_root = w
        outer = QVBoxLayout(w); outer.setContentsMargins(18,16,18,16); outer.setSpacing(12)
        top = QHBoxLayout(); top.setSpacing(8)
        t = T()
        # v2.0.0 顶栏（FD-HIG 层级铁律）：标识+标题+徽章 ｜ 账户下拉 ｜ 幽灵导航 ｜ 图标工具 ｜ 唯一主按钮
        self._logo_lbl = QLabel(); self._logo_lbl.setPixmap(icon_pixmap("logo", t["accent"], 26))
        top.addWidget(self._logo_lbl)
        self._title_lbl = QLabel("基金日报"); self._title_lbl.setFont(QFont(FONT,15,QFont.Bold)); self._title_lbl.setStyleSheet(f"color:{t['text']};")
        top.addWidget(self._title_lbl)
        self._ver_lbl = QLabel(f"v{APP_VERSION}"); self._ver_lbl.setFont(QFont(FONT,8))
        self._ver_lbl.setStyleSheet(f"color:{t['muted']};border:1px solid {t['card_border']};border-radius:4px;padding:1px 5px;")
        top.addWidget(self._ver_lbl)
        self.lbl_time = QLabel(""); self.lbl_time.setStyleSheet(f"color:{t['muted']};"); top.addWidget(self.lbl_time)
        top.addSpacing(8)
        self._acc_lbl = QLabel("账户"); self._acc_lbl.setStyleSheet(f"color:{t['muted']};"); top.addWidget(self._acc_lbl)
        self._account_combo = QComboBox(); self._account_combo.setFont(QFont(FONT,9))
        self._account_combo.setStyleSheet(combo_qss(t))  # v2.1.0 返工：构建时即上 Win11 风样式（此前只在 _restyle_all 里，首开时原生 Win98 样）
        self._account_combo.addItem("全部账户", "__all__")
        for _a in load_accounts():
            self._account_combo.addItem(_a, _a)
        self._account_combo.currentIndexChanged.connect(self._on_account_changed)
        top.addWidget(self._account_combo)
        top.addSpacing(6)
        # 导航群：统一幽灵按钮（去彩虹/去 emoji，手绘图标+文字）
        self._tb_navs = []
        for _k, _txt, _slot in (("hold","管理持仓",self._open_hold),("pnl","收益明细",self._open_pnl),
                                ("trades","交易记录",lambda: TradesDialog(self).exec()),("diag","诊断",self._show_diag)):
            _b = QPushButton(QIcon(icon_pixmap(_k, t["muted"], 15)), "  " + _txt)
            _b.setFont(QFont(FONT,9)); _b.clicked.connect(_slot); _b.setStyleSheet(ghost_btn_qss(t))
            self._tb_navs.append((_b, _k, _txt)); top.addWidget(_b)
        self.btn_hold, self.btn_pnl, self.btn_trades, self.btn_diag = [x[0] for x in self._tb_navs]  # 旧属性兼容
        top.addStretch()
        # 工具群：纯图标小按钮归组最右（悬停显名）
        self._tb_tools = []
        for _k, _tip, _slot in (("export","导出",self._open_export),("backup","备份",self._open_backup),
                                ("about","关于",self._open_about),("settings","设置",self._open_settings)):
            _b = QPushButton(QIcon(icon_pixmap(_k, t["muted"], 15)), "")
            _b.setFixedSize(30,30); _b.setToolTip(_tip); _b.clicked.connect(_slot)
            _b.setStyleSheet(ghost_btn_qss(t, pad="0px")); self._tb_tools.append((_b, _k)); top.addWidget(_b)
        self.btn_export, self.btn_backup, self.btn_about, self.btn_settings = [x[0] for x in self._tb_tools]
        # 主操作：全顶栏唯一强调色实底按钮
        self.btn_refresh = QPushButton(QIcon(icon_pixmap("refresh", "#ffffff", 15)), "  刷新数据"); self.btn_refresh.setFont(QFont(FONT,10,QFont.Bold))
        self.btn_refresh.setStyleSheet(primary_btn_qss(t))
        self.btn_refresh.clicked.connect(self._refresh_home); top.addWidget(self.btn_refresh); outer.addLayout(top)
        self.summary = QFrame(); self.summary.setStyleSheet(panel_qss())  # v1.7 令牌化
        sl = QHBoxLayout(self.summary); sl.setContentsMargins(16,12,16,12)
        self.lbl_total = QLabel("总持仓市值  —"); self.lbl_total.setFont(QFont(FONT,11,QFont.Bold))
        self.lbl_today = QLabel("今日盈亏  —"); self.lbl_today.setFont(QFont(FONT,11,QFont.Bold))
        self.lbl_cum = QLabel("累计收益  —"); self.lbl_cum.setFont(QFont(FONT,11,QFont.Bold))  # v1.2：总市值-总本金
        sl.addWidget(self.lbl_total); sl.addStretch(); sl.addWidget(self.lbl_today)
        _sep = QLabel("｜"); _sep.setStyleSheet(f"color:{T()['faint']};"); sl.addWidget(_sep)
        sl.addWidget(self.lbl_cum); outer.addWidget(self.summary)
        t = T()
        self.lbl_fix = QLabel(""); self.lbl_fix.setStyleSheet(f"QLabel{{color:{t['mid_val']};background:{t['hover_bg']};border:1px solid {t['card_border']};border-radius:8px;padding:8px 12px;}}")
        self.lbl_fix.setWordWrap(True); self.lbl_fix.hide(); outer.addWidget(self.lbl_fix)
        self.lbl_alert = QLabel(""); self.lbl_alert.setStyleSheet(f"QLabel{{color:{t['up']};background:{t['hover_bg']};border:1px solid {t['card_border']};border-radius:8px;padding:8px 12px;}}")
        self.lbl_alert.setWordWrap(True); self.lbl_alert.hide(); outer.addWidget(self.lbl_alert)
        self.lbl_noacc = QLabel("")  # v1.2：无持仓账户引导
        self.lbl_noacc.setStyleSheet(f"QLabel{{color:{t['accent']};background:{t['accent_soft']};border:1px solid {t['accent_soft_hover']};border-radius:8px;padding:8px 12px;font-size:12px;}}")
        self.lbl_noacc.setWordWrap(True); self.lbl_noacc.hide(); outer.addWidget(self.lbl_noacc)
        self.lbl_empty = QLabel("看板还是空的：在左上方『快速添加』输入 6 位基金代码，添加第一只基金后，这里就会变成你的持仓看板。")
        self.lbl_empty.setStyleSheet(f"QLabel{{color:{t['accent']};background:{t['accent_soft']};border:1px solid {t['accent_soft_hover']};border-radius:10px;padding:14px 16px;font-size:12px;}}")
        self.lbl_empty.setWordWrap(True); outer.addWidget(self.lbl_empty)
        split = QSplitter(Qt.Horizontal)
        split.setHandleWidth(2)  # v2.0.0：分栏手柄收窄并主题化（修“中间白条”）
        left_scroll = QScrollArea(); left_scroll.setWidgetResizable(True); left_scroll.setMinimumWidth(380)
        left_scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        left_scroll.viewport().setStyleSheet("background:transparent;")  # v2.0.0：同上，视口透明
        left_wrap = QWidget(); left_col = QVBoxLayout(left_wrap); left_col.setContentsMargins(0,0,0,0); left_col.setSpacing(10)
        add_box = QFrame(); add_box.setStyleSheet(board_qss())
        self._add_box = add_box
        abl = QHBoxLayout(add_box); abl.setContentsMargins(12,10,12,10)
        atitle = QLabel("快速添加"); atitle.setFont(QFont(FONT,9,QFont.Bold)); atitle.setStyleSheet(f"color:{t['text']};"); abl.addWidget(atitle)
        self._add_title = atitle
        self._add_input = QLineEdit(); self._add_input.setPlaceholderText("6位代码，或点「搜名字」"); self._add_input.setFont(QFont(FONT,9))
        self._add_input.setFixedWidth(150); self._add_input.setStyleSheet(f"QLineEdit{{padding:6px 8px;border:1px solid {t['card_border']};border-radius:6px;background:{t['card_bg']};color:{t['text']};}}")
        abl.addWidget(self._add_input)
        self._search_btn = QPushButton("搜名字"); self._search_btn.setFont(QFont(FONT,9))
        self._search_btn.setStyleSheet(soft_btn_qss(t))
        self._search_btn.clicked.connect(self._search_pick); abl.addWidget(self._search_btn)
        self._add_btn = QPushButton("添加"); self._add_btn.setFont(QFont(FONT,9))
        self._add_btn.setStyleSheet(soft_btn_qss(t))
        self._add_btn.clicked.connect(self._add_fund); abl.addWidget(self._add_btn)
        self._rm_btn = QPushButton("移除自加"); self._rm_btn.setFont(QFont(FONT,9))
        self._rm_btn.setStyleSheet(f"QPushButton{{padding:6px 12px;border-radius:6px;background:transparent;color:{t['up']};border:1px solid {t['card_border']};}}QPushButton:hover{{background:{t['hover_bg']};}}")
        self._rm_btn.clicked.connect(self._remove_custom_fund); abl.addWidget(self._rm_btn)
        abl.addStretch()
        left_col.addWidget(add_box)
        chart_box = QFrame(); chart_box.setStyleSheet(board_qss())
        cl = QVBoxLayout(chart_box); cl.setContentsMargins(8,8,8,4)
        ctitle = QLabel("今日涨跌一览（涨红跌绿）"); ctitle.setFont(QFont(FONT,10,QFont.Bold)); ctitle.setStyleSheet(f"color:{t['text']};"); cl.addWidget(ctitle)
        self._chart_title = ctitle  # v2.1.0 返工：存引用供 _restyle_all 重刷（防切主题后停在旧色）
        self.chart = BarChart(); self.chart.setFixedHeight(180); self.chart.apply_theme(); cl.addWidget(self.chart); left_col.addWidget(chart_box)
        self._chart_box = chart_box
        self.board = QFrame(); self.board.setStyleSheet(board_qss())
        bl = QHBoxLayout(self.board); bl.setContentsMargins(14,10,14,10)
        self._red_col = QVBoxLayout(); rh = QLabel("今日红榜"); rh.setFont(QFont(FONT,10,QFont.Bold)); rh.setStyleSheet(f"color:{T()['up']};"); self._red_col.addWidget(rh)
        self._red_rows = [QLabel("—") for _ in range(3)]
        for lb in self._red_rows: lb.setFont(QFont(FONT,9)); lb.setStyleSheet(f"color:{T()['up']};"); self._red_col.addWidget(lb)
        self._green_col = QVBoxLayout(); gh = QLabel("今日黑榜"); gh.setFont(QFont(FONT,10,QFont.Bold)); gh.setStyleSheet(f"color:{T()['down']};"); self._green_col.addWidget(gh)
        self._green_rows = [QLabel("—") for _ in range(3)]
        for lb in self._green_rows: lb.setFont(QFont(FONT,9)); lb.setStyleSheet(f"color:{T()['down']};"); self._green_col.addWidget(lb)
        bl.addLayout(self._red_col); bl.addSpacing(20); bl.addLayout(self._green_col); left_col.addWidget(self.board)
        # 快照全貌区已随 OCR 移除（lbl_snap_total / snap_box 不再创建）
        left_col.addStretch()
        left_scroll.setWidget(left_wrap); split.addWidget(left_scroll)
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setMinimumWidth(440)
        scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        scroll.viewport().setStyleSheet("background:transparent;")  # v2.0.0：视口透明，主题底色透出来（修“空白区不变色”）
        self.cards_wrap = QWidget(); self.cards_layout = QVBoxLayout(self.cards_wrap); self.cards_layout.setSpacing(8); self.cards_layout.addStretch()
        scroll.setWidget(self.cards_wrap)
        right_wrap = QWidget(); rw = QVBoxLayout(right_wrap); rw.setContentsMargins(0,0,0,0); rw.setSpacing(6)
        sort_row = QHBoxLayout()  # v1.5：卡片排序
        _sort_lbl = QLabel("卡片排序"); _sort_lbl.setFont(QFont(FONT,9)); _sort_lbl.setStyleSheet(f"color:{t['muted']};")
        self._sort_lbl = _sort_lbl
        self._sort_combo = QComboBox(); self._sort_combo.setFont(QFont(FONT,9))
        for _t2, _v in (("默认顺序","default"),("估值·低估优先","val_asc"),("估值·过热优先","val_desc"),("今日涨幅优先","chg_desc"),("持有金额优先","mv_desc")):
            self._sort_combo.addItem(_t2, _v)
        self._sort_combo.setStyleSheet(combo_qss(t))  # v2.1.0 返工：原 1.x 内联样式无子控件箭头，改 Win11 风 combo_qss
        _sm = (load_settings() or {}).get("sort_mode", "default")
        _six = self._sort_combo.findData(_sm)
        if _six >= 0: self._sort_combo.setCurrentIndex(_six)
        self._sort_combo.currentIndexChanged.connect(self._on_sort_change)
        sort_row.addWidget(_sort_lbl); sort_row.addWidget(self._sort_combo); sort_row.addStretch()
        rw.addLayout(sort_row); rw.addWidget(scroll, 1)
        split.addWidget(right_wrap)
        split.setChildrenCollapsible(False); split.setStretchFactor(0, 4); split.setStretchFactor(1, 6)
        split.setSizes([520, 760])
        self._split = split
        self._theme_split()
        outer.addWidget(split, 1)
        self.cards = {}
        self._cleared_codes = getattr(self, "_cleared_codes", load_show_state())
        for code,name in FUNDS:
            c = FundCard(code); c.lbl_name.setText(name)
            c.btn_detail.clicked.connect(lambda checked, cd=code: self._open_detail(cd))
            c.btn_clear.clicked.connect(lambda checked, cd=code: self._toggle_cleared(cd))
            if code in self._cleared_codes: c.set_cleared(True)
            _sp = (getattr(self, "_spark_cache", {}) or {}).get("data", {})
            if _sp.get(code): c.spark.set_points(_sp[code])  # v2.1.0：构建时命中缓存先画上
            self.cards_layout.insertWidget(self.cards_layout.count()-1,c); self.cards[code]=c
        self._sync_empty_banner()
        return w

    def _theme_split(self):
        """v2.0.0：分栏手柄随主题着色（构建与切换共用）。"""
        if hasattr(self, "_split"):
            self._split.setStyleSheet(f"QSplitter::handle{{background:{T()['card_border']};}}")

    def _open_settings(self):
        """v1.7：设置（主题切换）。"""
        d = SettingsDialog(self)
        if d.exec() == QDialog.Accepted:
            choice = d.combo.currentData()
            if choice and choice != _THEME:
                set_theme(choice)
                s = load_settings() or {}; s["theme"] = choice; save_settings(s)
                self._restyle_all()

    def _apply_global_qss(self):
        """v2.0.0：弹窗/表格类控件全局主题样式（四主题全应用：classic 不再豁免，保全程序样式统一）。"""
        self.setStyleSheet(global_qss(T()))

    def _restyle_all(self):
        """v1.7/v2.0.0：主题切换后全量重刷令牌化组件。"""
        t = T()
        self._apply_global_qss()
        root = getattr(self, "_home_root", None)
        if root is not None:
            root.setStyleSheet(f"#homeRoot{{background:{t['win_bg']};}}" if t["win_bg"] else "")
        # 顶栏（v2.0.0：图标/文字/徽章随主题重绘）
        if hasattr(self, "_logo_lbl"):
            self._logo_lbl.setPixmap(icon_pixmap("logo", t["accent"], 26))
            self._title_lbl.setStyleSheet(f"color:{t['text']};")
            self._ver_lbl.setStyleSheet(f"color:{t['muted']};border:1px solid {t['card_border']};border-radius:4px;padding:1px 5px;")
            self.lbl_time.setStyleSheet(f"color:{t['muted']};"); self._acc_lbl.setStyleSheet(f"color:{t['muted']};")
            self._account_combo.setStyleSheet(combo_qss(t))  # v2.1.0 返工：Win11 风（含箭头子控件），原 1.x 内联样式无子控件致首开 Win98 样
            for b, kind, _txt in self._tb_navs:
                b.setIcon(QIcon(icon_pixmap(kind, t["muted"], 15))); b.setStyleSheet(ghost_btn_qss(t))
            for b, kind in self._tb_tools:
                b.setIcon(QIcon(icon_pixmap(kind, t["muted"], 15))); b.setStyleSheet(ghost_btn_qss(t, pad="0px"))
            self.btn_refresh.setIcon(QIcon(icon_pixmap("refresh", "#ffffff", 15))); self.btn_refresh.setStyleSheet(primary_btn_qss(t))
        # 提示条/快速添加/排序（v2.0.0）
        if hasattr(self, "lbl_fix"):
            self.lbl_fix.setStyleSheet(f"QLabel{{color:{t['mid_val']};background:{t['hover_bg']};border:1px solid {t['card_border']};border-radius:8px;padding:8px 12px;}}")
            self.lbl_alert.setStyleSheet(f"QLabel{{color:{t['up']};background:{t['hover_bg']};border:1px solid {t['card_border']};border-radius:8px;padding:8px 12px;}}")
            _info = f"QLabel{{color:{t['accent']};background:{t['accent_soft']};border:1px solid {t['accent_soft_hover']};border-radius:8px;padding:8px 12px;font-size:12px;}}"
            self.lbl_noacc.setStyleSheet(_info); self.lbl_empty.setStyleSheet(_info.replace("8px;", "10px;").replace("padding:8px 12px", "padding:14px 16px"))
            self._add_input.setStyleSheet(f"QLineEdit{{padding:6px 8px;border:1px solid {t['card_border']};border-radius:6px;background:{t['card_bg']};color:{t['text']};}}")
            self._search_btn.setStyleSheet(soft_btn_qss(t)); self._add_btn.setStyleSheet(soft_btn_qss(t))
            self._rm_btn.setStyleSheet(f"QPushButton{{padding:6px 12px;border-radius:6px;background:transparent;color:{t['up']};border:1px solid {t['card_border']};}}QPushButton:hover{{background:{t['hover_bg']};}}")
            if hasattr(self, "_sort_combo"):
                self._sort_combo.setStyleSheet(combo_qss(t))
        if hasattr(self, "summary"): self.summary.setStyleSheet(panel_qss())
        if hasattr(self, "_split"): self._theme_split()
        if hasattr(self, "board"): self.board.setStyleSheet(board_qss())
        if hasattr(self, "_chart_box"): self._chart_box.setStyleSheet(board_qss())
        if hasattr(self, "_chart_title"): self._chart_title.setStyleSheet(f"color:{t['text']};")  # v2.1.0 返工：图表标题随主题重刷
        if hasattr(self, "chart"): self.chart.apply_theme()
        if hasattr(self, "_add_box"):  # v2.0.0 返工②：快速添加区随主题重刷
            self._add_box.setStyleSheet(board_qss())
            self._add_title.setStyleSheet(f"color:{t['text']};")
        if hasattr(self, "_sort_lbl"): self._sort_lbl.setStyleSheet(f"color:{t['muted']};")
        if hasattr(self, "_red_col"):
            for lb in [self._red_col.itemAt(i).widget() for i in range(self._red_col.count())]:
                if lb: lb.setStyleSheet(f"color:{t['up']};")
            for lb in [self._green_col.itemAt(i).widget() for i in range(self._green_col.count())]:
                if lb: lb.setStyleSheet(f"color:{t['down']};")
        for c in self.cards.values():
            c.setStyleSheet(card_qss_cleared() if getattr(c, "_cleared", False) else card_qss_normal())
            c.lbl_name.setStyleSheet(f"color:{t['text']};")  # v2.1.0 返工：基金名随主题重刷
            c.lbl_code.setStyleSheet(f"color:{t['muted']};")
            c.lbl_nav.setStyleSheet(f"color:{t['muted']};")
            c.lbl_mv.setStyleSheet(f"color:{t['text']};")
            c._sep1.setStyleSheet(f"color:{t['faint']};"); c._sep2.setStyleSheet(f"color:{t['faint']};")
            c.spark.update()  # v2.1.0：走势线颜色实时读令牌，重绘即随主题
            c._theme_btns()  # v2.0.0 返工②：卡片按钮随主题重刷
        # v2.0.0 返工③：估值信号标签按缓存重刷（颜色/档位随主题）
        _vc = self._val_cache.get("pct") or {}
        for code, c in self.cards.items():
            c.set_val(_vc.get(code))
        if hasattr(self, "detail"): self.detail._apply_theme()
        if self.last_results:
            self._apply_results()

    def _toggle_cleared(self, code):
        card = self.cards.get(code)
        if not card: return
        on = code not in self._cleared_codes
        if on: self._cleared_codes.add(code)
        else: self._cleared_codes.discard(code)
        save_show_state(self._cleared_codes)
        card.set_cleared(on)
        nm = dict(FUNDS).get(code, code)
        if on:
            QMessageBox.information(self, "已清仓",
                f"「{nm}」已盖灰章：不再计入柱状图/红黑榜/总持仓，卡片仅观察。\n持仓记录未动，点卡片上『恢复持有』可随时还原。")
        else:
            QMessageBox.information(self, "已恢复", f"「{nm}」已恢复持有，重新计入总账与榜单。")
        self._apply_results()
        for fn in (self._update_board, self._update_alert, self._update_summary):
            try:
                fn(self.last_results)
            except Exception:
                pass

    def _fade_cards(self):
        for c in self.cards.values():
            eff = QGraphicsOpacityEffect(c); c.setGraphicsEffect(eff); eff.setOpacity(0.0)
            a = QPropertyAnimation(eff, b"opacity", c); a.setDuration(260); a.setStartValue(0.0); a.setEndValue(1.0)
            c._fade_anim = a; a.start(QAbstractAnimation.DeleteWhenStopped)

    def _on_sort_change(self):
        """v1.5：排序选择持久化到 settings.json（schema 只增不改）并重排。"""
        if not hasattr(self, "_sort_combo"): return
        s = load_settings() or {}; s["sort_mode"] = self._sort_combo.currentData(); save_settings(s)
        self._apply_sort()

    def _apply_sort(self):
        """v1.5：按当前排序模式重排卡片（removeWidget 不删控件，仅重新插入）。"""
        if not hasattr(self, "_sort_combo") or not self.cards: return
        mode = self._sort_combo.currentData()
        if mode == "default":
            order = [c for c, _ in FUNDS if c in self.cards]
        else:
            val_map = {c: (info or {}).get("pct") for c, info in (self._val_cache.get("pct") or {}).items()}
            chg_map = {d["code"]: d.get("chg",0) for d in self.last_results if d.get("status")=="ok"}
            mv_map = {}
            for d in self.last_results:
                r2 = self.resolved.get(d["code"]); nav = d.get("nav",0)
                if r2 and r2.get("shares") and nav: mv_map[d["code"]] = float(r2["shares"]) * nav
            order = sort_card_codes(self.cards.keys(), mode, val_map, chg_map, mv_map)
        for c in self.cards.values(): self.cards_layout.removeWidget(c)
        for code in order:
            if code in self.cards: self.cards_layout.insertWidget(self.cards_layout.count()-1, self.cards[code])

    def _sync_empty_banner(self):
        self.lbl_empty.setVisible(not FUNDS)

    def _refresh_home(self):
        if not FUNDS:
            self._sync_empty_banner()
            return
        self.btn_refresh.setEnabled(False); self.btn_refresh.setText("抓取中…")
        self.lbl_time.setText(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.worker = Worker(); self.worker.done.connect(self._on_home_done); self.worker.start()

    def _on_home_done(self, results):
        self.last_results = results
        self._cleared_codes = load_show_state()
        for _cd, _c in self.cards.items(): _c.set_cleared(_cd in self._cleared_codes)
        self._apply_results(); self._fade_cards()
        self._start_val_worker(results)  # v1.3：估值红绿灯后台拉取
        self._start_spark_worker(results)  # v2.1.0：卡片迷你走势后台拉取
        self.btn_refresh.setEnabled(True); self.btn_refresh.setText("刷新数据")
        if sum(1 for r in results if r.get("status")=="ok") == 0:
            QMessageBox.warning(self,"没抓到数据","全部抓取失败，请点「🩺 诊断」。")

    def _start_val_worker(self, results):
        """v1.3：估值信号——当日已算过的用缓存，其余后台拉取（v1.4 指数走 PE/PB）。"""
        clr = getattr(self, "_cleared_codes", set()) or set()
        nav_map = {d["code"]: d.get("nav",0) for d in results
                   if d.get("status")=="ok" and d.get("nav") and d["code"] not in clr}
        if not nav_map: return
        name_map = {d["code"]: d.get("name","") for d in results if d["code"] in nav_map}
        today = datetime.now().strftime("%Y-%m-%d")
        if self._val_cache.get("date") != today:
            self._val_cache = {"date": today, "pct": {}}
        cache = self._val_cache["pct"]
        for c, info in cache.items():
            card = self.cards.get(c)
            if card and c in nav_map: card.set_val(info)
        todo = {c: n for c, n in nav_map.items() if c not in cache}
        if not todo: return
        for c in todo:
            card = self.cards.get(c)
            if card: card.set_val_loading()
        self._val_worker = ValWorker(todo, {c: name_map.get(c,"") for c in todo})
        self._val_worker.done.connect(self._on_val_done)
        self._val_worker.start()

    def _on_val_done(self, info_map):
        self._val_cache.setdefault("pct", {}).update(info_map)
        for c, info in info_map.items():
            card = self.cards.get(c)
            if card: card.set_val(info)
        self._apply_sort()  # v1.5：估值到达后按分位排序才有数据

    def _start_spark_worker(self, results):
        """v2.1.0：卡片迷你走势——当日缓存命中直接画，其余后台拉近 60 日净值。"""
        codes = [d["code"] for d in results if d.get("status") == "ok"]
        if not codes: return
        today = datetime.now().strftime("%Y-%m-%d")
        if self._spark_cache.get("date") != today:
            self._spark_cache = {"date": today, "data": {}}
        data = self._spark_cache["data"]
        for c in codes:
            card = self.cards.get(c)
            if card and data.get(c): card.spark.set_points(data[c])
        todo = [c for c in codes if not data.get(c)]
        if not todo: return
        w = SparkWorker(todo); w.done.connect(self._on_spark_done)
        self._spark_workers.append(w); w.start()

    def _on_spark_done(self, pts_map):
        self._spark_cache.setdefault("data", {}).update(pts_map)
        save_spark_cache(self._spark_cache)  # 落盘按日缓存，当日再刷不重复请求
        for c, pts in pts_map.items():
            card = self.cards.get(c)
            if card: card.spark.set_points(pts)

    def _apply_results(self):
        price_map = {d["code"]: d.get("nav",0) for d in self.last_results if d.get("status")=="ok"}
        account = self._account_combo.currentData() if hasattr(self, '_account_combo') else "__all__"
        if account == "__all__":
            holdings = load_holdings()
        else:
            holdings = load_holdings_for_account(account)
        self.resolved, self.corrected_codes = resolve_holdings(holdings, price_map)
        # v1.2：无持仓账户引导横幅（只在选中具体账户时判断）
        if hasattr(self, "lbl_noacc"):
            if account != "__all__":
                _has_hold = any((float((rec or {}).get("shares") or 0) > 0 or float((rec or {}).get("principal") or 0) > 0)
                                for rec in holdings.values() if isinstance(rec, dict))
                self.lbl_noacc.setText(f"👛 「{account}」账户还没有持仓：点「💼 管理持仓」切到该账户填写份额/成本，盈亏统计才会生效。")
                self.lbl_noacc.setVisible(not _has_hold)
            else:
                self.lbl_noacc.hide()
        self._cleared_codes = load_show_state()
        act = [d for d in self.last_results if d["code"] not in self._cleared_codes]
        for d in act:
            c = self.cards.get(d["code"])
            if c: c.update_data(d, self.resolved)
        self.chart.draw([(d.get("name",d["code"])[:8], d.get("chg",0)) for d in act], animate=True)
        self._update_summary(self.last_results)
        self._update_board(self.last_results)
        self._update_alert(self.last_results)
        if self.corrected_codes:
            self.lbl_fix.setText(f"🔧 检测到 {len(self.corrected_codes)} 只填的是『本金』(含份额≈1 的撞车行)，已按『每份成本』显示盈亏（卡片带🔧）。点「💼 管理持仓」核对并保存，即可永久修正。")
            self.lbl_fix.show()
        else:
            self.lbl_fix.hide()
        self._apply_sort()  # v1.5：新数据到达后维持排序

    def _update_board(self, results):
        clr = getattr(self, "_cleared_codes", set()) or set()
        results = [d for d in results if d["code"] not in clr]
        ok = [(d.get("name",d["code"]), d.get("chg",0)) for d in results if d.get("status")=="ok"]
        ok.sort(key=lambda x: x[1], reverse=True)
        up = [x for x in ok if x[1] > 0]
        dn = [x for x in reversed(ok) if x[1] < 0]
        for i,lb in enumerate(self._red_rows):
            lb.setText(f"{up[i][0][:10]}  {up[i][1]:+.2f}%  🚀" if i < len(up) else "—")
        bot = list(reversed(ok))
        for i,lb in enumerate(self._green_rows):
            lb.setText(f"{dn[i][0][:10]}  {dn[i][1]:+.2f}%  📉" if i < len(dn) else "—")

    def _update_alert(self, results):
        alerts = []
        clr = getattr(self, "_cleared_codes", set()) or set()
        for d in results:
            if d["code"] in clr: continue
            if d.get("status") != "ok": continue
            name = d.get("name",d["code"]); nav = d.get("nav",0); chg = d.get("chg",0)
            if chg <= -3:
                alerts.append(f"⚠ {name[:8]} 单日跌 {chg:.1f}%")
            r2 = self.resolved.get(d["code"])
            if r2 and r2.get("shares") and nav:
                cost = float(r2.get("cost") or 0)
                if cost > 0:
                    pct = (nav-cost)/cost*100
                    if pct <= -15:
                        alerts.append(f"🕳 {name[:8]} 已深套 {pct:.0f}%")
                    tp = take_profit_level(pct)  # v1.3：止盈参考线提醒
                    if tp:
                        alerts.append(f"🎯 {name[:8]} 累计 +{pct:.0f}%，触及 {tp}% 止盈参考线")
        if alerts:
            self.lbl_alert.setText("  ｜  ".join(alerts[:4])); self.lbl_alert.show()
        else:
            self.lbl_alert.hide()

    def _update_summary(self, results):
        total_mv=0.0; today_pnl=0.0; total_prin=0.0; has=False
        clr = getattr(self, "_cleared_codes", set()) or set()
        results = [d for d in results if d["code"] not in clr]
        for d in results:
            r2=self.resolved.get(d["code"]); nav=d.get("nav",0); chg=d.get("chg",0)
            if r2 and r2.get("shares") and nav:
                cost=float(r2.get("cost") or 0); sh=float(r2["shares"])
                if cost > 0:
                    mv=nav*sh; total_mv+=mv; today_pnl += mv*(chg/(100+chg)) if (100+chg) else 0; has=True
                    prin=float(r2.get("principal") or 0)  # v1.2：本金缺失时用 成本×份额 兑底
                    total_prin += prin if prin > 0 else cost*sh
        if has:
            self.lbl_total.setText(f"总持仓市值  ¥{total_mv:,.2f}"); self.lbl_total.setStyleSheet(f"color:{T()['text']};")
            pc=RED if today_pnl>=0 else GREEN
            _nds = [r.get("nav_date","") for r in results if r.get("status")=="ok" and r.get("nav_date")]
            _nd = max(_nds) if _nds else ""
            _tag = "今日盈亏" if (_nd == datetime.now().strftime("%Y-%m-%d")) else (f"盈亏(截至{_nd[5:]})" if _nd else "今日盈亏")
            self.lbl_today.setText(f"{_tag}  {today_pnl:+,.2f}元"); self.lbl_today.setStyleSheet(f"color:{pc};")
            cum = total_mv - total_prin  # v1.2：累计收益 = 总市值 - 总本金
            cpc = RED if cum >= 0 else GREEN
            pct = (cum/total_prin*100) if total_prin > 0 else 0.0
            self.lbl_cum.setText(f"累计收益  {cum:+,.2f}元 ({pct:+.2f}%)"); self.lbl_cum.setStyleSheet(f"color:{cpc};")
        else:
            # 无手填持仓 → 显示平均涨跌（OCR 快照回退已移除）
            self.lbl_total.setText("总持仓市值  未填持仓"); self.lbl_total.setStyleSheet(f"color:{T()['muted']};font-size:12px;")
            self.lbl_cum.setText("累计收益  —"); self.lbl_cum.setStyleSheet(f"color:{T()['muted']};")
            chgs=[d.get("chg",0) for d in results if d.get("status")=="ok"]
            if chgs:
                avg=sum(chgs)/len(chgs); pc=RED if avg>=0 else GREEN
                self.lbl_today.setText(f"今日平均涨跌  {avg:+.2f}%（{len(chgs)}只）"); self.lbl_today.setStyleSheet(f"color:{pc};")
            else:
                self.lbl_today.setText("今日盈亏  —"); self.lbl_today.setStyleSheet(f"color:{T()['muted']};")

    def _open_export(self):
        ExportDialog(self).exec()

    def _open_backup(self):
        """一键备份：把 5 个数据 json 打包成 zip。支持记住默认目录。"""
        files = [EXTRA_FILE, HOLD_FILE, TRADES_FILE, SHOW_FILE, ACCOUNTS_FILE, SETTINGS_FILE]
        existing = [f for f in files if os.path.exists(f)]
        if not existing:
            QMessageBox.information(self, "备份", "还没有任何数据文件，无需备份。")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"基金日报备份_{ts}.zip"
        settings = load_settings()
        backup_dir = settings.get("backup_dir", "")
        used_default = False
        if backup_dir and os.path.isdir(backup_dir):
            save_path = os.path.join(backup_dir, default_name)
            used_default = True
        else:
            save_path, _ = QFileDialog.getSaveFileName(self, "选择备份保存位置", default_name, "ZIP 压缩包 (*.zip)")
            if not save_path:
                return
        try:
            with zipfile.ZipFile(save_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for f in existing:
                    zf.write(f, os.path.basename(f))
        except Exception as e:
            QMessageBox.warning(self, "备份失败", f"写入失败：{e}")
            return
        if used_default:
            msg = QMessageBox(self)
            msg.setWindowTitle("备份成功")
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"✅ 已备份到：\n{save_path}\n\n（已存到你设置的默认目录）")
            b_change = msg.addButton("更换默认目录", QMessageBox.ActionRole)
            msg.addButton("好的", QMessageBox.AcceptRole)
            msg.exec()
            if msg.clickedButton() is b_change:
                new_dir = QFileDialog.getExistingDirectory(self, "选择新的默认备份目录", backup_dir)
                if new_dir:
                    settings["backup_dir"] = new_dir; save_settings(settings)
        else:
            msg = QMessageBox(self)
            msg.setWindowTitle("备份成功")
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"✅ 已备份到：\n{save_path}")
            b_remember = msg.addButton("记住这个位置", QMessageBox.ActionRole)
            msg.addButton("就这一次", QMessageBox.AcceptRole)
            msg.exec()
            if msg.clickedButton() is b_remember:
                settings["backup_dir"] = os.path.dirname(save_path); save_settings(settings)

    def _open_about(self):
        """v1.0.0 新增：关于本软件（版本信息 / GitHub 链接 / 隐私声明 / 免责声明 / 手动检查更新）。"""
        dlg = QDialog(self); dlg.setWindowTitle("关于本软件"); dlg.resize(460, 420)
        lay = QVBoxLayout(dlg); lay.setSpacing(8)
        t = QLabel("📊 基金日报"); t.setFont(QFont(FONT,18,QFont.Bold)); t.setAlignment(Qt.AlignCenter); lay.addWidget(t)
        v = QLabel(f"v{APP_VERSION} · 正式发布"); v.setAlignment(Qt.AlignCenter); v.setStyleSheet("color:#2563eb;font-weight:bold;"); lay.addWidget(v)
        desc = QLabel("Windows 个人基金看板：实时行情 / 持仓盈亏 / 收益日历 / 多账户 / 一键备份 / 导出 Excel。\n全部持仓与交易数据仅存本地磁盘，不上传。")
        desc.setAlignment(Qt.AlignCenter); desc.setWordWrap(True); desc.setStyleSheet("color:#555;"); lay.addWidget(desc)
        link = QLabel(f'<a href="https://github.com/{GITHUB_REPO}">github.com/{GITHUB_REPO}</a>')
        link.setTextFormat(Qt.RichText); link.setTextInteractionFlags(Qt.TextBrowserInteraction)
        link.setOpenExternalLinks(True); link.setAlignment(Qt.AlignCenter); lay.addWidget(link)
        b_check = QPushButton("🔄 立即检查更新")
        b_check.setStyleSheet("QPushButton{padding:8px 14px;border-radius:8px;background:#eef3ff;color:#2563eb;border:none;}QPushButton:hover{background:#dbe6ff;}QPushButton:disabled{background:#eee;color:#999;}")
        lay.addWidget(b_check)
        def do_check():
            b_check.setEnabled(False); b_check.setText("检查中…")
            def on_checked(has_new, tag, url):
                b_check.setEnabled(True); b_check.setText("🔄 立即检查更新")
                if has_new:
                    QMessageBox.information(dlg, "发现新版本", f"🎉 v{tag} 已发布（当前 v{APP_VERSION}）。\n请到 GitHub Releases 下载更新。")
                elif tag:
                    QMessageBox.information(dlg, "检查完成", f"✅ 已是最新版本（v{APP_VERSION}）。")
                else:
                    QMessageBox.information(dlg, "检查失败", "当前连不上 GitHub，请检查网络后重试。")
            w = UpdateWorker(timeout=8); w.checked.connect(on_checked)
            dlg._check_worker = w  # 防 worker 被提前回收
            w.start()
        b_check.clicked.connect(do_check)
        disc = QLabel("免责声明：本工具仅供个人记账参考，不构成任何投资建议；\n行情数据或有延迟，以官方渠道为准。")
        disc.setAlignment(Qt.AlignCenter); disc.setWordWrap(True); disc.setStyleSheet("color:#999;font-size:10px;"); lay.addWidget(disc)
        b_close = QPushButton("关闭"); b_close.clicked.connect(dlg.accept)
        b_close.setStyleSheet("QPushButton{padding:8px 24px;border-radius:8px;background:#f0f0f0;border:none;}QPushButton:hover{background:#e3e3e3;}")
        lay.addWidget(b_close, 0, Qt.AlignCenter)
        dlg.exec()

    def _open_pnl(self):
        self._pnl_dialog = PnlDialog(self); self._pnl_dialog.show(); self._pnl_dialog.start()

    def _refresh_account_combo(self):
        """重新填充主页账户下拉框（管理账户增删改后调用，免重启，v1.1 新增）。"""
        if not hasattr(self, "_account_combo"): return
        cur = self._account_combo.currentData()
        self._account_combo.blockSignals(True)
        self._account_combo.clear()
        self._account_combo.addItem("全部账户", "__all__")
        for a in load_accounts():
            self._account_combo.addItem(a, a)
        idx = self._account_combo.findData(cur)
        self._account_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self._account_combo.blockSignals(False)

    def _open_hold(self):
        price_map = {d["code"]: d.get("nav",0) for d in self.last_results if d.get("status")=="ok"}
        account = self._account_combo.currentData() if hasattr(self, '_account_combo') else "__all__"
        if account == "__all__":
            account = load_accounts()[0]
        holdings = load_holdings_for_account(account)
        resolved, corrected = resolve_holdings(holdings, price_map)
        dlg = HoldDialog(resolved, corrected, price_map, self, account)
        def _after(_):
            self._refresh_account_combo()  # 无论是否保存，账户结构可能已变
            if dlg.saved: self._refresh_home()
        dlg.finished.connect(_after)
        dlg.exec()

    def _on_account_changed(self):
        if self.last_results:
            self._apply_results()

    def _open_detail(self, code):
        # v2.2.0：把当日涨跌幅传给详情页副行（今日盈亏）
        _chg = next((d.get("chg") for d in self.last_results if d.get("code") == code and d.get("status") == "ok"), None)
        self.detail.load(code, self.resolved.get(code, {}), (self._val_cache.get("pct") or {}).get(code), _chg); self.stack.setCurrentIndex(1)
    def _go_home(self): self.stack.setCurrentIndex(0)

    def _redraw_aggregates(self):
        self.chart.draw([(d.get("name",d["code"])[:8], d.get("chg",0)) for d in self.last_results], animate=False)
        self._update_board(self.last_results); self._update_summary(self.last_results); self._update_alert(self.last_results)

    def _search_pick(self):
        d = QDialog(self); d.setWindowTitle("🔍 搜基金名字"); d.resize(460, 420)
        lay = QVBoxLayout(d)
        row = QHBoxLayout()
        ed = QLineEdit(); ed.setPlaceholderText("输入基金名字/代码，如 广发纳指 / 006479"); ed.setFont(QFont(FONT,9))
        ed.setText(self._add_input.text().strip())
        row.addWidget(ed, 1)
        b_go = QPushButton("搜索"); b_go.setStyleSheet("QPushButton{padding:6px 14px;border-radius:7px;background:#2563eb;color:#fff;border:none;}")
        row.addWidget(b_go); lay.addLayout(row)
        lst = QListWidget(); lst.setFont(QFont(FONT,9)); lay.addWidget(lst, 1)
        lbl = QLabel("双击某一行即可添加到看板。"); lbl.setStyleSheet("color:#888;font-size:10px;"); lay.addWidget(lbl)
        def do_search():
            lst.clear(); lst.addItem("⏳ 搜索中…")
            b_go.setEnabled(False)
            def on_done(res):
                b_go.setEnabled(True)
                lst.clear()
                if not res:
                    lst.addItem("（没搜到，或网络不通。可改输 6 位代码直接添加。）"); return
                for c, n in res:
                    lst.addItem(f"{n}（{c}）")
            w = SearchWorker(ed.text()); w.done.connect(on_done)
            d._worker = w  # 防 worker 被提前回收
            w.start()
        b_go.clicked.connect(do_search)
        def pick(item, *_):
            txt = item.text()
            m = re.search(r"（(\d{6})）", txt)
            if not m: return
            self._add_input.setText(m.group(1))
            d.accept()
        lst.itemDoubleClicked.connect(pick)
        d.exec()
        if d.result() == QDialog.Accepted:
            self._add_fund()

    def _add_fund(self):
        code = self._add_input.text().strip()
        if not re.search(r"^\d{6}$", code):
            QMessageBox.warning(self,"代码格式","请输入6位基金代码，例如 000001。"); return
        if code in self.cards or code in {c for c,_ in FUNDS}:
            QMessageBox.information(self,"已存在",f"{code} 已在列表里，无需重复添加。"); return
        self._add_btn.setEnabled(False); self._add_btn.setText("⏳ 识别中…"); QApplication.processEvents()
        r = fetch_one(code)
        self._add_btn.setEnabled(True); self._add_btn.setText("添加")
        name = (r.get("name") or "").strip()
        if not name or r.get("status") != "ok":
            QMessageBox.warning(self,"识别失败",f"抓不到 {code} 的名字/净值——代码可能无效，或网络暂时不通。未添加。"); return
        extra = _load_extra()
        if code not in {it["code"] for it in extra}:
            extra.append({"code": code, "name": name}); _save_extra(extra)
        FUNDS.append((code, name)); NAME_MAP[code] = name
        card = FundCard(code); card.lbl_name.setText(name)
        card.btn_detail.clicked.connect(lambda checked, cd=code: self._open_detail(cd))
        card.btn_clear.clicked.connect(lambda checked, cd=code: self._toggle_cleared(cd))
        self.cards_layout.insertWidget(self.cards_layout.count()-1, card)
        self.cards[code] = card
        card.update_data(r, self.resolved)
        self.last_results.append(r)
        w = SparkWorker([code]); w.done.connect(self._on_spark_done)  # v2.1.0：新加基金也拉走势
        self._spark_workers.append(w); w.start()
        self._redraw_aggregates(); self._add_input.clear()
        self._sync_empty_banner()
        QMessageBox.information(self,"已添加",
            f"✅ {name}（{code}）已加入看板并保存。\n\n该基金尚未填写持仓，暂不计入盈亏；\n在「💼 管理持仓」填写份额/成本后，盈亏统计才会包含它。")

    def _remove_custom_fund(self):
        extra = _load_extra()
        if not extra:
            QMessageBox.information(self,"没有可移除的基金","当前看板为空，没有可移除的基金。\n"); return
        d = QDialog(self); d.setWindowTitle("🗑 移除自加基金"); d.resize(460,380)
        lay = QVBoxLayout(d)
        lay.addWidget(QLabel("下面列出看板上的所有基金，点「移除」即从看板删除。\n"))
        list_box = QVBoxLayout(); lay.addLayout(list_box); lay.addStretch()
        b_close = QPushButton("关闭"); b_close.clicked.connect(d.accept)
        b_close.setStyleSheet("QPushButton{padding:8px 16px;border-radius:8px;background:#2563eb;color:#fff;border:none;}")
        lay.addWidget(b_close, 0, Qt.AlignRight)
        def rebuild():
            while list_box.count():
                it = list_box.takeAt(0); w = it.widget()
                if w: w.deleteLater()
            cur = _load_extra()
            if not cur:
                t = QLabel("（已全部移除）"); t.setStyleSheet("color:#999;"); list_box.addWidget(t); return
            holds = load_holdings()
            for _it in cur:
                c, n = _it["code"], _it["name"]
                roww = QWidget(); rl = QHBoxLayout(roww); rl.setContentsMargins(0,2,0,2)
                lab = QLabel(f"{n}  ({c})"); lab.setFont(QFont(FONT,9)); rl.addWidget(lab,1)
                if c in holds and ((holds[c].get("shares") or 0) or (holds[c].get("principal") or 0)):
                    warn = QLabel("⚠填过持仓"); warn.setStyleSheet("color:#b45309;font-size:8px;"); rl.addWidget(warn)
                bb = QPushButton("移除"); bb.setStyleSheet("QPushButton{padding:4px 12px;border-radius:6px;background:#fdecea;color:#b3261e;border:1px solid #f5b7b1;}")
                bb.clicked.connect(lambda checked, cc=c, nn=n: do_remove(cc, nn)); rl.addWidget(bb)
                list_box.addWidget(roww)
        def do_remove(c, n):
            holds = load_holdings()
            has_hold = c in holds and ((holds[c].get("shares") or 0) or (holds[c].get("principal") or 0))
            clear_hold = False
            if has_hold:
                ans = QMessageBox.question(d,"该只填过持仓",
                    f"「{n}」在管理表里填过份额/本金。\n移除时是否一并清空它的持仓记录？\n\n是＝连持仓一起删；否＝只从看板移除、持仓保留。",
                    QMessageBox.Yes|QMessageBox.No|QMessageBox.Cancel, QMessageBox.Cancel)
                if ans == QMessageBox.Cancel: return
                clear_hold = (ans == QMessageBox.Yes)
            _save_extra([it for it in _load_extra() if str(it.get("code","")).strip() != c])
            for i in range(len(FUNDS)-1,-1,-1):
                if FUNDS[i][0] == c: del FUNDS[i]
            NAME_MAP.pop(c, None)
            card = self.cards.pop(c, None)
            if card: self.cards_layout.removeWidget(card); card.deleteLater()
            self.last_results = [x for x in self.last_results if x.get("code") != c]
            if clear_hold:
                _remove_code_from_all_accounts(c)
            self._redraw_aggregates(); rebuild()
            self._sync_empty_banner()
        rebuild(); d.exec()

    def _show_diag(self):
        if not FUNDS:
            QMessageBox.information(self, "诊断", "看板是空的，没有可诊断的基金。\n请先在左上方『➕ 快速添加』添加基金。")
            return
        if not self.last_results:
            QMessageBox.information(self,"诊断","还没抓过数据，先点「刷新数据」。"); return
        d = QDialog(self); d.setWindowTitle("🩺 抓取诊断"); d.resize(620,420); lay = QVBoxLayout(d)
        lay.addWidget(QLabel("检测以下数据能否顺利抓取："))
        te = QTextEdit(); te.setReadOnly(True); te.setFont(QFont("Consolas",9))
        lines = [f"时间 {datetime.now():%Y-%m-%d %H:%M:%S}", "",
                 f"成功 {sum(1 for r in self.last_results if r.get('status')=='ok')} / 共 {len(self.last_results)}", "-"*50]
        for r in self.last_results:
            if r.get("status")=="ok":
                lines.append(f"[OK]   {r['code']} {r.get('name','')[:14]:14s} 净值{r['nav']:.4f} 涨跌{r['chg']:+.2f}%  via={r.get('via')}")
            else:
                lines.append(f"[FAIL] {r['code']} {r.get('name','')[:14]:14s}  -> {r.get('err','未知')}")
        # v1.3：持仓集中度（按当前市值，仅信息展示）
        pm = {r["code"]: r.get("nav",0) for r in self.last_results if r.get("status")=="ok"}
        cs = concentration_stats(self.resolved, pm)
        if cs:
            lines += ["-"*50, "持仓结构（按当前市值）",
                      f"持有 {cs['n']} 只基金，总市值 ¥{cs['total']:,.2f}",
                      f"最大单只：{cs['top1_name'][:12]}，占比 {cs['top1_pct']:.1f}%",
                      f"前三大占比：{cs['top3_pct']:.1f}%" + ("  ⚠ 偏集中（参考阈值 60%）" if cs['top3_pct'] >= 60 else "")]
        else:
            lines += ["-"*50, "持仓结构：暂无有效持仓（未填份额或无行情）"]
        te.setPlainText("\n".join(lines)); lay.addWidget(te)
        b = QPushButton("关闭"); b.clicked.connect(d.accept); lay.addWidget(b); d.exec()


if __name__ == "__main__":
    app = QApplication(sys.argv); win = MainWindow(); win.show(); sys.exit(app.exec())
